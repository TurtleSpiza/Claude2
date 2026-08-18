"""b43_build.py  -  v42 -> v43 (18-Aug-2026)
Combined enrichment build (Builds 1+2 per user direction):
  (1) Themes sheet restated as four years, stale controls repointed to live register reads,
      data-quality block added.
  (2) Site dimension: column DZ made EXHAUSTIVE (no row reads "(no site named)"), the source
      hierarchy extended to every register column that carries a printed park name
      (S01 Ev Line items ... S10 Src User Fld 2), depot / network / accounting / internal /
      capture-backlog classes labelled, multi-site holds extended.
Column T never written; control total unchanged; no green block touched.
One save -> one convert-route recalc (isolated LO profile, OOXMLRecalcMode=0) -> one exact-TRUE verify -> ship.
"""
import json, copy, collections, re, os, sys, subprocess, time
import openpyxl, pandas as pd
from openpyxl.utils import get_column_letter
from python_calamine import CalamineWorkbook

ROOT='/home/claude/b43'
SRC='/home/claude/up2/PS_WP_Transaction_Register_3FY_v42_HANDOVER.xlsx'
NEW='PS_WP_Transaction_Register_3FY_v43_HANDOVER.xlsx'
OUT_DIR=f'{ROOT}/out'; RECALC_DIR=f'{ROOT}/out/recalc'
os.makedirs(OUT_DIR,exist_ok=True); os.makedirs(RECALC_DIR,exist_ok=True)
CT=11377312.51
SITE_COL=129; BASIS_COL=130; SC='DY'; BC='DZ'
MULTI='Multiple sites named (not allocated)'
NONE_OLD='(no site named)'
FYS=['FY2023/24','FY2024/25','FY2025/26','FY2026/27']
EXTRACT={'FY2023/24':3481091.25,'FY2024/25':3627231.98,'FY2025/26':3429658.75}

# ---------------------------------------------------------------- 0. gazetteer + register read (calamine)
cwb=CalamineWorkbook.from_path(SRC)
sites_old=cwb.get_sheet_by_name('Sites').to_python()
gaz=[str(x[0]).strip() for x in sites_old[4:487] if str(x[0]).strip() not in ('','None')]
gaz=[g for g in gaz if g!='Undertake General Park']            # blocklisted facility-type label carried in at v42
CANON={g.lower():g for g in gaz}
GP=re.compile('('+'|'.join(re.escape(g) for g in sorted(gaz,key=len,reverse=True) if len(g)>5)+')',re.I)
DEPOT=re.compile(r'Parks Depot|177 Chambers Flat',re.I)

df=pd.read_excel(SRC,sheet_name='Register',header=3,engine='calamine').iloc[:21469]
assert len(df)==21469
amt=pd.to_numeric(df.iloc[:,19],errors='coerce').fillna(0)
assert abs(amt.sum()-CT)<0.005, amt.sum()
basis_old=df.iloc[:,BASIS_COL-1].fillna('').astype(str)
site_old=df.iloc[:,SITE_COL-1].fillna('').astype(str)

# ---------------------------------------------------------------- 1. source hierarchy over the unattributed rows (whole row, banned cols excluded)
SRC_COLS=[(111,'Sighted invoice line items (printed)'),
          (106,'Sighted invoice site (printed)'),
          (112,'Sighted invoice works carried out (printed)'),
          (108,'Sighted invoice work description'),
          (127,'Sighted invoice anomalies note'),
          ( 99,'Sighted invoice ship-to (printed)'),
          ( 42,'Electricity supply point'),
          ( 22,'GL narration'),
          ( 23,'Enquiry narration'),
          ( 80,'Src User Fld 2')]
BANNED_COLS={12,25,32,92,98,105}          # own analysis / derived / supplier premises / Council address
assert not ({c for c,_ in SRC_COLS} & BANNED_COLS)
DEPOT_LBL='Parks Depot, Marsden (non-park facility)'
L_DEPOT='Non-park Council facility (ship-to = Parks Depot)'
L_NET='City-wide network service (no site by design)'
L_ACCT='Accounting event (no site by construction)'
L_INT='Internal charge (not site-specific)'
L_HARV='Invoice sighted, printed site not in gazetteer (candidate harvest)'
L_ADDR='Supply-point address only (address route pending)'
L_CAPT='No invoice sighted (awaiting capture)'
L_NONE='No site determinable'
CLASS_LABELS=[MULTI,L_DEPOT,L_NET,L_ACCT,L_INT,L_HARV,L_ADDR,L_CAPT,L_NONE]

txt={c:df.iloc[:,c-1].fillna('').astype(str) for c,_ in SRC_COLS}
docu=df.iloc[:,8].fillna('').astype(str).str.upper()
narr=df.iloc[:,21].fillna('').astype(str)
natc=df.iloc[:,23].fillna('').astype(str)
stat=df.iloc[:,32].fillna('').astype(str)
supl=df.iloc[:,41].fillna('').astype(str).str.strip()
evid=df.iloc[:,87].fillna('').astype(str).str.strip()
sighted=~evid.isin(['','nan'])
jrnl=docu.str.contains('JOURNAL|JNL')|narr.str.contains(r'Accrue|Accrual|EOM|EOFY|Reversal|Recode|Incorrectly posted',case=False)
waste=natc.str.contains('Waste',case=False)|narr.str.contains('Cleanaway',case=False)
intern=natc.str.contains('Internal',case=False)
hasaddr=~supl.isin(['','nan'])

LEAK='Undertake General Park'   # facility-type label that leaked into Panel A at v42 (75 rows, all unsighted Partial)
new_site=list(site_old); new_basis=list(basis_old)
cnt=collections.Counter(); srcuse=collections.Counter(); verify_rows=[]; n_leak=0
for i in range(len(df)):
    if basis_old.iloc[i]!=NONE_OLD and site_old.iloc[i]!=LEAK: continue
    if site_old.iloc[i]==LEAK: n_leak+=1
    site=None; basis=None
    for c,lbl in SRC_COLS:
        t=txt[c].iloc[i]
        if not t.strip(): continue
        if c==99 and DEPOT.search(t):
            site=DEPOT_LBL; basis=L_DEPOT; srcuse[c]+=1; break
        found=sorted({CANON.get(m.lower(),m) for m in GP.findall(t)})
        if found:
            srcuse[c]+=1
            if len(found)>1: site='; '.join(found); basis=MULTI
            else: site=found[0]; basis=lbl
            verify_rows.append(dict(row=i+5,LineKey=str(df.iloc[i,0]),amount=float(amt.iloc[i]),site=site,basis=basis,source_col=c,source_text=t[:200]))
            break
    if site is None:
        if waste.iloc[i]: basis=L_NET
        elif jrnl.iloc[i]: basis=L_ACCT
        elif intern.iloc[i]: basis=L_INT
        elif sighted.iloc[i]: basis=L_HARV
        elif hasaddr.iloc[i]: basis=L_ADDR
        elif stat.iloc[i] in ('Partial','Pending evidence'): basis=L_CAPT
        else: basis=L_NONE
    new_site[i]=site or ''; new_basis[i]=basis; cnt[basis]+=1

# canonical-label uniqueness (v42 COUNTIF case-fold trap)
allparks=[s for s,b in zip(new_site,new_basis) if s and b not in CLASS_LABELS]
fold=collections.defaultdict(set)
for s in allparks: fold[s.lower()].add(s)
dup={k:v for k,v in fold.items() if len(v)>1}
assert not dup, f'case-folded label collision: {dup}'
assert not any(b==NONE_OLD for b in new_basis)
print('leaked-label rows reverted:',n_leak)
print('DZ classes on the former (no site named) rows:')
for k,v in cnt.most_common(): print(f'   {v:>6}  {k}')
print('source column use:',dict(srcuse))
json.dump(verify_rows,open(f'{ROOT}/b43_site_writes_for_verification.json','w'),indent=1)

# ---------------------------------------------------------------- 2. openpyxl writes
wb=openpyxl.load_workbook(SRC)
reg=wb['Register']; assert reg.max_row==21475
tmpl_d=reg.cell(5,42)
n_changed=0
for i in range(len(df)):
    if basis_old.iloc[i]!=NONE_OLD and site_old.iloc[i]!=LEAK: continue
    r=i+5
    a=reg.cell(r,SITE_COL); a.value=(new_site[i] or None); a._style=copy.copy(tmpl_d._style)   # openpyxl: ws.cell(r,c,None) does NOT clear
    b=reg.cell(r,BASIS_COL); b.value=new_basis[i];      b._style=copy.copy(tmpl_d._style)
    n_changed+=1
print('register rows rewritten (DY/DZ only):',n_changed)

# ---- park aggregates for Sites Panel A (from the new columns)
fyv=df.iloc[:,4].astype(str); pkv=df.iloc[:,16].astype(str)
park_fy=collections.defaultdict(lambda:collections.Counter()); park_pk=collections.defaultdict(collections.Counter); park_bas=collections.defaultdict(collections.Counter)
for s,b,f,p,a in zip(new_site,new_basis,fyv,pkv,amt):
    if s and b not in CLASS_LABELS:
        park_fy[s][f]+=a; park_pk[s][p]+=1; park_bas[s][b]+=1
n_single=sum(1 for s,b in zip(new_site,new_basis) if s and b not in CLASS_LABELS)
n_multi=sum(1 for b in new_basis if b==MULTI)
n_none=sum(1 for b in new_basis if b==L_NONE)
class_counts={c:sum(1 for b in new_basis if b==c) for c in CLASS_LABELS}
print(f'single {n_single} parks {len(park_fy)} | classes {class_counts}')

# ---------------------------------------------------------------- 3. Sites rebuild (same shape as v42, more group rows)
old=wb['Sites']
pb0=int(str(cwb.get_sheet_by_name('Config').to_python()[34][1]).split(':')[0])   # SITES_PANELB start
elec=[[old.cell(r,c).value for c in range(1,5)] for r in range(pb0,old.max_row+1)]
elec_styles=[[copy.copy(old.cell(r,c)._style) for c in range(1,5)] for r in range(pb0,old.max_row+1)]
hdr_style=copy.copy(old.cell(4,1)._style); ttl_style=copy.copy(old.cell(1,1)._style)
sub_style=copy.copy(old.cell(2,1)._style); data_style=copy.copy(old.cell(5,1)._style)
num_style=copy.copy(old.cell(5,3)._style); mon_style=copy.copy(old.cell(5,4)._style)
wb.remove(old)
ws=wb.create_sheet('Sites', wb.sheetnames.index('PK_Listing'))
ws.sheet_view.showGridLines=False
for i,w in enumerate([44,11,17,16,16,16,16,13,44,10],1): ws.column_dimensions[get_column_letter(i)].width=w
def put(sh,r,c,v,st=None,fmt=None):
    cell=sh.cell(r,c,v)
    if st is not None: cell._style=copy.copy(st)
    if fmt: cell.number_format=fmt
    return cell
put(ws,1,1,'Site dimension: cost per park / site, all loaded years',ttl_style)
put(ws,2,1,'Panel A keyed on Register column '+SC+' (Park / Site normalised); basis in column '+BC+'. v43: column '+BC+' is EXHAUSTIVE - every line carries either a park attribution basis or one of the closed-list classes below (multi-site hold, non-park facility, city-wide network service, accounting event, internal charge, candidate harvest, supply-point address only, awaiting capture, no site determinable). Nothing is inferred: a park is written only where a source field prints a FULL PARK NAME matching the gazetteer. Sources are read in evidence order over the whole register row: sighted invoice line items (col 111) > site details (106) > works carried out (112) > work description (108) > anomalies note (127) > ship-to (99, depot only) > electricity supply point (42) > GL narration (22) > enquiry narration (23) > Src User Fld 2 (80). Columns 12, 25, 32, 92, 98 and 105 are never read (own analysis, derived, supplier premises, Council address). Multi-site lines are HELD, never split.',sub_style)
put(ws,3,1,'Panel A: cost per park / site (live)',hdr_style)
HDRS=['Park / Site','Lines (live)','$ ex GST (live)']+FYS+['Primary PK','Primary basis','Bases']
for j,h in enumerate(HDRS,1): put(ws,4,j,h,hdr_style)
parks=sorted(park_fy.keys(), key=lambda p:-sum(park_fy[p].values()))
PA_START=5; r=PA_START
for p in parks:
    put(ws,r,1,p,data_style)
    put(ws,r,2,f'=COUNTIF(Register!${SC}$5:${SC}$21473,$A{r})',num_style,'#,##0;-#,##0;"-"')
    put(ws,r,3,f'=SUMIF(Register!${SC}$5:${SC}$21473,$A{r},Register!$T$5:$T$21473)',mon_style,'$#,##0.00;($#,##0.00);"-"')
    for k,fy in enumerate(FYS):
        put(ws,r,4+k,f'=SUMIFS(Register!$T$5:$T$21473,Register!${SC}$5:${SC}$21473,$A{r},Register!$E$5:$E$21473,"{fy}")',mon_style,'$#,##0.00;($#,##0.00);"-"')
    put(ws,r,8,max(park_pk[p].items(),key=lambda x:x[1])[0],data_style)
    put(ws,r,9,max(park_bas[p].items(),key=lambda x:x[1])[0],data_style)
    put(ws,r,10,len(park_bas[p]),num_style,'#,##0')
    r+=1
PA_END=r-1
GRP_NOTES={MULTI:'Held: several parks named on one line; never allocated',
  L_DEPOT:'Ship-to prints Logan City Council Parks Depot, 177 Chambers Flat Rd Marsden',
  L_NET:'Cleanaway public place bins; whole-of-city, no site exists',
  L_ACCT:'Accrual / EOM / EOFY / reversal / recode; nets negative here',
  L_INT:'Plant and fleet hire, stores issue, rates and levies',
  L_HARV:'Green-blocked; printed name not yet in the 483-name gazetteer',
  L_ADDR:'Street address only; property-key or G-NAF route pending',
  L_CAPT:'Partial / Pending; the site arrives with the invoice',
  L_NONE:'Residual'}
GRP_START=r
for lbl in CLASS_LABELS:
    put(ws,r,1,lbl,data_style)
    put(ws,r,2,f'=COUNTIF(Register!${BC}$5:${BC}$21473,$A{r})',num_style,'#,##0;-#,##0;"-"')
    put(ws,r,3,f'=SUMIF(Register!${BC}$5:${BC}$21473,$A{r},Register!$T$5:$T$21473)',mon_style,'$#,##0.00;($#,##0.00);"-"')
    for k,fy in enumerate(FYS):
        put(ws,r,4+k,f'=SUMIFS(Register!$T$5:$T$21473,Register!${BC}$5:${BC}$21473,$A{r},Register!$E$5:$E$21473,"{fy}")',mon_style,'$#,##0.00;($#,##0.00);"-"')
    put(ws,r,9,GRP_NOTES[lbl],data_style)
    r+=1
GRP_END=r-1; TOT_ROW=r
put(ws,r,1,'TOTAL (must tie to the register)',hdr_style)
put(ws,r,2,f'=SUM(B{PA_START}:B{GRP_END})',hdr_style,'#,##0')
put(ws,r,3,f'=SUM(C{PA_START}:C{GRP_END})',hdr_style,'$#,##0.00;($#,##0.00);"-"')
for k in range(4,8): put(ws,r,k,f'=SUM({get_column_letter(k)}{PA_START}:{get_column_letter(k)}{GRP_END})',hdr_style,'$#,##0.00;($#,##0.00);"-"')
for k in (8,9,10): put(ws,r,k,None,hdr_style)
CHK_ROW=r+1
put(ws,CHK_ROW,1,'Tie check (lines = 21,469 and $ = Register!$T$21475)',hdr_style)
put(ws,CHK_ROW,2,f'=IF(B{TOT_ROW}=21469,"TRUE","FALSE")',hdr_style)
put(ws,CHK_ROW,3,f'=IF(ROUND(C{TOT_ROW},2)=ROUND(Register!$T$21475,2),"TRUE","FALSE")',hdr_style)
put(ws,CHK_ROW,4,f'=IF(ROUND(SUM(D{TOT_ROW}:G{TOT_ROW}),2)=ROUND(C{TOT_ROW},2),"TRUE","FALSE")',hdr_style)
PB_TITLE=CHK_ROW+2
put(ws,PB_TITLE,1,'Panel B: electricity by retailer account, all loaded years (relocated unchanged)',hdr_style)
for j,h in enumerate(['Retailer account','Site','Lines','$ ex GST (net, all years)'],1): put(ws,PB_TITLE+1,j,h,hdr_style)
r=PB_TITLE+2
for vals,sts in zip(elec,elec_styles):
    for j,v in enumerate(vals,1): put(ws,r,j,v,sts[j-1])
    r+=1
PB_END=r-1
print(f'Sites: PanelA {PA_START}:{PA_END} ({len(parks)}), groups {GRP_START}:{GRP_END}, total {TOT_ROW}, check {CHK_ROW}, PanelB {PB_TITLE+2}:{PB_END}')

# ---------------------------------------------------------------- 4. Themes restated (four years, live controls)
th=wb['Themes']
for mr in list(th.merged_cells.ranges): th.unmerge_cells(str(mr))
t_ttl=copy.copy(th.cell(1,1)._style); t_sub=copy.copy(th.cell(2,1)._style); t_hdr=copy.copy(th.cell(4,1)._style)
t_lab=copy.copy(th.cell(5,1)._style); t_mon=copy.copy(th.cell(5,2)._style); t_tot=copy.copy(th.cell(18,1)._style)
t_ctl=copy.copy(th.cell(21,1)._style); t_note=copy.copy(th.cell(27,1)._style)
old_note=th.cell(27,1).value
themes=[th.cell(r,1).value for r in range(5,18)]
for row in th.iter_rows(min_row=1,max_row=th.max_row):
    for c in row: c.value=None
put(th,1,1,'Thematic analysis, four loaded years (restated v43)',t_ttl)
put(th,2,1,'Left panel: everything loaded, by FY (Doc) column E, in year order. Right panel: PK000022. Scopes differ by year (see Method 1.0): FY2026/27 is a part-year extract and grows with each load; its control ties live to the register rather than to a frozen subtotal. T9 accounting events are excluded from thematic spend; read T10 waste net per year.',t_sub)
TH_H=['Theme']+[f'{f} (all loaded)' for f in FYS]+['All years','']+[f'PK000022 {f}' for f in FYS]+['PK000022 all']
for j,h in enumerate(TH_H,1): put(th,4,j,h,t_hdr)
r=5
for t in themes:
    put(th,r,1,t,t_lab)
    for k,fy in enumerate(FYS):
        put(th,r,2+k,f'=SUMIFS(Register!$T$5:$T$21473,Register!$Z$5:$Z$21473,$A{r},Register!$E$5:$E$21473,"{fy}")',t_mon,'$#,##0.00;($#,##0.00);"-"')
    put(th,r,6,f'=SUM(B{r}:E{r})',t_mon,'$#,##0.00;($#,##0.00);"-"')
    for k,fy in enumerate(FYS):
        put(th,r,8+k,f'=SUMIFS(Register!$T$5:$T$21473,Register!$Z$5:$Z$21473,$A{r},Register!$E$5:$E$21473,"{fy}",Register!$P$5:$P$21473,"PK000022")',t_mon,'$#,##0.00;($#,##0.00);"-"')
    put(th,r,12,f'=SUM(H{r}:K{r})',t_mon,'$#,##0.00;($#,##0.00);"-"')
    r+=1
TH_TOT=r
put(th,r,1,'Total (nets to each ledger extract)',t_tot)
for c in list(range(2,7))+list(range(8,13)):
    L=get_column_letter(c); put(th,r,c,f'=SUM({L}5:{L}{r-1})',t_tot,'$#,##0.00;($#,##0.00);"-"')
r+=2; put(th,r,1,'Control checks (all must read TRUE)',t_hdr); r+=1
TH_C0=r
put(th,r,1,'FY2023/24 = 24SLACT extract $3,481,091.25',t_ctl); put(th,r,2,f'=IF(ROUND(B{TH_TOT},2)=3481091.25,"TRUE","FALSE")',t_ctl); r+=1
put(th,r,1,'FY2024/25 = 25SLACT extract $3,627,231.98',t_ctl); put(th,r,2,f'=IF(ROUND(C{TH_TOT},2)=3627231.98,"TRUE","FALSE")',t_ctl); r+=1
put(th,r,1,'FY2025/26 = 26SLACT extract $3,429,658.75',t_ctl); put(th,r,2,f'=IF(ROUND(D{TH_TOT},2)=3429658.75,"TRUE","FALSE")',t_ctl); r+=1
put(th,r,1,'FY2026/27 = register minus the three prior-year extracts (live; part-year, grows with each load)',t_ctl); put(th,r,2,f'=IF(ROUND(E{TH_TOT},2)=ROUND(Register!$T$21475-3481091.25-3627231.98-3429658.75,2),"TRUE","FALSE")',t_ctl); r+=1
put(th,r,1,'All years = Register!$T$21475 (control total, live)',t_ctl); put(th,r,2,f'=IF(ROUND(F{TH_TOT},2)=ROUND(Register!$T$21475,2),"TRUE","FALSE")',t_ctl); r+=1
put(th,r,1,'PK000022 all years = SUMIFS over the register (live)',t_ctl); put(th,r,2,f'=IF(ROUND(L{TH_TOT},2)=ROUND(SUMIFS(Register!$T$5:$T$21473,Register!$P$5:$P$21473,"PK000022"),2),"TRUE","FALSE")',t_ctl); r+=1
TH_C1=r-1
r+=1; put(th,r,1,'Data quality: FY (Service) values outside the four loaded years (col G), for information',t_hdr); r+=1
put(th,r,1,'FY (Service)',t_hdr); put(th,r,2,'Lines',t_hdr); put(th,r,3,'$ ex GST',t_hdr); r+=1
TH_DQ0=r
for f in ['FY1999/00','FY2021/22','FY2022/23','FY2027/28','Unstated']:
    put(th,r,1,f,t_lab)
    put(th,r,2,f'=COUNTIF(Register!$G$5:$G$21473,$A{r})',t_mon,'#,##0')
    put(th,r,3,f'=SUMIF(Register!$G$5:$G$21473,$A{r},Register!$T$5:$T$21473)',t_mon,'$#,##0.00;($#,##0.00);"-"')
    r+=1
r+=1; put(th,r,1,old_note,t_note)
th.column_dimensions['A'].width=52
for c in range(2,13): th.column_dimensions[get_column_letter(c)].width=17
print(f'Themes: rows 5:{TH_TOT-1}, total {TH_TOT}, controls {TH_C0}:{TH_C1}, dq {TH_DQ0}:{r-2}')

# ---------------------------------------------------------------- 5. Config
cf=wb['Config']
crow={str(cf.cell(i,1).value):i for i in range(1,cf.max_row+1) if cf.cell(i,1).value}
cf.cell(crow['WORKBOOK_VERSION'],2,'v43')
cf.cell(crow['SITES_PANELA'],2,f'{PA_START}:{PA_END}')
cf.cell(crow['SITES_GROUP_ROWS'],2,f'{GRP_START}:{GRP_END}')
cf.cell(crow['SITES_TOTAL_ROW'],2,TOT_ROW); cf.cell(crow['SITES_CHECK_ROW'],2,CHK_ROW)
cf.cell(crow['SITES_PANELB'],2,f'{PB_TITLE+2}:{PB_END}')
cf.cell(crow['PARK_COUNT'],2,len(parks)); cf.cell(crow['SITE_ATTRIBUTED_LINES'],2,n_single); cf.cell(crow['SITE_MULTI_LINES'],2,n_multi)
cst_a=copy.copy(cf.cell(cf.max_row,1)._style); cst_b=copy.copy(cf.cell(cf.max_row,2)._style)
nr=cf.max_row+1
for k,v in [('SITE_CLASS_LABELS',len(CLASS_LABELS)),('SITE_NONE_LINES',n_none),('SITE_SOURCE_COLS','111,106,112,108,127,99,42,22,23,80'),
            ('SITE_BANNED_COLS','12,25,32,92,98,105'),('THEMES_TOTAL_ROW',TH_TOT),('THEMES_CONTROL_ROWS',f'{TH_C0}:{TH_C1}')]:
    a=cf.cell(nr,1,k); a._style=copy.copy(cst_a); b=cf.cell(nr,2,v); b._style=copy.copy(cst_b); nr+=1

# ---------------------------------------------------------------- 6. Handover / Method / DA / Open_Items
h=wb['Handover']; hr=h.max_row+1; hs=[copy.copy(h.cell(h.max_row,c)._style) for c in range(1,4)]
h.cell(hr,1,f'v43 change (18-Aug-2026): (1) Themes sheet restated as four years with live controls (the v14-era grand-total control $10,677,585.33 and the FY2026/27 P1 control were reading FALSE); (2) site dimension made exhaustive - column {BC} carries a closed-list class on every line, no line reads "(no site named)"; source hierarchy extended to every register column that prints a park name (line items col 111 and works narrative col 112 were never read at v42 and carried $1.26M of names). Enrichment build: column T untouched, no green block written, control total unchanged at ${CT:,.2f}.')._style=hs[0]
h.cell(hr,2,f'{n_single} lines attributed to {len(parks)} named parks (v42: 5,452 / 483); {n_multi} multi-site holds (v42: 319); {class_counts[L_DEPOT]} depot; {class_counts[L_NET]} network service; {class_counts[L_ACCT]} accounting event; {class_counts[L_INT]} internal; {class_counts[L_HARV]} candidate harvest; {class_counts[L_ADDR]} address only; {class_counts[L_CAPT]} awaiting capture; {n_none} no site determinable. The {len(verify_rows)} new park writes are listed in b43_site_writes_for_verification.json for the user check (Open Items #129).')._style=hs[1]
h.cell(hr,3,None)._style=hs[2]
m=wb['Method']; mr=m.max_row+1; ms=copy.copy(m.cell(m.max_row,1)._style)
METH=[
 f"42.0 Themes restated (v43). The sheet had accumulated three FALSE-reading controls: a grand-total test frozen at $10,677,585.33 (v14), an FY2026/27 P1 test frozen at $139,603.35, and a header still reading 'three loaded years' with FY2025/26 bolted on as a ninth column. Restated as four FY columns in year order plus an all-years column, PK000022 mirrored, and six controls that read the register live: three prior-year extracts (frozen, correct), FY2026/27 = register minus those three (live, part-year), all years = Register!$T$21475, PK000022 = SUMIFS. A data-quality block lists FY (Service) values outside the loaded years. Standing rule: a control that tests a frozen figure for a part-year scope will go FALSE on the next load; part-year controls tie live.",
 f"42.1 Site basis is exhaustive (v43). Column {BC} now carries one of: a park attribution source, '{MULTI}', or a closed-list class - '{L_DEPOT}', '{L_NET}', '{L_ACCT}', '{L_INT}', '{L_HARV}', '{L_ADDR}', '{L_CAPT}', '{L_NONE}'. Every group is COUNTIF/SUMIF-able on the Sites sheet and the tie check remains exhaustive. The unattributed majority is now explained by class rather than left blank: at v43 {class_counts[L_NET]} lines / $3.07M are a whole-of-city network service, {class_counts[L_CAPT]} lines / $2.67M await invoice capture, {class_counts[L_ACCT]} lines are accounting events. A park attribution to any of these would be fabrication.",
 f"42.2 Whole-row source hierarchy (v43). The v42 build read six source fields. A scan of all 130 register columns on the 15,698 unattributed rows found park names in columns never read: 111 Ev Line items (printed) - 92 lines / $845,336 - and 112 Ev Works carried out (printed) - 87 lines / $416,278 - together 92% of the recoverable names; also 99 Ev Ship To (21 lines, all the Parks Depot), 80 Src User Fld 2, 127 Ev Anomalies. Order is now: 111 > 106 > 112 > 108 > 127 > 99 (depot only) > 42 > 22 > 23 > 80, first hit wins, source recorded as the basis. Six columns are BANNED as sources: 12 Contractor and 32 Coding Note (own analysis, circular), 25 Nature Detail (derived from the sighted lines already read at 111), 92 Ev Vendor address (supplier premises - 20 false hits on '15 Helios St, Shailer Park'), 98 Ev Bill To (Council's address), 105 Ev PK normalised. Standing rule: a supplier's premises, Council's billing address and Claude's own labels are never a site source.",
 f"42.3 Multi-site holds extended (v43). {n_multi} lines now held (v42 319). Precedent from the new sources: the Pool Shop QLD monthly standing-order invoices print 'Alexander Park; Darlington Parklands; Logan Gardens' in the line items field on every invoice, so all fifteen ($394,264.81) are held as multi-site rather than attributed to PK000022's header. Ship-to = Parks Depot is a non-park Council facility ({class_counts[L_DEPOT]} lines) and is classed, not attributed. Two blocklist additions: the label 'Undertake General Park' (a facility-type construction that had leaked into Panel A at v42) is removed from the gazetteer; harvest prefixes Mobile / Storm / Maintenance / Supply / Install are blocked.",
 f"42.4 Verification of the new writes (v43). The {len(verify_rows)} park writes made from the newly read columns are exact gazetteer-name matches (the same evidence class the v42 build wrote from col 106) and are listed line by line in b43_site_writes_for_verification.json (register row, LineKey, amount, site, basis, source column, source text). Open Items #129 holds them for the user's check; any that fail revert to '{L_HARV}' at the next build. The 45 gazetteer candidates harvested from the '{L_HARV}' class (Quandong Park $30,255, Horseshoe Park $21,950, Lavell Park -> Lavelle Park, Tygum Skate Park -> Tygum Park, and others) ship in PSWP_Site_Attribution_Worklist_v42_fullscan.xlsx for adjudication and are NOT written.",
 f"42.5 Address route deferred (v43). {class_counts[L_ADDR]} lines / $117,089 carry a supply-point street address and no park name (Origin Energy dominates, 3,026 lines). A street-name join to the Parks Directory was tested and REJECTED: 26 directory streets serve two or more parks (Underwood Rd four, Paradise Rd four), and OpenStreetMap geocoding returns the road centre with no house-number precision, so a 200 m point-in-polygon test disagreed with the directory on the three largest supply points. The route reopens on property-level coordinates only: G-NAF address points (QLD_ADDRESS_DEFAULT_GEOCODE) or Council's own cadastre keyed on Property_Key, with point-in-polygon against the LoganHub Existing parks layer at zero buffer and every hit flagged for user check.",
]
for i,t in enumerate(METH): m.cell(mr+i,1,t)._style=ms
d=wb['Data_Acquisition']; dr=d.max_row+1; ds=copy.copy(d.cell(d.max_row,1)._style)
DA=[' |  G4  |  Logan City Council Parks Directory (public open data): 411 park points, EPSG:4326, with address, suburb and a facility inventory (playground 288, picnic 318, electric BBQ 91, toilets 68, fitness 75, basketball 99, skate 28, dog off-leash 44). Held as logan_park_gazetteer.csv in the session zip. Coordinate source for the map; NOT a site-attribution source (rule 19.9).  |  2  |  Held  | ',
    ' |  G5  |  LoganHub Logan Planning Scheme v9.2 TLPI No.1 2024 MapServer (arcgis.lcc.wspdigital.com), layer 121 Existing parks: 516 polygons with Park_Name, Hierarchy, Function; plus LPS2015 Suburbs (70), Divisions (12), LCC Boundary from services5.arcgis.com. Held as GeoJSON in the session zip. Reference geometry only.  |  2  |  Held  | ',
    ' |  G6  |  Parks Strategy Layer (LGIP, 5-Nov-2025): 2,880 parcel rows, 1,275 park names, Division, Suburb, Hierarchy, Function, Embellishment, Ownership, Status, Leases, three area columns (Property_Area_sqm 135.1M, Shape_Area_sqm 138.0M, Park_Unique_Area_sqm 333.2M - authoritative column UNRESOLVED; no per-hectare figure ships until it is). Exported without geometry (SHAPE column reads the literal "Polygon"). Request a re-export with geometry.  |  2  |  Held  | ',
    ' |  G7  |  G-NAF August 2026 (PSMA/Geoscape open data): user holds the national release. Requested: QLD_ADDRESS_DETAIL, QLD_ADDRESS_DEFAULT_GEOCODE, QLD_STREET_LOCALITY, QLD_LOCALITY, QLD_STREET_LOCALITY_ALIAS and the Authority Code folder, to be filtered to the 70 Logan localities on load and shipped as logan_gnaf_addresses.csv. Purpose: property-level geocode of the supply-point addresses for point-in-polygon against layer 121, every hit flagged for user check.  |  1  |  Requested  | ',
    f'v43 update (18-Aug-2026): enrichment build, no invoice capture; sighted count unchanged at 1,877; control total unchanged at ${CT:,.2f}. Themes restated four-year with live controls; site basis made exhaustive over all 130 columns; {n_single} lines on {len(parks)} parks, {n_multi} multi-site holds; {len(verify_rows)} new park writes queued for user verification (Open Items #129).']
for i,t in enumerate(DA): d.cell(dr+i,1,t)._style=ds
oi=wb['Open_Items']; orow=oi.max_row+1; os_=[copy.copy(oi.cell(oi.max_row,c)._style) for c in range(1,8)]
ITEMS=[
 (129,'All','Open','Site dimension',len(verify_rows),round(sum(v['amount'] for v in verify_rows),2),f'NEW PARK WRITES PENDING USER CHECK. {len(verify_rows)} lines attributed at v43 from register columns not read at v42 (111 line items, 112 works narrative, 108, 127, 22, 80). Exact gazetteer-name matches, listed in b43_site_writes_for_verification.json with source text. User to check; failures revert to the candidate-harvest class next build.'),
 (130,'All','Open','Site dimension',class_counts[L_HARV],0,f'CANDIDATE HARVEST. {class_counts[L_HARV]} green-blocked lines print a site name absent from the 483-name gazetteer. 45 candidates ship in PSWP_Site_Attribution_Worklist_v42_fullscan.xlsx (Quandong Park $30,255; Horseshoe Park $21,950 - NOT Herses Park; Lavell Park -> Lavelle Park; Tygum Skate Park -> Tygum Park (facility within); Jimmy Phillips Park new). Adjudicate, extend the gazetteer, rerun.'),
 (131,'All','Open','Site dimension',class_counts[L_ADDR],117089.00,'ADDRESS ROUTE DEFERRED. Supply-point addresses with no park name (Origin Energy 3,026 lines). Street-name join rejected (26 ambiguous streets; road-centre geocodes disagree with the directory). Reopen on G-NAF property-level points or Council cadastre keyed on Property_Key, zero-buffer point-in-polygon, every hit user-flagged.'),
 (132,'All','Open','Parks Depot',class_counts[L_DEPOT],0,'DEPOT SPEND ON PARKS PKs. Ship-to prints Logan City Council Parks Depot, 177 Chambers Flat Rd Marsden on green-blocked invoices charged to park PKs; a further 26 unsighted lines / $119,080 narrate Parks Depot. Classed as a non-park facility. Confirm whether depot supplies and works should sit on a depot cost centre rather than PK000022.'),
 (133,'FY2024/25','Open','T11 Plumbing & water assets',1,-154596.00,'LARGE REVERSAL INSIDE AN OPERATIONAL THEME. Jul-2024 carries a T11 net of -$154,596 (accrual reversal landing in an operational theme rather than T9). Themes reads T11 FY2024/25 understated by that amount; recode to T9 or note in Method 1.0.'),
 (134,'All','Open','Themes',0,0,'THEMES CONTROLS WERE FALSE FROM v29 TO v42. The grand-total control tested $10,677,585.33 (v14) and the FY2026/27 control tested a P1 subtotal, so both read FALSE for thirteen builds without being caught, because the verify audited only the EI controls and the Coverage tie. Repaired v43; the verify now asserts every Themes control exact TRUE. Audit any other sheet-level control not yet in the verify.'),
]
for i,it in enumerate(ITEMS):
    for j,v in enumerate(it,1): c=oi.cell(orow+i,j,v); c._style=copy.copy(os_[j-1])

# ---------------------------------------------------------------- 7. save -> recalc (convert route) -> verify -> ship
pre=f'{OUT_DIR}/{NEW}'; wb.save(pre); print('saved pre-recalc')
prof=f'{ROOT}/lo_profile/user'; os.makedirs(prof,exist_ok=True)
open(f'{prof}/registrymodifications.xcu','w').write('''<?xml version="1.0" encoding="UTF-8"?>
<oor:items xmlns:oor="http://openoffice.org/2001/registry" xmlns:xs="http://www.w3.org/2001/XMLSchema" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
<item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop></item>
<item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop></item>
</oor:items>''')
t0=time.time()
res=subprocess.run(['soffice',f'-env:UserInstallation=file://{ROOT}/lo_profile','--headless','--calc','--convert-to','xlsx','--outdir',RECALC_DIR,pre],capture_output=True,text=True,timeout=1800)
final=f'{RECALC_DIR}/{NEW}'; assert os.path.exists(final), res.stdout+res.stderr
print(f'recalc {time.time()-t0:.0f}s')
json.dump({'PA':[PA_START,PA_END],'GRP':[GRP_START,GRP_END],'TOT_ROW':TOT_ROW,'CHK_ROW':CHK_ROW,'PB':[PB_TITLE+2,PB_END],
           'parks':len(parks),'single':n_single,'multi':n_multi,'none':n_none,'classes':class_counts,'CLASS_LABELS':CLASS_LABELS,
           'TH_TOT':TH_TOT,'TH_C':[TH_C0,TH_C1],'writes':len(verify_rows),'final':final,'CT':CT,'gb_count':int(sighted.sum()),'leak':n_leak},open(f'{ROOT}/b43_pos.json','w'),indent=1)
r=subprocess.run([sys.executable,f'{ROOT}/b43_verify.py'],capture_output=True,text=True); print(r.stdout); print(r.stderr[-2000:])
if 'FAILS: none' in r.stdout:
    os.makedirs('/mnt/user-data/outputs',exist_ok=True)
    subprocess.run(['cp',final,f'/mnt/user-data/outputs/{NEW}']); print('SHIPPED',NEW)
else:
    print('NOT SHIPPED'); sys.exit(1)
