#!/usr/bin/env python3
"""b37_build.py - v40 one-call build (rules 18/19.7/20) for Batch 37 (binder_batch_20260818, 64 invoices).
Self-contained: the pswp_build_batch.py driver was not in the session uploads, so this script reproduces its write set,
citation-map rewrites, control retargeting, convert-route recalc and exact-TRUE verify. Positions read from Config (19.6).
Inputs: corpus_b39.json, brief_b39.json, pi_v10.txt, v39 workbook. Output: v40 workbook (recalculated) + build_summary_b39.json.
Console: counts only."""
import json, re, copy, os, sys, subprocess, shutil, time, datetime
from decimal import Decimal, ROUND_HALF_UP
import openpyxl
from openpyxl.utils import get_column_letter as CL

ROOT = '/home/claude/b39'
SRC = '/mnt/user-data/outputs/PS_WP_Transaction_Register_3FY_v40_HANDOVER.xlsx'
OUT_DIR = f'{ROOT}/out'; RECALC_DIR = f'{ROOT}/out/recalc'
NEW_NAME = 'PS_WP_Transaction_Register_3FY_v41_HANDOVER.xlsx'
corpus = json.load(open(f'{ROOT}/corpus_b39.json')); brief = json.load(open(f'{ROOT}/brief_b39.json'))
pi_v10 = open(f'{ROOT}/pi_v10.txt').read().rstrip('\n').split('\n')
D = lambda x: Decimal(str(x))
STAMP_DATE = '18-Aug-2026'

def condense(s):
    s = re.sub(r'[ \t]{3,}', '  ', s); s = re.sub(r'\n{3,}', '\n\n', s); s = s.replace('\f', '')
    return s.strip()

def dt(s):  # 'YYYY-MM-DD' -> datetime
    y, m, d = map(int, s.split('-')); return datetime.datetime(y, m, d)

def money_s(x):
    return f'${D(x).quantize(Decimal("0.01")):,.2f}'

t0 = time.time()
wb = openpyxl.load_workbook(SRC)
reg, ei, eil = wb['Register'], wb['Evidence_Invoices'], wb['Evidence_Invoice_Lines']
cfg = wb['Config']
CFG = {}; cfg_row = {}
for r in range(1, cfg.max_row + 1):
    k = cfg.cell(r, 1).value
    if k: CFG[k] = cfg.cell(r, 2).value; cfg_row[k] = r
assert CFG['WORKBOOK_VERSION'] == 'v40'
EI_END = int(CFG['EI_DATA'].split(':')[1]); EI_TOT = int(CFG['EI_TOTALS_ROW']); EI_C0, EI_C1 = map(int, CFG['EI_CONTROL_ROWS'].split(':'))
EIL_END = int(CFG['EIL_END']); R0, R1 = map(int, CFG['RECON_ROWS'].split(':')); RECON_COUNT = int(CFG['RECON_COUNT'])
REG_END = int(CFG['REGISTER_DATA'].split(':')[1]); REG_TOT = int(CFG['REGISTER_TOTAL_ROW']); CTRL_TOTAL = D(CFG['CONTROL_TOTAL'])
SIGHTED = int(CFG['SIGHTED_COUNT']); VMIN = int(CFG['VARIANT_TAG_MIN']); GSTINC = D(CFG['GSTINC_ALLOWANCE']); BPK = int(CFG['BOILERPLATE_KEYS'])
assert (EI_END, EI_TOT, EI_C0, EI_C1, EIL_END, R0, R1, RECON_COUNT) == (1956, 1958, 1964, 1977, 4785, 4789, 6740, 1952)
for ws in (ei, eil, reg):
    assert not [m for m in ws.merged_cells.ranges if m.min_row >= 1950 or (ws is eil and m.min_row >= 4780)], ws.title

N_INV = len(corpus); N_LINES = sum(len(i['lines']) for i in corpus)
VAR = brief.get('variants', {})
NEW_EI_END = EI_END + N_INV                 # 1892
NEW_EI_TOT = NEW_EI_END + 2                 # 1894
SHIFT = N_INV                               # +40 for the EI tail
NEW_EIL_END = EIL_END + N_LINES             # 4626
NEW_R0 = NEW_EIL_END + 4                    # 4630
NEW_RH = NEW_R0 - 1                         # 4629 header
NEW_RECON_COUNT = RECON_COUNT + N_INV       # 1888
NEW_R1 = NEW_R0 + NEW_RECON_COUNT - 1       # 6517
UNMATCHED = set(brief.get('unmatched', []))
N_GREEN = sum(1 for i in corpus if i['id'] not in UNMATCHED)
NEW_SIGHTED = SIGHTED + N_GREEN
NEW_VMIN = VMIN + brief['variant_count_new']  # 873
NEW_GSTINC = GSTINC + D(brief['gstinc_add'])  # 40890.42
NEW_BPK = BPK + len(brief['new_bp'])        # 75

# ------------------------------------------------------------------ helpers
def copy_style(src, dst):
    dst._style = copy.copy(src._style)

def row_snapshot(ws, r, ncol):
    return [(ws.cell(r, c).value, copy.copy(ws.cell(r, c)._style)) for c in range(1, ncol + 1)]

def row_restore(ws, r, snap):
    for c, (v, st) in enumerate(snap, 1):
        cell = ws.cell(r, c); cell.value = v; cell._style = st

def clear_row(ws, r, ncol):
    for c in range(1, ncol + 1): ws.cell(r, c).value = None

# ------------------------------------------------------------------ Evidence_Invoice_Lines: append lines, relocate recon block
eil_tmpl = row_snapshot(eil, EIL_END, 11)
recon_snap = [row_snapshot(eil, r, 11) for r in range(R0 - 1, R1 + 1)]      # header + 1848 rows
old_header_text = eil.cell(R0 - 1, 3).value
for r in range(EIL_END + 1, R1 + 1): clear_row(eil, r, 11)
line_rows = {}   # evid -> (first,last)
r = EIL_END
for inv in corpus:
    m = brief['matches'][inv['id']]
    first = r + 1
    for ln in inv['lines']:
        r += 1
        lk = '' if inv['id'] in UNMATCHED else m['linekey']
        note = ln['notes']
        if inv['id'] in UNMATCHED:
            note = (note + '; ' if note else '') + 'No PS/WP register line (rule 16 capture only); LineKey blank.'
        vals = [inv['evid'], ln['n'], ln['desc'], ln['qty'] if ln['qty'] != '(not printed)' else '', float(D(ln['unit'])) if ln['unit'] != '(not printed)' else '',
                ln['gst'], float(D(ln['amount'])), ln['pk_printed'], ln['pk_norm'], lk, note]
        for c, v in enumerate(vals, 1):
            cell = eil.cell(r, c); cell.value = v; cell._style = copy.copy(eil_tmpl[c - 1][1])
    line_rows[inv['evid']] = (first, r)
assert r == NEW_EIL_END
# recon block
row_restore(eil, NEW_RH, recon_snap[0])
eil.cell(NEW_RH, 3).value = f'Per-invoice reconciliation (rule 16b): invoice | captured line sum | ties the printed target (v41 block; all ranges full: $A$5:$A${NEW_EIL_END})'
recon_map = {}
rr = NEW_R0 - 1
for snap in recon_snap[1:]:
    rr += 1
    row_restore(eil, rr, snap)
    evid = eil.cell(rr, 3).value
    eil.cell(rr, 4).value = f'=ROUND(SUMIF($A$5:$A${NEW_EIL_END},"{evid}",$G$5:$G${NEW_EIL_END}),2)'
    e = eil.cell(rr, 5).value
    assert isinstance(e, str) and e.startswith('=IF(') and re.search(r'D\d+', e), (rr, e)
    eil.cell(rr, 5).value = re.sub(r'D\d+', f'D{rr}', e)
    assert evid not in recon_map, evid; recon_map[evid] = rr
recon_tmpl = recon_snap[-1]
for inv in corpus:
    rr += 1
    row_restore(eil, rr, recon_tmpl)
    evid = inv['evid']; tgt = D(inv['total']) if inv.get('gst_inclusive_lines') else D(inv['subtotal'])
    eil.cell(rr, 3).value = evid
    eil.cell(rr, 4).value = f'=ROUND(SUMIF($A$5:$A${NEW_EIL_END},"{evid}",$G$5:$G${NEW_EIL_END}),2)'
    tv = f'{tgt.normalize():f}'
    if VAR.get(inv['id'], {}).get('check1') == 'tol1c':
        # Ratified check-1 per-line rounding tolerance (v41): the printed subtotal sums
        # unrounded line values, so the recon row ties within one cent.
        eil.cell(rr, 5).value = f'=IF(ROUND(ABS(D{rr}-{tv}),2)<=0.01,"TRUE","FALSE")'
    elif inv.get('gst_inclusive_lines'):
        eil.cell(rr, 5).value = f'=IF(ROUND(ABS(D{rr}-{tv}),2)=0,"TRUE","FALSE")'
    else:
        eil.cell(rr, 5).value = f'=IF(D{rr}={tv},"TRUE","FALSE")'
    recon_map[evid] = rr
assert rr == NEW_R1 and len(recon_map) == NEW_RECON_COUNT
for r2 in range(NEW_R1 + 1, NEW_R1 + 3): clear_row(eil, r2, 11)

# ------------------------------------------------------------------ Register: green blocks, analysis cols, companions
tmpl_row = 14937
gb_tmpl = {c: copy.copy(reg.cell(tmpl_row, c)._style) for c in range(88, 128)}
assert reg.cell(tmpl_row, 88).value == 'INV-2526'
def kv(fam): return brief['families'][fam]
green_rows, companion_rows, identity_fixed, evid_of_row = [], [], [], {}
TEMPLATE_NOTE = brief['template_notes']
for inv in corpus:
    if inv['id'] in UNMATCHED:
        continue
    m = brief['matches'][inv['id']]; r = m['row']; fam = kv(inv['family']); evid = inv['evid']
    printed_ref = str(reg.cell(r, 8).value).strip()
    assert printed_ref in (inv['id'], '380CFR1', str(inv['id']).lstrip('0')) and reg.cell(r, 88).value in (None, ''), (r, inv['id'])
    assert abs(D(reg.cell(r, 20).value) - D(inv['subtotal'])) <= Decimal('0.01')
    for held in brief['held_rows']: assert r != held
    # identity (Tier 1 from the sighted invoice); label must match Vendor_Series exactly
    if reg.cell(r, 11).value in (None, '') or reg.cell(r, 12).value != fam['label'] or reg.cell(r, 13).value in (None, ''):
        reg.cell(r, 11).value = fam['creditor']; reg.cell(r, 12).value = fam['label']; reg.cell(r, 13).value = fam['abn']; identity_fixed.append(r)
    nat_detail = inv['nature_detail']
    if inv['site'] not in ('(not printed)', '(blank as printed)'):
        nat_detail = f"{nat_detail} - {inv['site']}"
    nat_detail = f"{nat_detail} ({evid})"
    if len(nat_detail) > 1200: nat_detail = nat_detail[:1197] + '...'
    reg.cell(r, 24).value = inv['nature_category']
    reg.cell(r, 25).value = nat_detail
    reg.cell(r, 27).value = 'Sighted invoice line'
    a, b = line_rows[evid]
    ev = (reg.cell(r, 28).value or '').rstrip()
    ev_add = (f" Sighted tax invoice {evid}, {inv['vendor']}, ABN {inv['abn']}; line capture Evidence_Invoice_Lines rows {a}-{b}; reconciliation row {recon_map[evid]}.")
    if r in identity_fixed:
        ev_add = f" Identified Tier 1 from the sighted invoice ({fam['creditor']}, {inv['vendor']}, ABN {inv['abn']})." + ev_add
    reg.cell(r, 28).value = (ev + ('' if ev.endswith('.') or not ev else '.') + ev_add).strip()
    reg.cell(r, 29).value = 1
    reg.cell(r, 31).value = brief['verdicts'][inv['id']]
    reg.cell(r, 32).value = brief['coding_notes'][inv['id']]
    reg.cell(r, 33).value = 'Confirmed'
    fu = brief['followups'][inv['id']]
    reg.cell(r, 34).value = fu if fu else None
    # parsed references (write only where blank)
    def put(c, v):
        if v and v not in ('(not printed)', '(blank as printed)') and reg.cell(r, c).value in (None, ''): reg.cell(r, c).value = v
    put(35, inv['contract'])
    put(38, inv['cr_wo'])
    if inv['requesting_officer'] not in ('(not printed)', '(blank as printed)'):
        put(43, inv['requesting_officer'])
    # green block
    def item_str(l):
        if l['qty'] != '(not printed)' and l['unit'] != '(not printed)':
            return f"{l['n']}. {l['desc']} | {l['qty']} @ {money_s(l['unit'])} = {money_s(l['amount'])}."
        return f"{l['n']}. {l['desc']} = {money_s(l['amount'])}."
    items = ' '.join(item_str(l) for l in inv['lines'])
    anomalies = list(inv['anomalies'])
    anomalies.append(TEMPLATE_NOTE[inv['family']])
    for l in inv['lines']:
        if len(l['desc_rows']) > 1:
            anomalies.append(f"Line {l['n']} description printed across {len(l['desc_rows'])} layout rows and joined verbatim (F12).")
    if inv['id'] in VAR and VAR[inv['id']].get('note'):
        anomalies.append(VAR[inv['id']]['note'])
    anomalies.append(f"Payment details cited {inv['payment_bp']}; terms cited {inv['terms_bp']} (character-exact against the page text, whitespace runs condensed).")
    gb = {88: evid, 89: inv['vendor'], 90: inv['abn'], 91: inv['acn'], 92: inv['address'], 93: inv['phone'], 94: dt(inv['invoice_date']), 95: (dt(inv['due_date']) if inv['due_date'] and inv['due_date'] not in ('(not printed)',) else '(not printed)'),
          96: inv['purchase_order'], 97: inv['contract'], 98: inv['bill_to'], 99: inv['ship_to'], 100: inv['requesting_officer'], 101: inv['request_date'], 102: inv['request_via'],
          103: inv['cr_wo'], 104: inv['pk_printed'], 105: inv['pk_norm'], 106: inv['site'], 107: inv['site_contact'], 108: inv['work_description'], 109: inv['technician'],
          110: len(inv['lines']), 111: items, 112: condense(inv['narrative']), 113: float(D(inv['subtotal'])), 114: float(D(inv['gst'])), 115: float(D(inv['total'])),
          116: float(D(inv['payments_made'])) if inv['payments_made'] != '(not printed)' else '(not printed)',
          117: float(D(inv['balance_due'])) if inv['balance_due'] != '(not printed)' else '(not printed)',
          118: inv['payment_bp'], 119: inv['terms_bp'], 120: inv['pages'],
          121: f'=ROUND(SUMIF(Evidence_Invoice_Lines!$A$5:$A${NEW_EIL_END},CJ{r},Evidence_Invoice_Lines!$G$5:$G${NEW_EIL_END}),2)',
          122: (f'=IF(ROUND(ABS(DQ{r}-DI{r}),2)<=0.01,"TRUE","FALSE")' if VAR.get(inv['id'], {}).get('check1') == 'tol1c'
                else f'=IF(DQ{r}=DI{r},"TRUE","FALSE")'),
          123: (f'=IF(ROUND(ABS(ROUND(T{r},2)-DI{r}),2)<=0.01,"TRUE","FALSE")' if VAR.get(inv['id'], {}).get('check2') == 'tol1c'
                else f'=IF(ROUND(T{r},2)=DI{r},"TRUE","FALSE")'),
          124: (f'=IF(ROUND(ABS(DJ{r}-ROUND(DI{r}*0.1,2)),2)<=0.01,"TRUE","FALSE")' if VAR.get(inv['id'], {}).get('check3') == 'tol1cgst'
                else f'=IF(DJ{r}=ROUND(DI{r}*0.1,2),"TRUE","FALSE")'),
          125: f"{brief['batch_label']} (source file {inv['source_file']})", 126: f'{STAMP_DATE}, rule 17 capture (v41 batch build)', 127: condense(' '.join(anomalies))}
    for c, v in gb.items():
        cell = reg.cell(r, c); cell.value = v; cell._style = copy.copy(gb_tmpl[c])
    green_rows.append(r); evid_of_row[r] = evid
    for cr in m['companions']:
        assert reg.cell(cr, 8).value == inv['id'] and D(reg.cell(cr, 20).value) == 0 and reg.cell(cr, 88).value in (None, '')
        reg.cell(cr, 32).value = (f'Zero-amount companion AP row of LineKey {m["linekey"]} (register row {r}), whose sighted invoice {evid} carries the full amount {money_s(inv["subtotal"])}; '
                                  f'no green block on a $0.00 line (rule 17). Supplier and invoice sighted.')
        companion_rows.append(cr)
assert len(green_rows) == N_GREEN, (len(green_rows), N_GREEN)

# ------------------------------------------------------------------ Evidence_Invoices: append rows, relocate tail, retarget controls
ei_tail = [row_snapshot(ei, r, 32) for r in range(EI_END + 1, EI_C1 + 1)]   # 1853..1873
ei_tmpl = row_snapshot(ei, EI_END, 32)
for r in range(EI_END + 1, EI_C1 + 1): clear_row(ei, r, 32)
r = EI_END
STAMP = f"{brief['batch_label']}, captured {STAMP_DATE} (v41)"
for inv in corpus:
    r += 1; m = brief['matches'][inv['id']]; evid = inv['evid']
    lcc_ref, cq = inv['id'], inv['contract']
    if inv['id'] in UNMATCHED:
        lcc_ref = f"(no register match; Open Items #{brief['unmatched_open_item']})"
        ver = (f"Lines reconcile to the printed subtotal to the cent; reconciliation at "
               f"Evidence_Invoice_Lines row {recon_map[evid]}. Register row (no register match).")
        staged = brief['unmatched_staging']
        vals = {1: evid, 2: inv['vendor'], 3: inv['abn'], 4: dt(inv['invoice_date']),
                5: '(not printed)', 6: lcc_ref, 7: cq, 8: inv['site'], 9: len(inv['lines']),
                10: float(D(inv['subtotal'])), 11: float(D(inv['gst'])), 12: float(D(inv['total'])),
                13: ver, 14: staged, 32: STAMP}
        for c in range(1, 33):
            cell = ei.cell(r, c); cell.value = vals.get(c); cell._style = copy.copy(ei_tmpl[c - 1][1])
        continue
    ver = (f"Lines reconcile to the printed subtotal to the cent; reconciliation at "
           f"Evidence_Invoice_Lines row {recon_map[evid]}. Rule 17 checks live on Register row {m['row']}.")
    vals = {1: evid, 2: inv['vendor'], 3: inv['abn'], 4: dt(inv['invoice_date']),
            5: (dt(inv['due_date']) if inv['due_date'] and inv['due_date'] != '(not printed)' else '(not printed)'),
            6: lcc_ref, 7: cq, 8: inv['site'], 9: len(inv['lines']),
            10: float(D(inv['subtotal'])), 11: float(D(inv['gst'])), 12: float(D(inv['total'])),
            13: ver, 14: 'Complete on the register line (green block, columns CJ to DW)', 32: STAMP}
    for c in range(1, 33):
        cell = ei.cell(r, c); cell.value = vals.get(c); cell._style = copy.copy(ei_tmpl[c - 1][1])
assert r == NEW_EI_END
for i, snap in enumerate(ei_tail):
    row_restore(ei, EI_END + 1 + SHIFT + i, snap)
# totals row explicit
for c in (9, 10, 11, 12):
    L = CL(c); ei.cell(NEW_EI_TOT, c).value = f'=SUM({L}5:{L}{NEW_EI_END})'
NEW_C0, NEW_C1 = EI_C0 + SHIFT, EI_C1 + SHIFT
ei.cell(NEW_C0 - 1, 1).value = 'Rule 17 controls (v41)'
def retarget(f):
    f = re.sub(r'\b([CJ])(19[5-8]\d)\b', lambda mm: f'{mm.group(1)}{int(mm.group(2)) + SHIFT}', f)
    f = f.replace(f'$AF$5:$AF${EI_END + 1}', f'$AF$5:$AF${NEW_EI_END + 1}')
    f = f.replace(f'$G$5:$G${EIL_END}', f'$G$5:$G${NEW_EIL_END}').replace(f'{GSTINC:f}', f'{NEW_GSTINC:f}')
    f = f.replace(f'$E${R0}:$E${R1}', f'$E${NEW_R0}:$E${NEW_R1}').replace(f'={RECON_COUNT},', f'={NEW_RECON_COUNT},')
    f = f.replace(f'={SIGHTED},', f'={NEW_SIGHTED},').replace(f'>={VMIN},', f'>={NEW_VMIN},')
    return f
for rr in range(NEW_C0, NEW_C1 + 1):
    f = ei.cell(rr, 3).value
    if isinstance(f, str) and f.startswith('='): ei.cell(rr, 3).value = retarget(f)
    lab = ei.cell(rr, 1).value
    if isinstance(lab, str):
        lab = lab.replace(f'{SIGHTED:,}', f'{NEW_SIGHTED:,}').replace(f'{RECON_COUNT:,} of {RECON_COUNT:,}', f'{NEW_RECON_COUNT:,} of {NEW_RECON_COUNT:,}')
        lab = lab.replace(f'>= {VMIN}', f'>= {NEW_VMIN}').replace(f'${GSTINC:,.2f}', f'${NEW_GSTINC:,.2f}')
        ei.cell(rr, 1).value = lab
assert f'C{NEW_C0}={NEW_SIGHTED}' in ei.cell(NEW_C0 + 2, 3).value and f'$E${NEW_R0}:$E${NEW_R1}' in ei.cell(NEW_C0 + 9, 3).value and f'{NEW_GSTINC:f}' in ei.cell(NEW_C0 + 7, 3).value
assert f'>={NEW_VMIN},' in ei.cell(NEW_C0 + 10, 3).value and f'$AF$5:$AF${NEW_EI_END + 1}' in ei.cell(NEW_C1, 3).value and f'J{NEW_EI_TOT}' in ei.cell(NEW_C0 + 7, 3).value

# ------------------------------------------------------------------ Citation limbs (rule 19.10): both limbs from the maps
reg_map = {}
for rr in range(5, REG_END + 1):
    v = reg.cell(rr, 88).value
    if v not in (None, ''): reg_map.setdefault(str(v).strip(), []).append(rr)
recon_rewritten = 0; reg_limb_checked = 0; reg_limb_fixed = 0
for rr in range(5, NEW_EI_END + 1):
    evid = str(ei.cell(rr, 1).value).strip(); t = ei.cell(rr, 13).value
    if not isinstance(t, str): continue
    if evid in recon_map and re.search(r'Evidence_Invoice_Lines row \d+', t):
        t2 = re.sub(r'Evidence_Invoice_Lines row \d+', f'Evidence_Invoice_Lines row {recon_map[evid]}', t)
        if t2 != t: recon_rewritten += 1
        t = t2
    mm = re.search(r'([Rr]egister row(?:\(s\))?\s+)((?:\d+)(?:(?:,\s*| and )\d+)*)', t)
    if mm:
        reg_limb_checked += 1
        cited = sorted({int(x) for x in re.findall(r'\d+', mm.group(2))}); actual = sorted(reg_map.get(evid, []))
        if actual and cited != actual:
            t = t[:mm.start(2)] + ' and '.join(str(x) for x in actual) + t[mm.end(2):]; reg_limb_fixed += 1
    ei.cell(rr, 13).value = t
col28_rewritten = 0
for rr, evid in ((rr, str(reg.cell(rr, 88).value).strip()) for rr in range(5, REG_END + 1) if reg.cell(rr, 88).value not in (None, '')):
    t = reg.cell(rr, 28).value
    if isinstance(t, str) and evid in recon_map and re.search(r'recon(?:ciliation)? row \d+', t):
        t2 = re.sub(r'(recon(?:ciliation)? row )\d+', lambda mm: f'{mm.group(1)}{recon_map[evid]}', t)
        if t2 != t: col28_rewritten += 1; reg.cell(rr, 28).value = t2

# ------------------------------------------------------------------ Vendor_Boilerplate
vb = wb['Vendor_Boilerplate']
vb_last = max(r for r in range(1, vb.max_row + 1) if str(vb.cell(r, 1).value or '').startswith('BP:'))
vb_keys = {str(vb.cell(r, 1).value): r for r in range(5, vb.max_row + 1) if str(vb.cell(r, 1).value or '').startswith('BP:')}
assert len(vb_keys) == BPK
new_keys = {nb['key'] for nb in brief['new_bp']}
for k, n in brief['bp_bumps'].items():
    if k in new_keys:
        continue          # minted below with their sighting count
    vb.cell(vb_keys[k], 5).value = int(vb.cell(vb_keys[k], 5).value) + n
vb_tmpl = row_snapshot(vb, vb_last, 6)
cited = {}
for inv in corpus:
    for k in (inv['payment_bp'], inv['terms_bp']):
        if k and k != '(not printed)': cited[k] = cited.get(k, 0) + 1
for i, nb in enumerate(brief['new_bp'], 1):
    r = vb_last + i; row_restore(vb, r, vb_tmpl)
    for c, v in enumerate([nb['key'], nb['field'], nb['vendor'], nb['first'], cited[nb['key']], nb['text']], 1): vb.cell(r, c).value = v
    vb_keys[nb['key']] = r
for k in cited: assert k in vb_keys, k
NEW_VB_LAST = vb_last + len(brief['new_bp'])

# ------------------------------------------------------------------ Narrative sheets: Handover, Method, Data_Acquisition, Open_Items, Vendor_Series
S = brief['summary']
def append_rows(ws, rows, ncol=None, tmpl_row=None, gap=0):
    ncol = ncol or ws.max_column; last = ws.max_row
    while last > 1 and all(ws.cell(last, c).value in (None, '') for c in range(1, ncol + 1)): last -= 1
    tmpl = row_snapshot(ws, tmpl_row or last, ncol); r = last + gap
    for vals in rows:
        r += 1; row_restore(ws, r, tmpl)
        for c, v in enumerate(vals, 1): ws.cell(r, c).value = v
    return r
tot_ex = money_s(S['ex_gst'])
handover_txt = (
    f"Batch 38 plus two closed Open Items. levi_bind_3 (one source PDF, 43 pages, 43 documents, corpus md5 {brief['corpus_md5_supplied']}) "
    f"gated in-session: 41 in-scope invoices captured (T & H Levai 39, 4Park 1, TradeTools 1) and two 2019-vintage Levai invoices "
    f"left uncaptured as outside the register scope. Open Item #107 CLOSED: BLC 001291 ($885.00, register row 2020) recovered from the "
    f"page text of the merged Batch 37 record and captured. Open Item #108 CLOSED: the check-1 per-line rounding tolerance is RATIFIED, "
    f"releasing the four held Play Force invoices. Total {N_INV} invoices, {N_LINES} lines, {tot_ex} ex GST; {N_GREEN} register rows "
    f"green-blocked Confirmed, four check variants. Register total unchanged at {money_s(CTRL_TOTAL)}; sighted count {NEW_SIGHTED:,}. "
    f"Two new Tier 1 identities (4Park Pty Ltd, TradeTools Pty Ltd) replace Unidentified series labels. Findings Open Items #113-#116.")
append_rows(wb['Handover'], [[f'v41 change ({STAMP_DATE}): Batch 38 build (levi_bind_3) with Open Items #107 and #108 closed', handover_txt]], ncol=2)
method_rows = [
    [f'40.0 v41 build ({STAMP_DATE}): Batch 38 (F42), extraction corpus levi_bind_3 (md5 {brief["corpus_md5_supplied"]}, Microsoft 365 Copilot / GPT-5 reasoning) over "levi bind 3.pdf", 43 pages, 43 document records, full 43-page OCR coverage. '
     f'Gated and repaired in-session (b38_gate.py). {N_INV} invoices captured across three cohorts - Batch 38 proper (41), the Open Item #107 recovery (1) and the Open Item #108 release (4) - {N_LINES} lines, {tot_ex} ex GST; {N_GREEN} green blocks, four check variants. '
     f'EI data 5:{NEW_EI_END}, totals {NEW_EI_TOT}, controls {NEW_C0}:{NEW_C1}; EIL data 5:{NEW_EIL_END}, recon {NEW_R0}:{NEW_R1} ({NEW_RECON_COUNT}); sighted {NEW_SIGHTED:,}; VARIANT_TAG_MIN {NEW_VMIN}; BOILERPLATE_KEYS {NEW_BPK}. Config restated in the same build.'],
    ['40.1 RATIFIED: check-1 per-line rounding tolerance (Method 40.1; fold into instructions rule 17). Where a vendor prints each line amount rounded to two decimals but computes the printed subtotal from the UNROUNDED values, the captured lines cannot equal the printed subtotal and the printed document does not self-tie. '
     'Check 1 then takes the standard tolerance form =IF(ROUND(ABS(DQ{r}-DI{r}),2)<=0.01,"TRUE","FALSE"), tagged [check variant], with the delta and a worked example stated in the anomalies field. This is the same species as the already-ratified check-2 one-cent tolerance (TechOne ex-GST derivation) and check-3 rounding tolerances, and completes the set across all three checks. '
     'Precedent and first use: Play Force Australia INV-4412, INV-6739, INV-7609 and INV-7759 (5.50 x $105.41 = $579.755, printed $579.75), held uncaptured at v40 and released here. The reconciliation rows for these four take the ratified tolerance E-form. VARIANT_TAG_MIN 876 -> 880.'],
    ['40.2 Open Item #107 CLOSED - recovering an invoice the extractor never emitted. BLC 001291 ($885.00, 9-May-2025, Windaroo Lake Park wire balustrade reinstatement, CR 3682017, PO 707879) was absent from the Batch 37 corpus because document segmentation merged it into record 001290 (Binder111.pdf pages 13-16, of which 15-16 are this separate invoice). '
     'It was recovered from the retained page text of that merged record. The page-15 text layer inserts spurious spaces inside words ("L o g a n City Council", "Invoice n u m b e r : 001 291"), so every field was de-spaced and then cross-checked field by field against the eleven clean sibling BLC invoices captured at v40, which share the template and boilerplate verbatim; the printed figures carry no spacing artefact and were read directly. '
     'The basis is stated in the anomalies field per rule 16a: this is reconstruction of a defective text layer, not paraphrase. ROUTE: where a text layer is character-spaced but structurally intact, de-space and cross-check against a clean sibling of the same template rather than discarding the capture.'],
    ['40.3 Levai template and identities. The Levai (Xero) layout prints a DESCRIPTION / AMOUNT table with NO qty and NO unit-price column, so the extractor classified every line NARRATIVE and four documents captured $0.00 - the whole invoice lost, not a line (defect family F9, now seen on five documents across five vendors in two batches; it is the single highest-value fix for extraction prompt v3). '
     'The Levai letterhead address changes from "Lot 2 Industrial Ave" to "41 Industrial Ave" between May-2024 and Jun-2024 under the same ABN, so the address is parsed per invoice. Two new Tier 1 identities: 4Park Pty Ltd (56 657 333 296, trading as Forpark, WA) and TradeTools Pty Ltd (67 010 700 053), both previously carried as Unidentified series labels on account 72111. '
     'The 4Park printed Amount column is GST-INCLUSIVE and its extractor record lumped both printed lines into one $2,940.00 summary row; the TradeTools line amount was attached to a footer warranty clause. Both restored to the printed line tables (rule 16, a summary row in place of line capture is a critical failure).'],
    ['40.4 Findings: T & H Levai and Play Force Australia print the SAME premises, 41 Industrial Ave Logan Village, with adjacent landlines ((07) 3803 0032 and (07) 3803 1788) across 123 and 118 sighted green blocks respectively, on different ABNs - a related-party question for Finance (#113). '
     'Two 2019-vintage Levai invoices (00021470 $918.00 and 00022061 $2,775.00, the latter on PK000391) sit in the binder outside the register scope with no register line and are NOT captured; 00021470 has no text layer at all (#114). INV-36682 and INV-36683 bill an identical $1,520.00 on the same day and are confirmed as distinct jobs from their printed job lines (#115). '
     'Two Levai target rows were legacy Confirmed on the narration with no rule 17 evidence (the Open Items #95 pattern) and are completed here (#116).']]
append_rows(wb['Method'], method_rows, ncol=1)
da = wb['Data_Acquisition']
da_rows = [[None],
    [f" |  F42  |  Batch 38: extraction corpus levi_bind_3 (md5 {brief['corpus_md5_supplied']}) over \"levi bind 3.pdf\", 43 pages, 43 document records, 3,051 line records, full 43-page OCR coverage (PASS). Extraction tool identity: Microsoft 365 Copilot, GPT-5 reasoning model. "
     f"The supplied capture report declared 5 documents OUT. In-session gate and repair (b38_gate.py, b38_prepare.py): four F9 amount-column-only line tables recovered, one lumped two-line summary row split, one line amount moved off a footer row, one invoice identified from its OCR layer. "
     f"Rule 12 screen against Evidence_Invoices column A: ZERO re-sightings despite 48 Levai invoices captured at v37. Register match: 41 of 41 in-scope documents amount-corroborated, no reference-only matches. Corpus corpus_b38_repaired.json; driver corpus corpus_b39.json; brief brief_b39.json; build script b39_build.py."],
    [f"v41 update ({STAMP_DATE}): {N_INV} invoices captured under rules 16/17 ({tot_ex} ex GST, {N_LINES} lines): {N_GREEN} register rows green-blocked Confirmed, four check variants, {len(companion_rows)} companion row cross-referenced. Open Items #107 and #108 CLOSED. Findings #113-#116. "
     f"Outstanding: the two out-of-scope 2019 Levai invoices (#114) and the extraction prompt v3 defect queue (#112)."]]
append_rows(da, da_rows, ncol=1, tmpl_row=169)
oi = wb['Open_Items']
oi_rows = [[o['n'], o['fy'], o['status'], o['who'], o['lines'], float(D(o['amt'])), o['text']] for o in brief['open_items']]
append_rows(oi, oi_rows, ncol=7)
vs = wb['Vendor_Series']
append_rows(vs, [[f"Footnote, {STAMP_DATE} (v41): Batch 38 (levi_bind_3, {N_INV} invoices, {tot_ex} ex GST). T & H Levai Pty Ltd (65 100 395 480) sighted on 39 further invoices; its letterhead address changes from Lot 2 to 41 Industrial Ave, Logan Village between May-2024 and Jun-2024 under the same ABN. "
                  f"TWO NEW Tier 1 labels replace Unidentified series rows on account 72111: 4Park Pty Ltd (56 657 333 296, trading as Forpark, 36 Adams Drive Welshpool WA) and TradeTools Pty Ltd (67 010 700 053, 109 Park Road Slacks Creek). Both ABNs pass the ATO checksum and both were sighted for the exact register amount. "
                  f"BLC Queensland Pty Ltd gains invoice 001291 (Open Item #107). Play Force Australia Pty Ltd gains four invoices released by the ratified check-1 tolerance (Open Item #108). "
                  f"NOTE: T & H Levai Pty Ltd and Play Force Australia Pty Ltd print the SAME premises (41 Industrial Ave, Logan Village QLD 4207) with adjacent landlines on different ABNs - separate entities, related-party question raised at Open Items #113."]], ncol=1, gap=1)

# ------------------------------------------------------------------ Project_Instructions v10 (same-step rule)
pi = wb['Project_Instructions']
pi_tmpl = copy.copy(pi.cell(3, 1)._style)
for r in range(1, pi.max_row + 1): pi.cell(r, 1).value = None
for i, line in enumerate(pi_v10, 1):
    pi.cell(i, 1).value = line if line else None; pi.cell(i, 1)._style = copy.copy(pi_tmpl)

# ------------------------------------------------------------------ Config
newcfg = {'WORKBOOK_VERSION': 'v41', 'EIL_DATA': f'5:{NEW_EIL_END}', 'EIL_END': NEW_EIL_END, 'RECON_ROWS': f'{NEW_R0}:{NEW_R1}', 'RECON_COUNT': NEW_RECON_COUNT,
          'EI_DATA': f'5:{NEW_EI_END}', 'EI_TOTALS_ROW': NEW_EI_TOT, 'EI_CONTROL_ROWS': f'{NEW_C0}:{NEW_C1}', 'SIGHTED_COUNT': NEW_SIGHTED,
          'VARIANT_TAG_MIN': NEW_VMIN, 'GSTINC_ALLOWANCE': float(NEW_GSTINC), 'BOILERPLATE_KEYS': NEW_BPK}
for k, v in newcfg.items(): cfg.cell(cfg_row[k], 2).value = v

# ------------------------------------------------------------------ save -> recalc -> verify -> ship (one call)
os.makedirs(OUT_DIR, exist_ok=True); os.makedirs(RECALC_DIR, exist_ok=True)
pre = f'{OUT_DIR}/{NEW_NAME}'
wb.save(pre)
t_save = time.time()
res = subprocess.run(['soffice', f'-env:UserInstallation=file://{ROOT}/lo_profile', '--headless', '--calc', '--convert-to', 'xlsx', '--outdir', RECALC_DIR, pre],
                     capture_output=True, text=True, timeout=1500)
final = f'{RECALC_DIR}/{NEW_NAME}'
assert os.path.exists(final), res.stdout + res.stderr
t_recalc = time.time()

# ---- verify (python-calamine on the RECALCULATED file; exact "TRUE")
from python_calamine import CalamineWorkbook
cw = CalamineWorkbook.from_path(final)
V = {}
def sheet(n): return cw.get_sheet_by_name(n).to_python()
EIv, EILv, REGv, COVv, CFGv, VBv = sheet('Evidence_Invoices'), sheet('Evidence_Invoice_Lines'), sheet('Register'), sheet('Coverage'), sheet('Config'), sheet('Vendor_Boilerplate')
fails = []
def chk(name, cond, detail=''):
    V[name] = bool(cond)
    if not cond: fails.append((name, detail))
ctrl = [(r, EIv[r - 1][0], EIv[r - 1][2]) for r in range(NEW_C0, NEW_C1 + 1)]
chk('ei_controls_all_TRUE', all((c[2] == 'TRUE') if not isinstance(c[2], (int, float)) else int(c[2]) == NEW_SIGHTED for c in ctrl), [c for c in ctrl if not ((c[2] == 'TRUE') if not isinstance(c[2], (int, float)) else int(c[2]) == NEW_SIGHTED)])
chk('ei_controls_count_14', len(ctrl) == 14)
sighted = sum(1 for r in REGv[4:REG_END] if r[26] == 'Sighted invoice line')
chk('sighted_count', sighted == NEW_SIGHTED, sighted)
chk('register_total', D(REGv[REG_TOT - 1][19]).quantize(Decimal('0.01')) == CTRL_TOTAL, REGv[REG_TOT - 1][19])
chk('coverage_tie', D(COVv[141][2]).quantize(Decimal('0.01')) == CTRL_TOTAL, COVv[141][2])
bad3 = [(r, REGv[r - 1][121], REGv[r - 1][122], REGv[r - 1][123]) for r in green_rows if not (REGv[r - 1][121] == REGv[r - 1][122] == REGv[r - 1][123] == 'TRUE')]
chk('new_rows_three_checks', not bad3, bad3[:5])
allbad = [i + 1 for i, r in enumerate(REGv[4:REG_END], 4) if r[26] == 'Sighted invoice line' and not (r[121] == r[122] == r[123] == 'TRUE')]
chk('all_sighted_three_checks', not allbad, allbad[:5])
recon = EILv[NEW_R0 - 1:NEW_R1]
chk('recon_all_TRUE', len(recon) == NEW_RECON_COUNT and all(r[4] == 'TRUE' for r in recon), [ (NEW_R0 + i, r[2], r[3], r[4]) for i, r in enumerate(recon) if r[4] != 'TRUE'][:5])
chk('recon_ids_unique', len({r[2] for r in recon}) == NEW_RECON_COUNT)
# citation limbs
recon_actual = {r[2]: NEW_R0 + i for i, r in enumerate(recon)}
reg_actual = {}
for i, r in enumerate(REGv[4:REG_END], 5):
    if r[87] != '': reg_actual.setdefault(str(r[87]).strip(), []).append(i)
c13_bad = []
for i, r in enumerate(EIv[4:NEW_EI_END], 5):
    evid = str(r[0]).strip(); t = str(r[12])
    mm = re.search(r'Evidence_Invoice_Lines row (\d+)', t)
    if mm and evid in recon_actual and int(mm.group(1)) != recon_actual[evid]: c13_bad.append((i, 'recon', mm.group(1), recon_actual[evid]))
    mm = re.search(r'([Rr]egister row(?:\(s\))?\s+)((?:\d+)(?:(?:,\s*| and )\d+)*)', t)
    if mm:
        cited_rows = sorted({int(x) for x in re.findall(r'\d+', mm.group(2))})
        if cited_rows != sorted(reg_actual.get(evid, [])): c13_bad.append((i, 'register', cited_rows, reg_actual.get(evid)))
chk('ei_col13_both_limbs', not c13_bad, c13_bad[:5])
c28_bad = []
for i, r in enumerate(REGv[4:REG_END], 5):
    if r[87] != '':
        mm = re.search(r'recon(?:ciliation)? row (\d+)', str(r[27]))
        if mm and int(mm.group(1)) != recon_actual.get(str(r[87]).strip()): c28_bad.append((i, mm.group(1), recon_actual.get(str(r[87]).strip())))
chk('register_col28_recon_limb', not c28_bad, c28_bad[:5])
chk('held_rows_no_evid', all(REGv[h - 1][87] == '' for h in brief['held_rows']))
chk('companions_no_evid_and_noted', all(REGv[c - 1][87] == '' and 'Zero-amount companion' in str(REGv[c - 1][31]) for c in companion_rows))
chk('new_rows_confirmed_evid', all(REGv[r - 1][32] == 'Confirmed' and REGv[r - 1][87] == evid_of_row[r] and REGv[r - 1][28] in (1, 1.0) and REGv[r - 1][11] != '' for r in green_rows))
stale = [r for r in green_rows if any(str(REGv[r - 1][31]).startswith(p) for p in brief['stale_note_prefixes']) or any(str(REGv[r - 1][33]).startswith(p) for p in brief['stale_followup_prefixes'])]
chk('no_stale_notes_on_targets', not stale, stale)
chk('nature_detail_nonempty', all(len(str(REGv[r - 1][24])) > 20 for r in green_rows))
vbkeys = {str(r[0]) for r in VBv[4:] if str(r[0]).startswith('BP:')}
chk('bp_keys_count', len(vbkeys) == NEW_BPK, len(vbkeys))
chk('bp_cited_exist', all(k in vbkeys for k in cited))
ei_new = EIv[EI_END:NEW_EI_END]
chk('ei_new_rows', len(ei_new) == N_INV and all(str(r[13]).startswith('Complete on the register line') or str(r[13]).startswith('Staged') for r in ei_new))
unm_evid = {i['evid'] for i in corpus if i['id'] in UNMATCHED}
chk('unmatched_no_register_evid', all(str(rr[87]).strip() not in unm_evid for rr in REGv[4:REG_END]))
chk('unmatched_staged', all(any(str(rr[0]).strip() == e and str(rr[13]).startswith('Staged') for rr in EIv[4:NEW_EI_END]) for e in unm_evid))
chk('ei_totals_row', str(EIv[NEW_EI_TOT - 1][0]).startswith('Totals') and abs(D(EIv[NEW_EI_TOT - 1][9]) - sum(D(str(r[9])) for r in EIv[4:NEW_EI_END] if isinstance(r[9], (int, float)))) < Decimal('0.01'), EIv[NEW_EI_TOT - 1][:12])
cfgd = {r[0]: r[1] for r in CFGv if r and r[0]}
chk('config_v39', cfgd.get('WORKBOOK_VERSION') == 'v41' and int(cfgd['EIL_END']) == NEW_EIL_END and cfgd['RECON_ROWS'] == f'{NEW_R0}:{NEW_R1}' and int(cfgd['SIGHTED_COUNT']) == NEW_SIGHTED)
variant_cnt = sum(1 for r in REGv[4:REG_END] if '[check variant]' in str(r[126]))
chk('variant_count', variant_cnt >= NEW_VMIN, variant_cnt)
lk = [r[0] for r in REGv[4:REG_END]]
chk('linekeys_unique', len(lk) == len(set(lk)) == REG_END - 4)
errs = sum(1 for sh in ('Register', 'Evidence_Invoices', 'Evidence_Invoice_Lines', 'Coverage', 'Vendor_Series') for row in sheet(sh) for c in row if isinstance(c, str) and c.startswith('#') and c[1:4] in ('REF', 'NAM', 'VAL', 'DIV', 'N/A'))
chk('no_formula_errors', errs == 0, errs)
summary = dict(version='v41', invoices=N_INV, lines=N_LINES, ex_gst=S['ex_gst'], green_rows=green_rows, companion_rows=companion_rows, identity_fixed=identity_fixed,
               positions=newcfg, ei_controls=[(c[0], c[2]) for c in ctrl], recon_rewritten_ei_col13=recon_rewritten, register_limb_checked=reg_limb_checked, register_limb_fixed=reg_limb_fixed,
               col28_rewritten=col28_rewritten, variant_count=variant_cnt, sighted=sighted, verify=V, fails=fails,
               timings=dict(load_build=round(t_save - t0), recalc=round(t_recalc - t_save), verify=round(time.time() - t_recalc)))
json.dump(summary, open(f'{ROOT}/build_summary_b39.json', 'w'), indent=1, default=str)
if fails:
    print('VERIFY FAILED', len(fails), fails[:6]); sys.exit(2)
os.makedirs('/mnt/user-data/outputs', exist_ok=True)
shutil.copy(final, f'/mnt/user-data/outputs/{NEW_NAME}')
print(f'SHIPPED {NEW_NAME} invoices {N_INV} lines {N_LINES} green {len(green_rows)} companions {len(companion_rows)} sighted {sighted} controls_TRUE {sum(1 for c in ctrl if c[2]=="TRUE")}/12+2 counts '
      f'recon {NEW_RECON_COUNT} col13_rewritten {recon_rewritten} reg_limb_fixed {reg_limb_fixed} col28_rewritten {col28_rewritten} variants {variant_cnt} '
      f'timings {summary["timings"]}')
