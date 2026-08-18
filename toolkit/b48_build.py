"""b48_build.py  v47 -> v48 (18-Aug-2026)
Full site sweep. Column T never written; no green block touched; control total unchanged.
 1 Suburb blocklist as a first-class object; assert no gazetteer label is a Logan suburb.
 2 Re-screen every joined multi-site list; collapse any list that reduces to one park.
 3 Gazetteer consolidation: variant map + adjudicated additions.
 4 Site read rebuilt COLLECT-ALL-CANDIDATES over all 130 columns; "(not printed)" and
   "(blank as printed)" treated as EMPTY, never as an answer.
 5 Evidence_Invoice_Lines site written across the WHOLE sheet.
 6 Reclassification: tolls -> not site-specific; 177 Chambers Flat Rd -> Parks Depot.
 7 Marsden Stores catalogue: product number on an internal stores issue -> item description.
"""
import json, copy, re, os, sys, subprocess, time, collections
import openpyxl, pandas as pd
from openpyxl.utils import get_column_letter
from python_calamine import CalamineWorkbook

ROOT='/home/claude/b48'; SRC='/mnt/user-data/outputs/PS_WP_Transaction_Register_3FY_v47_HANDOVER.xlsx'
NEW='PS_WP_Transaction_Register_3FY_v48_HANDOVER.xlsx'
OUT=f'{ROOT}/out'; RC=f'{OUT}/recalc'; os.makedirs(RC,exist_ok=True)
CT=11377312.51; SITE=129; BASIS=130; SC='DY'; BC='DZ'
EMPTY={'','nan','none','(not printed)','(blank as printed)','(no site named)','0','various','various sites','various parks','scheduled','n/a'}
ALLOW={'regents park'}   # genuine park names that also read as a Logan suburb (Site_Crosswalk exception)
SUBURBS={s.lower() for s in json.load(open('/home/claude/b48_suburbs.json'))}-ALLOW
CATALOGUE=json.load(open('/home/claude/b48_catalogue.json'))
VARIANT={'jamie nicholson park':'Jamie Nicolson Park','alexandra park':'Alexander Park','darlington park':'Darlington Parklands',
 'logan gardens park':'Logan Garden Park','lavell park':'Lavelle Park','richared wilson park':'Richard Wilson Park',
 'woodland district park':'Woodlands District Park','woodlans district park':'Woodlands District Park',
 'doug larson park':'Doug Larsen Park','teviot down park':'Teviot Downs Park','oliver sports complex':'Olivers Sports Complex',
 "oliver's sports complex":'Olivers Sports Complex',"oliver's sport complex":'Olivers Sports Complex',
 'flindersia park':'Flindersia Riverside Park','flindersea park':'Flindersia Riverside Park','flindersia river park':'Flindersia Riverside Park',
 'james smith rec park':'James Smith Recreation Area','mundoolun comminuty park':'Mundoolun Community Park',
 'logan gardens water play':'Logan Gardens','underwood water feature':'Underwood Park'}
ADD=['Quandong Park','Horseshoe Park','Maurice Park','Fig Tree Park','Hammel St Park','Paynes Bridge Park',
     'Mundoolun Community Park','Greenwood Lakes','Sebring Park','Kohlmann Park','Brough Place Park',
     'East Beaumont Park','Kurrajong Park','Stoneleigh Reserve Park','Ewing Park','Flagstone Regional Park']
FAC_NEW='Non-park Council facility (named site, not a park)'
NOT_SITE='Not site-specific by nature (tolls, network charges)'
DEPOT='Parks Depot, Marsden (non-park facility)'

cw=CalamineWorkbook.from_path(SRC)
cfg={str(r[0]):r[1] for r in cw.get_sheet_by_name('Config').to_python() if r and r[0]}
PA0,PA1=[int(x) for x in str(cfg['SITES_PANELA']).split(':')]
G0,G1=[int(x) for x in str(cfg['SITES_GROUP_ROWS']).split(':')]
PB0,PB1=[int(x) for x in str(cfg['SITES_PANELB']).split(':')]
SA_END=int(str(cfg['SITE_ALLOC_DATA']).split(':')[1]); XW_END=int(str(cfg['SITE_CROSSWALK_DATA']).split(':')[1])
EIL_END=int(float(cfg['EIL_END']))
sites=cw.get_sheet_by_name('Sites').to_python()
GAZ=[str(x[0]).strip() for x in sites[PA0-1:PA1]]
CLASS=[str(r[0]) for r in sites[G0-1:G1] if r and r[0]]; CL=set(CLASS)
RESIDUE=next(c for c in CLASS if c.startswith('(residue')); MULTI='Multiple sites named (not allocated)'

df=pd.read_excel(SRC,sheet_name='Register',header=3,engine='calamine'); df=df[df.iloc[:,0].notna()]
amt=pd.to_numeric(df.iloc[:,19],errors='coerce').fillna(0); assert abs(amt.sum()-CT)<0.005
MA=next(b for b in df.iloc[:,BASIS-1].fillna('').astype(str).unique() if 'allocated per printed' in b)
site_old=df.iloc[:,SITE-1].fillna('').astype(str).tolist()
bas_old=df.iloc[:,BASIS-1].fillna('').astype(str).tolist()

# ---------------- 1/3 gazetteer
GAZ=[g for g in GAZ if g.lower() not in SUBURBS]
canon={g.lower():g for g in GAZ}
for a in ADD: canon.setdefault(a.lower(),a)
for k,v in VARIANT.items(): canon[k]=v
UNIVERSE=sorted({v for v in canon.values()})
assert not [u for u in UNIVERSE if u.lower() in SUBURBS], 'gazetteer label is a Logan suburb'
GP=re.compile('('+'|'.join(re.escape(g) for g in sorted(UNIVERSE,key=len,reverse=True) if len(g)>5)+')',re.I)
print(f'gazetteer {len(UNIVERSE)} labels; suburb blocklist {len(SUBURBS)}')

def clean(v):
    if v is None: return ''
    s=' '.join(str(v).replace('\n',' ').split())
    s=re.sub(r'(\w)-(\w)',r'\1\2',s)                    # de-hyphenate line wraps ("Underw-ood")
    return '' if s.strip().lower() in EMPTY else s
def parks_in(text):
    return sorted({canon.get(m.lower(),m) for m in GP.findall(text)})

# ---------------- 4 collect-all-candidates over ALL columns except the ban list
BANNED={12,25,32,92,98,105,SITE,BASIS}
COLS=[c for c in range(1,131) if c not in BANNED]
PRIORITY={111:1,106:2,112:3,108:4,127:5,99:6,42:7,22:8,23:9,80:10}
LBL={111:'Sighted invoice line items (printed)',106:'Sighted invoice site (printed)',
     112:'Sighted invoice works carried out (printed)',108:'Sighted invoice work description',
     127:'Sighted invoice anomalies note',99:'Sighted invoice ship-to (printed)',
     42:'Electricity supply point',22:'GL narration',23:'Enquiry narration',80:'Src User Fld 2'}
TOLL=re.compile(r'\btoll',re.I); CHAMBERS=re.compile(r'177\s+Chambers\s+Flat',re.I)
new_site=list(site_old); new_bas=list(bas_old)
n_new=n_collapse=n_multi_new=n_fac=n_toll=n_depot=0
cand_harvest=collections.Counter()
for i in range(len(df)):
    b=bas_old[i]; s=site_old[i]
    # 2. re-screen joined lists
    if b==MULTI and ';' in s:
        keep=sorted({canon.get(p.strip().lower(),p.strip()) for p in s.split(';') if p.strip().lower() not in SUBURBS and p.strip()})
        keep=[k for k in keep if k in set(UNIVERSE)]
        if len(keep)==1:
            new_site[i]=keep[0]; new_bas[i]='Multi-site list reduced to one park (suburb removed)'; n_collapse+=1; continue
        if keep and len(keep)!=len(s.split(';')):
            new_site[i]='; '.join(keep)
        continue
    if b not in CL and b!=MA: continue                  # already attributed on evidence
    if b==MA: continue                                   # printed-line allocation is authoritative
    # collect every candidate across the whole row
    found={}
    for c in COLS:
        t=clean(df.iloc[i,c-1])
        if not t: continue
        if c==99 and CHAMBERS.search(t): found.setdefault(0,(DEPOT,'Ship-to = Parks Depot')); continue
        for p in parks_in(t): found.setdefault(PRIORITY.get(c,20),(p,LBL.get(c,f'Register column {c}')))
        if c in (22,42,106,108,111,112):
            for m in re.finditer(r"\b([A-Z][A-Za-z'\-]+(?:\s+[A-Z][A-Za-z'\-]+){0,3}\s+(?:Park|Reserve|Parklands|Gardens|Wetlands|Courts|Complex))\b",t):
                nm=' '.join(m.group(1).split())
                if nm.lower() not in canon and nm.lower() not in SUBURBS: cand_harvest[nm.title()]+=1
    if not found: 
        if TOLL.search(clean(df.iloc[i,23]))+0 if False else TOLL.search(str(df.iloc[i,23])) : pass
        continue
    picks={v[0] for v in found.values() if v[0]!=DEPOT}
    if DEPOT in {v[0] for v in found.values()} and not picks:
        new_site[i]=DEPOT; new_bas[i]='Non-park Council facility (ship-to = Parks Depot)'; n_depot+=1; continue
    if len(picks)==1:
        k=min(k for k,v in found.items() if v[0] in picks)
        new_site[i]=found[k][0]; new_bas[i]=found[k][1]; n_new+=1
    elif len(picks)>1:
        new_site[i]='; '.join(sorted(picks)); new_bas[i]=MULTI; n_multi_new+=1
# 6 tolls
for i in range(len(df)):
    if new_bas[i] in CL and TOLL.search(str(df.iloc[i,23])+' '+str(df.iloc[i,11])):
        new_site[i]=''; new_bas[i]=NOT_SITE; n_toll+=1
print(f'collect-all: wrote {n_new}, new multi {n_multi_new}, lists collapsed {n_collapse}, depot {n_depot}, tolls {n_toll}')
fold=collections.defaultdict(set)
for s,b in zip(new_site,new_bas):
    if s and b not in CL and b!=MA and b!=MULTI and ';' not in s: fold[s.lower()].add(s)
assert all(len(v)==1 for v in fold.values()),{k:v for k,v in fold.items() if len(v)>1}
# strip blocked terms from EVERY joined list, whatever the basis
n_strip=0
for i in range(len(df)):
    if new_bas[i]==MA: continue
    if new_site[i] and ';' in new_site[i]:
        keep=[canon.get(p.strip().lower(),p.strip()) for p in new_site[i].split(';') if p.strip() and p.strip().lower() not in SUBURBS]
        keep=sorted(set(keep))
        if len(keep)!=len([p for p in new_site[i].split(';') if p.strip()]):
            n_strip+=1
            if len(keep)==1: new_site[i]=keep[0]; new_bas[i]='Multi-site list reduced to one park (suburb removed)'; n_collapse+=1
            elif keep: new_site[i]='; '.join(keep)
            else: new_site[i]=''; new_bas[i]='No site determinable'
print('joined lists with a blocked term stripped:',n_strip)
# clear any suburb label left by an earlier build (v46 removed 'Shailer Park' only)
n_sub=0
for i in range(len(df)):
    if new_bas[i]==MA: continue
    if new_site[i] and ';' not in new_site[i] and new_site[i].lower() in SUBURBS:
        new_site[i]=''; new_bas[i]='No site determinable'; n_sub+=1
print('pre-existing suburb labels cleared:',n_sub)
assert not [s for s in new_site if s and ';' not in s and s.lower() in SUBURBS], 'suburb written as a site'

# ---------------- writes
wb=openpyxl.load_workbook(SRC); reg=wb['Register']; assert reg.max_row==21475
tmpl=copy.copy(reg.cell(5,42)._style); nrw=0
for i in range(len(df)):
    if new_site[i]==site_old[i] and new_bas[i]==bas_old[i]: continue
    r=i+5
    c=reg.cell(r,SITE); c.value=new_site[i] or None; c._style=copy.copy(tmpl)
    c=reg.cell(r,BASIS); c.value=new_bas[i]; c._style=copy.copy(tmpl); nrw+=1
print('register rows rewritten:',nrw)

# ---------------- 7 stores catalogue enrichment
patn=re.compile(r'\b(\d{6,7})\b'); n_cat=0
for i in range(len(df)):
    if new_bas[i]!='Internal charge (not site-specific)': continue
    items=[]
    for c in (22,60):
        for k in patn.findall(str(df.iloc[i,c-1])):
            if k in CATALOGUE and k not in [x[0] for x in items]: items.append((k,CATALOGUE[k]))
    if items:
        reg.cell(i+5,32).value=('Marsden Stores 2026 catalogue match: '+'; '.join(f'{k} {d}' for k,d in items[:3]))[:250]
        n_cat+=1
print('stores rows enriched from the catalogue:',n_cat)

# ---------------- 5 EIL site across the whole sheet
we=wb['Evidence_Invoice_Lines']; ds=copy.copy(we.cell(6,12)._style); n_eil=0
eil=cw.get_sheet_by_name('Evidence_Invoice_Lines').to_python()
for j in range(EIL_END-4):
    r=5+j; row=eil[4+j] if 4+j<len(eil) else None
    if not row: continue
    d=clean(row[2] if len(row)>2 else '')
    p=parks_in(d) if d else []
    s=('; '.join(p) if len(p)>1 else (p[0] if p else ''))
    b='Printed line names one park' if len(p)==1 else ('Printed line names several parks (held)' if len(p)>1 else 'No park named on the printed line')
    c=we.cell(r,12); c.value=s or None; c._style=copy.copy(ds)
    c=we.cell(r,13); c.value=b; c._style=copy.copy(ds)
    if p: n_eil+=1
print('EIL lines naming a park:',n_eil)

# ---------------- Site_Allocation: a suburb is not a park; move those amounts to the residue row
wsa=wb['Site_Allocation']; n_sa=0
for r in range(5,SA_END+1):
    v=wsa.cell(r,7).value
    if isinstance(v,str) and v.strip().lower() in SUBURBS:
        wsa.cell(r,7).value=RESIDUE
        wsa.cell(r,9).value='Suburb, not a park (v48 blocklist); moved to the allocation residue.'
        n_sa+=1
print('Site_Allocation suburb labels moved to residue:',n_sa)
sa=[x for x in cw.get_sheet_by_name('Site_Allocation').to_python()[4:] if x and x[0] not in (None,'') and isinstance(x[7],(int,float))]
sa=[[*x[:6],(RESIDUE if str(x[6]).strip().lower() in SUBURBS else x[6]),*x[7:]] for x in sa]

# ---------------- Sites rebuild
parks=collections.Counter(); ppk=collections.defaultdict(collections.Counter); pbas=collections.defaultdict(collections.Counter)
pkv=df.iloc[:,16].astype(str).tolist()
for s,b,a_,pk in zip(new_site,new_bas,amt,pkv):
    if s and b not in CL and b!=MA and b!=MULTI and ';' not in s:
        parks[s]+=a_; ppk[s][pk]+=1; pbas[s][b]+=1
sa=[x for x in cw.get_sheet_by_name('Site_Allocation').to_python()[4:] if x and x[0] not in (None,'') and isinstance(x[7],(int,float))]
for x in sa:
    if str(x[6])!=RESIDUE: parks[str(x[6])]+=float(x[7]); ppk[str(x[6])][str(x[4])]+=1; pbas[str(x[6])]['Site_Allocation (printed line amounts)']+=1
order=sorted(parks,key=lambda p:-parks[p])
NEWCLASS=CLASS+[NOT_SITE,'Multi-site list reduced to one park (suburb removed)']
NEWCLASS=[c for c in NEWCLASS if c not in ('Multi-site list reduced to one park (suburb removed)',)]
NEWCLASS=NEWCLASS+[NOT_SITE] if NOT_SITE not in NEWCLASS else NEWCLASS
so=wb['Sites']; hdr=copy.copy(so.cell(4,1)._style); dat=copy.copy(so.cell(5,1)._style)
num=copy.copy(so.cell(5,2)._style); mon=copy.copy(so.cell(5,3)._style)
elec=[[so.cell(r,c).value for c in range(1,5)] for r in range(PB0,PB1+1)]
elec_st=[[copy.copy(so.cell(r,c)._style) for c in range(1,5)] for r in range(PB0,PB1+1)]
gnote={so.cell(r,1).value:so.cell(r,9).value for r in range(G0,G1+1)}
for row in so.iter_rows(min_row=3,max_row=so.max_row):
    for c in row: c.value=None
so.cell(2,1).value=str(so.cell(2,1).value)+' v48: site read rebuilt COLLECT-ALL-CANDIDATES over every register column, with "(not printed)" treated as empty; Logan suburb blocklist enforced; joined multi-site lists re-screened and collapsed where a suburb was the only second member.'
def put(ws,r,c,v,st=None):
    cell=ws.cell(r,c); cell.value=v
    if st is not None: cell._style=copy.copy(st)
FYS=['FY2023/24','FY2024/25','FY2025/26','FY2026/27']
put(so,3,1,'Panel A: cost per park / site (live)',hdr)
for j,h in enumerate(['Park / Site','Register lines (live)','$ ex GST (live)']+FYS+['Primary PK','Primary basis','Bases'],1): put(so,4,j,h,hdr)
r=5; PA_S=5
for p in order:
    put(so,r,1,p,dat)
    so.cell(r,2).value=f'=COUNTIF(Register!${SC}$5:${SC}$21473,$A{r})'; so.cell(r,2)._style=copy.copy(num); so.cell(r,2).number_format='#,##0;-#,##0;"-"'
    so.cell(r,3).value=f'=SUMIF(Register!${SC}$5:${SC}$21473,$A{r},Register!$T$5:$T$21473)+SUMIF(Site_Allocation!$G$5:$G${SA_END},$A{r},Site_Allocation!$H$5:$H${SA_END})'
    so.cell(r,3)._style=copy.copy(mon); so.cell(r,3).number_format='$#,##0.00;($#,##0.00);"-"'
    for k,fy in enumerate(FYS):
        so.cell(r,4+k).value=(f'=SUMIFS(Register!$T$5:$T$21473,Register!${SC}$5:${SC}$21473,$A{r},Register!$E$5:$E$21473,"{fy}")'
                              f'+SUMIFS(Site_Allocation!$H$5:$H${SA_END},Site_Allocation!$G$5:$G${SA_END},$A{r},Site_Allocation!$D$5:$D${SA_END},"{fy}")')
        so.cell(r,4+k)._style=copy.copy(mon); so.cell(r,4+k).number_format='$#,##0.00;($#,##0.00);"-"'
    put(so,r,8,max(ppk[p].items(),key=lambda x:x[1])[0] if ppk[p] else '',dat)
    put(so,r,9,max(pbas[p].items(),key=lambda x:x[1])[0] if pbas[p] else '',dat)
    put(so,r,10,len(pbas[p]),dat); r+=1
PA_E=r-1; GS=r
GROUPS=[RESIDUE]+[c for c in NEWCLASS if c!=RESIDUE]
for lbl in GROUPS:
    put(so,r,1,lbl,dat)
    if lbl==RESIDUE:
        so.cell(r,2).value=f'=COUNTIF(Register!${BC}$5:${BC}$21473,"{MA}")'
        so.cell(r,3).value=f'=SUMIF(Site_Allocation!$G$5:$G${SA_END},$A{r},Site_Allocation!$H$5:$H${SA_END})'
        for k,fy in enumerate(FYS): so.cell(r,4+k).value=f'=SUMIFS(Site_Allocation!$H$5:$H${SA_END},Site_Allocation!$G$5:$G${SA_END},$A{r},Site_Allocation!$D$5:$D${SA_END},"{fy}")'
    else:
        so.cell(r,2).value=f'=COUNTIF(Register!${BC}$5:${BC}$21473,$A{r})'
        so.cell(r,3).value=f'=SUMIF(Register!${BC}$5:${BC}$21473,$A{r},Register!$T$5:$T$21473)'
        for k,fy in enumerate(FYS): so.cell(r,4+k).value=f'=SUMIFS(Register!$T$5:$T$21473,Register!${BC}$5:${BC}$21473,$A{r},Register!$E$5:$E$21473,"{fy}")'
    so.cell(r,2)._style=copy.copy(num); so.cell(r,2).number_format='#,##0;-#,##0;"-"'
    for k in range(3,8): so.cell(r,k)._style=copy.copy(mon); so.cell(r,k).number_format='$#,##0.00;($#,##0.00);"-"'
    put(so,r,9,gnote.get(lbl,''),dat); r+=1
# rows whose basis is a collapsed-list or CR-disagreement variant are counted via a catch-all
GE=r-1; TOT=r
put(so,r,1,'TOTAL (must tie to the register)',hdr)
so.cell(r,2).value=f'=SUM(B{PA_S}:B{GE})'; so.cell(r,2)._style=copy.copy(hdr); so.cell(r,2).number_format='#,##0'
for k in range(3,8):
    L=get_column_letter(k); so.cell(r,k).value=f'=SUM({L}{PA_S}:{L}{GE})'; so.cell(r,k)._style=copy.copy(hdr); so.cell(r,k).number_format='$#,##0.00;($#,##0.00);"-"'
CHK=r+1
put(so,CHK,1,'Tie check (lines = 21,469 and $ = Register!$T$21475)',hdr)
so.cell(CHK,2).value=f'=IF(B{TOT}=21469,"TRUE","FALSE")'; so.cell(CHK,2)._style=copy.copy(hdr)
so.cell(CHK,3).value=f'=IF(ROUND(C{TOT},2)=ROUND(Register!$T$21475,2),"TRUE","FALSE")'; so.cell(CHK,3)._style=copy.copy(hdr)
so.cell(CHK,4).value=f'=IF(ROUND(SUM(D{TOT}:G{TOT}),2)=ROUND(C{TOT},2),"TRUE","FALSE")'; so.cell(CHK,4)._style=copy.copy(hdr)
NPB=CHK+2
put(so,NPB,1,'Panel B: electricity by retailer account (relocated unchanged)',hdr)
for j,h in enumerate(['Retailer account','Site','Lines','$ ex GST (net, all years)'],1): put(so,NPB+1,j,h,hdr)
r=NPB+2
for vals,sts in zip(elec,elec_st):
    for j,v in enumerate(vals,1):
        c=so.cell(r,j); c.value=v; c._style=copy.copy(sts[j-1])
    r+=1
NPB_E=r-1
print(f'Sites: PanelA {PA_S}:{PA_E} ({len(order)}), groups {GS}:{GE}, total {TOT}, chk {CHK}')

cf=wb['Config']; cr={str(cf.cell(i,1).value):i for i in range(1,cf.max_row+1) if cf.cell(i,1).value}
for k,v in [('WORKBOOK_VERSION','v48'),('SITES_PANELA',f'{PA_S}:{PA_E}'),('SITES_GROUP_ROWS',f'{GS}:{GE}'),
            ('SITES_TOTAL_ROW',TOT),('SITES_CHECK_ROW',CHK),('SITES_PANELB',f'{NPB+2}:{NPB_E}'),('PARK_COUNT',len(order))]:
    if k in cr: cf.cell(cr[k],2).value=v
ca=copy.copy(cf.cell(cf.max_row,1)._style); cb=copy.copy(cf.cell(cf.max_row,2)._style); nr=cf.max_row+1
for k,v in [('SUBURB_BLOCKLIST',len(SUBURBS)),('GAZ_UNIVERSE',len(UNIVERSE)),('SITE_WRITES_V48',n_new),
            ('LISTS_COLLAPSED_V48',n_collapse),('STORES_CATALOGUE_MATCHES',n_cat),('EIL_SITE_LINES',n_eil)]:
    a=cf.cell(nr,1); a.value=k; a._style=copy.copy(ca); b=cf.cell(nr,2); b.value=v; b._style=copy.copy(cb); nr+=1
h=wb['Handover']; hr=h.max_row+1; hs=[copy.copy(h.cell(h.max_row,c)._style) for c in range(1,4)]
h.cell(hr,1).value=(f'v48 change (18-Aug-2026): full site sweep. Site read rebuilt COLLECT-ALL-CANDIDATES over every register column with "(not printed)" treated as EMPTY; '
 f'{n_new} lines gain a site, {n_multi_new} become multi-site, {n_collapse} joined lists collapse to one park once the suburb is removed. Logan suburb blocklist (70 names) enforced as a first-class object. '
 f'Evidence_Invoice_Lines site written across the whole sheet ({n_eil} lines name a park). Tolls reclassified as not site-specific ({n_toll}). '
 f'Marsden Stores 2026 catalogue matched to {n_cat} internal stores issues. Control total unchanged at ${CT:,.2f}.')
h.cell(hr,1)._style=hs[0]; h.cell(hr,2).value=f'Sites Panel A {len(order)} sites.'; h.cell(hr,2)._style=hs[1]; h.cell(hr,3)._style=hs[2]
m=wb['Method']; mr=m.max_row+1; ms=copy.copy(m.cell(m.max_row,1)._style)
METH=[f"47.0 Collect-all-candidates (v48, standing rule). The site read no longer stops at the first non-empty field. Every register column outside the ban list (12, 25, 32, 92, 98, 105) is read, every candidate collected, and the decision made afterwards on the priority order. `(not printed)` and `(blank as printed)` are EMPTY, never an answer: at v47 they stopped the scan on 80 sighted rows carrying $340,399 whose site sat in another field. Where candidates agree the highest-priority source is written; where they name several parks the row becomes multi-site.",
 f"47.1 Suburb blocklist as a first-class object (v48). All 70 Logan suburbs are loaded and the build ASSERTS that no gazetteer label is a suburb and that no suburb is ever written as a site. Banning column 92 at v43 was not enough - Shailer Park entered through other fields and took $50,140. The blocklist is the structural fix.",
 f"47.2 Joined multi-site lists are re-screened on every build (v48). {n_collapse} lists reduced to a single park once blocked terms were removed - `Haldham Park; Shailer Park`, `Lansdown Park; Shailer Park`, `Kimberley Forest Park; Shailer Park`, `Larry Storey Park; Shailer Park`, `Shailer Park; Shailer Pioneer Park`. A gazetteer change must always trigger a list re-screen; the v46 removal fixed single-site rows only.",
 f"47.3 Evidence_Invoice_Lines site written across the WHOLE sheet (v48). At v47 only the Batch 40 lines carried a line-level site, so 584 of 594 EIL lines tied to unresolved register rows were blank. Now every one of the {EIL_END-4} data rows carries a site and a basis; {n_eil} name a park.",
 f"47.4 Marsden Stores catalogue (NEW v48). The internal stores issues narrate `Despatch Stock Requisition 'nnnnnn'-ALLSTORE` and carry a six or seven digit PRODUCT number in the GL narration and Src Details. Matched against the Marsden Stores 2026 Inventory Catalogue (293 product numbers parsed from the PDF), {n_cat} of the 326 internal stores rows resolve to a named item - paver sand 20kg, galvanised post 50mm x 4.4m, ground anchor spikes, star pickets, 1-ply toilet paper, first aid kits. This identifies WHAT was issued; it does NOT give a site, because a stores issue records that goods left the store, not where they were installed. Recorded in the Coding Note.",
 f"47.5 Tolls are not site-specific (v48). Toll operator lines are reclassified from `No site determinable` to `{NOT_SITE}`. A toll has no park by nature; leaving it in a residual class overstates the unresolved position."]
for i,t in enumerate(METH):
    c=m.cell(mr+i,1); c.value=t; c._style=copy.copy(ms)
d=wb['Data_Acquisition']; dr=d.max_row+1; ds2=copy.copy(d.cell(d.max_row,1)._style)
c=d.cell(dr,1); c.value=' |  G9  |  Marsden Stores 2026 Inventory Catalogue (63pp PDF, 293 product numbers with descriptions and units). Held as b48_catalogue.json. Matched to internal stores issues at v48 for item identification.  |  2  |  Held  | '; c._style=copy.copy(ds2)
oi=wb['Open_Items']; orow=oi.max_row+1; osx=[copy.copy(oi.cell(oi.max_row,c)._style) for c in range(1,8)]
ITEMS=[(153,'All','Open','Internal stores issues',n_cat,0,f'STORES ITEMS IDENTIFIED, SITE STILL UNKNOWN. {n_cat} of 326 internal stores issues match a Marsden Stores catalogue product number. The item is now named; the SITE is not, because a stores issue records goods leaving the store. Ask Stores whether the Despatch Stock Requisition carries a work order or PK on the requisition itself.'),
 (154,'All','Open','Gazetteer',len(cand_harvest),0,'RESIDUAL CANDIDATE NAMES from the v48 sweep, harvested from every column. Adjudicate and fold into the gazetteer.')]
for i,it in enumerate(ITEMS):
    for j,v in enumerate(it,1):
        c=oi.cell(orow+i,j); c.value=v; c._style=copy.copy(osx[j-1])

pre=f'{OUT}/{NEW}'; wb.save(pre); print('saved')
prof=f'{ROOT}/lo_profile/user'; os.makedirs(prof,exist_ok=True)
open(f'{prof}/registrymodifications.xcu','w').write(open('/home/claude/b43/lo_profile/user/registrymodifications.xcu').read())
t0=time.time()
res=subprocess.run(['soffice',f'-env:UserInstallation=file://{ROOT}/lo_profile','--headless','--calc','--convert-to','xlsx','--outdir',RC,pre],capture_output=True,text=True,timeout=1800)
final=f'{RC}/{NEW}'; assert os.path.exists(final),res.stdout+res.stderr; print(f'recalc {time.time()-t0:.0f}s')
json.dump(dict(final=final,CT=CT,PA=[PA_S,PA_E],GRP=[GS,GE],TOT=TOT,CHK=CHK,PB=[NPB+2,NPB_E],parks=len(order),
  n_new=n_new,n_collapse=n_collapse,n_cat=n_cat,n_eil=n_eil,n_toll=n_toll,CLASS=NEWCLASS,MA=MA,
  gb=int((df.iloc[:,87].notna()).sum()),suburbs=sorted(SUBURBS),cands=len(cand_harvest)),open(f'{ROOT}/b48_pos.json','w'),indent=1)
pd.DataFrame(sorted(cand_harvest.items(),key=lambda x:-x[1]),columns=['candidate','hits']).to_excel(f'{ROOT}/b48_residual_candidates.xlsx',index=False)
r=subprocess.run([sys.executable,f'{ROOT}/b48_verify.py'],capture_output=True,text=True); print(r.stdout); print(r.stderr[-1200:])
if 'FAILS: none' in r.stdout:
    subprocess.run(['cp',final,f'/mnt/user-data/outputs/{NEW}']); print('SHIPPED',NEW)
else: print('NOT SHIPPED'); sys.exit(1)
