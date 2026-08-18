import json, openpyxl, subprocess, os
from openpyxl.styles import Font, PatternFill, Alignment
E=json.load(open('/home/claude/b51/b51_elec.json')); wb=openpyxl.Workbook()
F=Font(name='Cambria',size=10); H=Font(name='Cambria',size=10,bold=True,color='FFFFFF'); FILL=PatternFill('solid',fgColor='1F4E78')
MON='$#,##0.00;($#,##0.00);"-"'
def sheet(ws,hdr,rows,widths):
    ws.append(hdr)
    for c in ws[1]: c.font=H; c.fill=FILL; c.alignment=Alignment(wrap_text=True,vertical='top')
    for r in rows: ws.append(r)
    for row in ws.iter_rows(min_row=2):
        for c in row: c.font=F
    for i,w in enumerate(widths,1): ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
    ws.freeze_panes='A2'
ws=wb.active; ws.title='Statements'
rows=[[k,v['issue'],v['accounts'],v['total_incl'],v['total_gst'],round(v['total_incl']-v['total_gst'],2),v['parks_groups'],v['parks_ex'],v['md5'],v['file']] for k,v in E['stmt'].items()]
sheet(ws,['Statement','Issue date','Sub-accounts on statement','Current charges incl GST','GST','Ex GST','Sub-accounts on Parks register (tied)','Parks register $ ex GST','md5','File'],rows,[18,12,14,18,14,16,18,18,34,34])
n=len(rows)+1; ws.cell(n+1,1,'TOTAL').font=H
for c,L in ((3,'C'),(4,'D'),(5,'E'),(6,'F'),(7,'G'),(8,'H')): ws.cell(n+1,c,f'=SUM({L}2:{L}{n})').font=H
for r in ws.iter_rows(min_row=2,min_col=4,max_col=8):
    for c in r: c.number_format=MON
for r in ws.iter_rows(min_row=2,min_col=3,max_col=3): r[0].number_format='#,##0'
for r in ws.iter_rows(min_row=2,min_col=7,max_col=7): r[0].number_format='#,##0'
ws=wb.create_sheet('Recon')
rows=[[x['stmt'],x['nmi'],x['sub'],x['inv'],x['stmt_addr'],x['reg_addr'],x['stmt_incl'],round(x['stmt_incl']-x['stmt_ex'],2) if x['stmt_ex'] is not None else None,x['stmt_ex'],x['reg_ex'],x['lines'],x['pk'],x['na'],x['dy'],x['dz'],x['rows'] and ','.join(str(i) for i in x['rows'])] for x in E['recon']]
sheet(ws,['Statement','NMI (register col AO)','Statement sub-account','Statement invoice number','Supply address (statement, full)','Supply point (register, 40-char label)','Statement charges incl GST','Statement GST','Statement ex GST','Register $ ex GST','Register lines','PK charged','Natural account','Register site (DY)','Register site basis (DZ)','Register rows','Tie (register ex = statement ex, +/- 2c)'],rows,[16,14,14,14,44,40,14,10,14,14,8,12,10,26,34,30,12])
n=len(rows)+1
for r in range(2,n+1): ws.cell(r,17,f'=IF(ROUND(ABS(J{r}-I{r}),2)<=0.02,"TRUE","FALSE")').font=F
for r in ws.iter_rows(min_row=2,max_row=n,min_col=7,max_col=10):
    for c in r: c.number_format=MON
ws.cell(n+2,1,'TOTAL').font=H
for c,L in ((7,'G'),(8,'H'),(9,'I'),(10,'J'),(11,'K')): ws.cell(n+2,c,f'=SUM({L}2:{L}{n})').font=H
ws.cell(n+3,1,'Groups tied').font=H; ws.cell(n+3,2,f'=COUNTIF(Q2:Q{n},"TRUE")').font=H; ws.cell(n+3,3,f'=IF(B{n+3}={n-1},"TRUE","FALSE")').font=H
ws.cell(n+4,1,'Register total on the six statements ties Recon J').font=H; ws.cell(n+4,2,round(sum(x['reg_ex'] for x in E['recon']),2)).font=H; ws.cell(n+4,3,f'=IF(ROUND(B{n+4},2)=ROUND(J{n+2},2),"TRUE","FALSE")').font=H
ws=wb.create_sheet('Sub_accounts')
onreg={(x['stmt'],x['sub']) for x in E['recon'] if x['tie']=='TRUE'}
rows=[[k,r['sub'],r['inv'],r['addr'],r['incl'],r['gst'],r['ex'],'Y' if (k,r['sub']) in onreg else 'N'] for k,v in E['stmt'].items() for r in v['rows']]
sheet(ws,['Statement','Sub-account','Invoice number','Supply address','Charges incl GST','GST','Ex GST','On the PS/WP register'],rows,[16,14,14,54,14,10,14,12])
n=len(rows)+1; ws.cell(n+2,1,'TOTAL').font=H
for c,L in ((5,'E'),(6,'F'),(7,'G')): ws.cell(n+2,c,f'=SUM({L}2:{L}{n})').font=H
for r in ws.iter_rows(min_row=2,min_col=5,max_col=7):
    for c in r: c.number_format=MON
ws=wb.create_sheet('Method'); ws.column_dimensions['A'].width=140
for t in ['Batch 43 (18-Aug-2026, v51): Origin collective account summaries A-FE35E476-022 to -027 (issues 1-Sep-2025 to 4-Feb-2026), parsed from the CURRENT CHARGES - SUB ACCOUNT SUMMARY of each PDF (pdftotext -layout). Per statement the row count is asserted against the printed "Number of accounts" and the row sums against the printed "Total current charges" (incl and GST) and the account-summary "Current charges" line.',
 'Register side: every FY2025/26 line whose Reference is one of the six statements, grouped by (statement, NMI in column AO). Match: the register supply-point label is the Origin supply address cut at 40 characters, so the register label (minus a truncated tail token) must be a prefix of the normalised statement address; then register ex GST must equal statement (charges incl GST - GST) within 2c. 375 of 375 groups tie.',
 'The statement supply addresses add no park name the truncated register labels did not already carry, so no site write follows from this batch. Sub-accounts marked N are Council-wide accounts outside the Parks register (sewerage pumps, offices, cemeteries, zero-charge accounts).',
 'Amounts: statement figures incl GST as printed; ex GST = incl - GST (statement); register ex GST from column T. All figures $ AUD.']:
    ws.append([t]); ws.cell(ws.max_row,1).font=F; ws.cell(ws.max_row,1).alignment=Alignment(wrap_text=True,vertical='top')
p='/home/claude/b51/out/PSWP_Batch43_Electricity_Reconciliation_022_027.xlsx'; wb.save(p)
res=subprocess.run(['soffice','-env:UserInstallation=file:///home/claude/b51/lo_profile','--headless','--calc','--convert-to','xlsx','--outdir','/home/claude/b51/out/recalc',p],capture_output=True,text=True,timeout=600)
from python_calamine import CalamineWorkbook
w=CalamineWorkbook.from_path('/home/claude/b51/out/recalc/PSWP_Batch43_Electricity_Reconciliation_022_027.xlsx')
rc=w.get_sheet_by_name('Recon').to_python(); print(rc[-2][:3], rc[-1][:3]); print(w.get_sheet_by_name('Statements').to_python()[7][:8])
err=[c for s in w.sheet_names for row in w.get_sheet_by_name(s).to_python() for c in row if isinstance(c,str) and c.startswith('#')]; print('errors',err[:3])
