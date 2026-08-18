"""b46_build.py  v45 -> v46 (18-Aug-2026)
IDENTIFICATION build + site corrections. Column T never written; no green block touched; control total unchanged.
(1) Five APLEDGER creditor histories (ELE032, QPO001, SEC022, SEA029, HIB003) appended to Creditor_Lines;
    485 register rows identified where the creditor transaction amount (INCL GST) ties the register ex GST at 1.1
    within 2c (v35 amount-corroboration rule). Confirmed rows skipped. 20 references held on the amount test.
(2) Site corrections per user directive: 'Shailer Park' removed from the gazetteer (it is a SUBURB and David
    Robinson Landscaping's business address); three manual overrides; a new Site_Crosswalk sheet that OUTRANKS
    the matcher permanently; non-park Council facilities recognised as sites.
"""
import json, copy, re, os, sys, subprocess, time, collections
import openpyxl, pandas as pd  # noqa
from openpyxl.utils import get_column_letter
from python_calamine import CalamineWorkbook

ROOT='/home/claude/b46'; SRC='/mnt/user-data/outputs/PS_WP_Transaction_Register_3FY_v45_HANDOVER.xlsx'
NEW='PS_WP_Transaction_Register_3FY_v46_HANDOVER.xlsx'
OUT=f'{ROOT}/out'; RC=f'{OUT}/recalc'; os.makedirs(RC,exist_ok=True)
CT=11377312.51; SITE_COL=129; BASIS_COL=130; SC='DY'; BC='DZ'
P=json.load(open(f'{ROOT}/b46_parse.json'))
TGT={t['row']:t for t in P['targets']}
cw=CalamineWorkbook.from_path(SRC)
cfg={str(r[0]):r[1] for r in cw.get_sheet_by_name('Config').to_python() if r and r[0]}
PA0,PA1=[int(x) for x in str(cfg['SITES_PANELA']).split(':')]
G0,G1=[int(x) for x in str(cfg['SITES_GROUP_ROWS']).split(':')]
TOT_ROW=int(float(cfg['SITES_TOTAL_ROW'])); CHK_ROW=int(float(cfg['SITES_CHECK_ROW']))
PB0,PB1=[int(x) for x in str(cfg['SITES_PANELB']).split(':')]
SA_END=int(str(cfg['SITE_ALLOC_DATA']).split(':')[1])
sites=cw.get_sheet_by_name('Sites').to_python()
CLASS=[str(r[0]) for r in sites[G0-1:G1] if r and r[0]]
CL=set(CLASS); MULTI='Multiple sites named (not allocated)'
_bas=set(pd.read_excel(SRC,sheet_name='Register',header=3,engine='calamine').iloc[:,BASIS_COL-1].fillna('').astype(str))
MA=next(c for c in _bas if 'allocated per printed' in c)   # basis label, NOT a Sites group row (residue stands for it)
RESIDUE=next(c for c in CLASS if c.startswith('(residue'))
FAC_CLASS='Non-park Council facility (ship-to = Parks Depot)'
NEWFAC='Non-park Council facility (named site, not a park)'
XW_BASIS='Manual site crosswalk (user-adjudicated, outranks the matcher)'

df=pd.read_excel(SRC,sheet_name='Register',header=3,engine='calamine'); df=df[df.iloc[:,0].notna()]
amt=pd.to_numeric(df.iloc[:,19],errors='coerce').fillna(0); assert abs(amt.sum()-CT)<0.005
site_old=df.iloc[:,SITE_COL-1].fillna('').astype(str).tolist()
bas_old=df.iloc[:,BASIS_COL-1].fillna('').astype(str).tolist()
narr=df.iloc[:,21].fillna('').astype(str).str.replace('\n',' ',regex=False).tolist()

wb=openpyxl.load_workbook(SRC); reg=wb['Register']; assert reg.max_row==21475
tmpl=copy.copy(reg.cell(5,42)._style)

# ---------------------------------------------------------------- 1. identification writes
n_id=0
for row,t in sorted(TGT.items()):
    r=row
    reg.cell(r,11).value=t['acct']                                   # creditor code
    reg.cell(r,12).value=t['label']                                  # contractor
    reg.cell(r,13).value=t['abn']
    reg.cell(r,28).value=(f"APLEDGER creditor history {t['acct']} pulled 18-Aug-2026: reference {t['ref']} present, "
                          f"transaction amount ${t['incl']:,.2f} incl GST = register ${t['amount']:,.2f} ex GST at 1.1 (amount corroborated).")
    reg.cell(r,29).value='Tier 1'
    reg.cell(r,34).value='Supplier identified from creditor history; sight the invoice to confirm nature and coding.'
    n_id+=1
print('identification rows written:',n_id)

# ---------------------------------------------------------------- 2. site corrections
REMOVE=set(P['gaz_remove']); OVR=P['override']; FAC=set(P['facility'])
new_site=list(site_old); new_bas=list(bas_old)
n_rm=n_ovr=n_fac=0
for i in range(len(df)):
    s=site_old[i]
    if s in REMOVE:
        new_site[i]=''; new_bas[i]='No site determinable'; n_rm+=1
for i in range(len(df)):
    t=narr[i]
    for key,park in OVR.items():
        if key.lower() in t.lower():
            new_site[i]=park
            new_bas[i]=XW_BASIS if park not in FAC else NEWFAC
            n_ovr+=1; break
# named non-park facilities present in the narration and currently unattributed
FACPAT={'West Logan Netball Courts':re.compile(r'west\s+logan\s+netball\s+courts',re.I),
        'Logan and District Services Club':re.compile(r'logan\s+and\s+district\s+services',re.I),
        'Historical Society Amenities (223 Main St)':re.compile(r'historical\s+society\s+amenities',re.I)}
for i in range(len(df)):
    if new_bas[i] not in CL: continue
    for lbl,pat in FACPAT.items():
        if pat.search(narr[i]):
            new_site[i]=lbl; new_bas[i]=NEWFAC; n_fac+=1; break
print(f'site: removed {n_rm}, overrides {n_ovr}, facilities {n_fac}')
n_w=0
for i in range(len(df)):
    if new_site[i]==site_old[i] and new_bas[i]==bas_old[i]: continue
    r=i+5
    c=reg.cell(r,SITE_COL); c.value=new_site[i] or None; c._style=copy.copy(tmpl)
    c=reg.cell(r,BASIS_COL); c.value=new_bas[i]; c._style=copy.copy(tmpl); n_w+=1
print('site rows rewritten:',n_w)
fold=collections.defaultdict(set)
for s,b in zip(new_site,new_bas):
    if s and b not in CL and b!=MA and b!=NEWFAC: fold[s.lower()].add(s)
assert all(len(v)==1 for v in fold.values()),{k:v for k,v in fold.items() if len(v)>1}

# ---------------------------------------------------------------- 3. Creditor_Lines append
cl=wb['Creditor_Lines']
last=max(i for i in range(5,cl.max_row+1) if cl.cell(i,1).value not in (None,''))
cst=[copy.copy(cl.cell(last,c)._style) for c in range(1,15)]
rr=last+1
for a,label,abn,ref,date,dtype,det,amt_,due,age,md5 in P['cred']:
    vals=[f'{a} ({label})','(history pull)',ref,date,date,'N','N',date,dtype,det,0,amt_,amt_,due]
    for c,v in enumerate(vals,1):
        cell=cl.cell(rr,c); cell.value=v; cell._style=copy.copy(cst[c-1])
    rr+=1
CL_END=rr-1
print('Creditor_Lines ->',CL_END)

# ---------------------------------------------------------------- 4. Site_Crosswalk (permanent, outranks the matcher)
xw=wb.create_sheet('Site_Crosswalk', wb.sheetnames.index('Sites')+1)
xw.sheet_view.showGridLines=False
so=wb['Sites']; ttl=copy.copy(so.cell(1,1)._style); sub=copy.copy(so.cell(2,1)._style)
hdr=copy.copy(so.cell(4,1)._style); dat=copy.copy(so.cell(5,1)._style)
def put(ws,r,c,v,st=None):
    cell=ws.cell(r,c); cell.value=v
    if st is not None: cell._style=copy.copy(st)
    return cell
put(xw,1,1,'Site crosswalk: manual, user-adjudicated site decisions (v46)',ttl)
put(xw,2,1,'These decisions OUTRANK the automatic matcher and survive every future build. A published Parks Directory address is a starting point, not an authority: the directory lists Ewing Park at Netball Drive and Rainbow Park at Ewing Road, so a street match sent all 237 lines at 99 Ewing Rd to the wrong park. Street-only matching is banned outright (it sent 293 Logan St Eagleby to Doug Larsen Park, which is 41 Logan Street BEENLEIGH). Non-park Council facilities are valid sites and are recorded as such, not forced onto a park.',sub)
for j,h in enumerate(['Source text (narration / field)','Decided site','Type','Basis of the decision','Decided by','Date'],1): put(xw,4,j,h,hdr)
for j,w in enumerate([46,34,22,66,18,13],1): xw.column_dimensions[get_column_letter(j)].width=w
rows=[('99 Ewing Rd Woodridge','Ewing Park','Park','Directory lists Ewing Park at Netball Drive and Rainbow Park at Ewing Road; local knowledge confirms the supply point is Ewing Park.','User','18-Aug-2026'),
      ('293 Logan St Eagleby (Little Athletic Club)','Olivers Sports Complex','Non-park Council facility','Directory lists Olivers Sports Complex at Logan Street, Eagleby. The matcher returned Doug Larsen Park (41 Logan Street, BEENLEIGH) - a cross-suburb error.','User','18-Aug-2026'),
      ('71 Flagstonian Dr (Pathway Lighting)','Flagstone Regional Park','Park','Directory lists Flagstone Regional Park at Flagstonian Drive, Flagstone.','User','18-Aug-2026'),
      ('Shailer Park (as a site label)','(removed)','Suburb, not a park','Shailer Park is a SUBURB and David Robinson Landscaping\u2019s business address (15 Helios St, Shailer Park, per PO 802122). It had taken $50,140 across 16 lines and outranked Homestead Park (Shailer Park) on an exact match.','Claude/User','18-Aug-2026'),
      ('West Logan Netball Courts, 155 Middle Rd','West Logan Netball Courts','Non-park Council facility','Named Council sporting facility appearing under three spellings in the electricity narration; a real site, not a park.','Claude','18-Aug-2026'),
      ('Historical Society Amenities, 223 Main St','Historical Society Amenities (223 Main St)','Non-park Council facility','Named facility on the electricity narration.','Claude','18-Aug-2026'),
      ('Logan and District Services Club, 5 Croydon','Logan and District Services Club','Non-park Council facility','Named facility on the electricity narration.','Claude','18-Aug-2026')]
r=5
for v in rows:
    for j,x in enumerate(v,1): put(xw,r,j,x,dat)
    r+=1
XW_END=r-1

# ---------------------------------------------------------------- 5. Sites Panel A rebuild
parks=collections.Counter(); plines=collections.Counter(); ppk=collections.defaultdict(collections.Counter); pbas=collections.defaultdict(collections.Counter)
pkv=df.iloc[:,16].astype(str).tolist()
for s,b,a_,pk in zip(new_site,new_bas,amt,pkv):
    if s and b not in CL and b!=MA: parks[s]+=a_; plines[s]+=1; ppk[s][pk]+=1; pbas[s][b]+=1
sa=[x for x in cw.get_sheet_by_name('Site_Allocation').to_python()[4:] if x and x[0] not in (None,'') and isinstance(x[7],(int,float))]
for x in sa:
    if str(x[6])!=RESIDUE: parks[str(x[6])]+=float(x[7]); ppk[str(x[6])][str(x[4])]+=1; pbas[str(x[6])]['Site_Allocation (printed line amounts)']+=1
order=sorted(parks,key=lambda p:-parks[p])
elec=[[so.cell(r,c).value for c in range(1,5)] for r in range(PB0,PB1+1)]
elec_st=[[copy.copy(so.cell(r,c)._style) for c in range(1,5)] for r in range(PB0,PB1+1)]
grp_note={so.cell(r,1).value:so.cell(r,9).value for r in range(G0,G1+1)}
num=copy.copy(so.cell(5,2)._style); mon=copy.copy(so.cell(5,3)._style)
for row in so.iter_rows(min_row=3,max_row=so.max_row):
    for c in row: c.value=None
so.cell(2,1).value=str(so.cell(2,1).value)+' v46: Site_Crosswalk decisions outrank the matcher; "Shailer Park" removed (suburb / vendor address); named non-park Council facilities recognised as sites in their own right.'
put(so,3,1,'Panel A: cost per park / site (live; register single-site + Site_Allocation + crosswalk + facilities)',hdr)
FYS=['FY2023/24','FY2024/25','FY2025/26','FY2026/27']
for j,h in enumerate(['Park / Site','Register lines (live)','$ ex GST (live)']+FYS+['Primary PK','Primary basis','Bases'],1): put(so,4,j,h,hdr)
r=5; PA_START=5
for p in order:
    put(so,r,1,p,dat)
    so.cell(r,2).value=f'=COUNTIF(Register!${SC}$5:${SC}$21473,$A{r})'; so.cell(r,2)._style=copy.copy(num); so.cell(r,2).number_format='#,##0;-#,##0;"-"'
    so.cell(r,3).value=f'=SUMIF(Register!${SC}$5:${SC}$21473,$A{r},Register!$T$5:$T$21473)+SUMIF(Site_Allocation!$G$5:$G${SA_END},$A{r},Site_Allocation!$H$5:$H${SA_END})'
    so.cell(r,3)._style=copy.copy(mon); so.cell(r,3).number_format='$#,##0.00;($#,##0.00);"-"'
    for k,fy in enumerate(FYS):
        so.cell(r,4+k).value=(f'=SUMIFS(Register!$T$5:$T$21473,Register!${SC}$5:${SC}$21473,$A{r},Register!$E$5:$E$21473,"{fy}")'
                              f'+SUMIFS(Site_Allocation!$H$5:$H${SA_END},Site_Allocation!$G$5:$G${SA_END},$A{r},Site_Allocation!$D$5:$D${SA_END},"{fy}")')
        so.cell(r,4+k)._style=copy.copy(mon); so.cell(r,4+k).number_format='$#,##0.00;($#,##0.00);"-"'
    put(so,r,8,max(ppk[p].items(),key=lambda x:x[1])[0] if ppk[p] else '',dat)
    put(so,r,9,max(pbas[p].items(),key=lambda x:x[1])[0] if pbas[p] else '',dat)
    put(so,r,10,len(pbas[p]),dat)
    r+=1
PA_END=r-1; GRP_START=r
GROUPS=[RESIDUE]+[c for c in CLASS if c!=RESIDUE]
for lbl in GROUPS:
    put(so,r,1,lbl,dat)
    if lbl==RESIDUE:
        so.cell(r,2).value=f'=COUNTIF(Register!${BC}$5:${BC}$21473,"{MA}")'
        so.cell(r,3).value=f'=SUMIF(Site_Allocation!$G$5:$G${SA_END},$A{r},Site_Allocation!$H$5:$H${SA_END})'
        for k,fy in enumerate(FYS): so.cell(r,4+k).value=f'=SUMIFS(Site_Allocation!$H$5:$H${SA_END},Site_Allocation!$G$5:$G${SA_END},$A{r},Site_Allocation!$D$5:$D${SA_END},"{fy}")'
    else:
        so.cell(r,2).value=f'=COUNTIF(Register!${BC}$5:${BC}$21473,$A{r})'
        so.cell(r,3).value=f'=SUMIF(Register!${BC}$5:${BC}$21473,$A{r},Register!$T$5:$T$21473)'
        for k,fy in enumerate(FYS): so.cell(r,4+k).value=f'=SUMIFS(Register!$T$5:$T$21473,Register!${BC}$5:${BC}$21473,$A{r},Register!$E$5:$E$21473,"{fy}")'
    so.cell(r,2)._style=copy.copy(num); so.cell(r,2).number_format='#,##0;-#,##0;"-"'
    for k in range(3,8): so.cell(r,k)._style=copy.copy(mon); so.cell(r,k).number_format='$#,##0.00;($#,##0.00);"-"'
    put(so,r,9,grp_note.get(lbl,''),dat); r+=1
GRP_END=r-1; NTOT=r
put(so,r,1,'TOTAL (must tie to the register)',hdr)
so.cell(r,2).value=f'=SUM(B{PA_START}:B{GRP_END})'; so.cell(r,2)._style=copy.copy(hdr); so.cell(r,2).number_format='#,##0'
for k in range(3,8):
    L=get_column_letter(k); so.cell(r,k).value=f'=SUM({L}{PA_START}:{L}{GRP_END})'; so.cell(r,k)._style=copy.copy(hdr); so.cell(r,k).number_format='$#,##0.00;($#,##0.00);"-"'
NCHK=r+1
put(so,NCHK,1,'Tie check (lines = 21,469 and $ = Register!$T$21475)',hdr)
so.cell(NCHK,2).value=f'=IF(B{NTOT}=21469,"TRUE","FALSE")'; so.cell(NCHK,2)._style=copy.copy(hdr)
so.cell(NCHK,3).value=f'=IF(ROUND(C{NTOT},2)=ROUND(Register!$T$21475,2),"TRUE","FALSE")'; so.cell(NCHK,3)._style=copy.copy(hdr)
so.cell(NCHK,4).value=f'=IF(ROUND(SUM(D{NTOT}:G{NTOT}),2)=ROUND(C{NTOT},2),"TRUE","FALSE")'; so.cell(NCHK,4)._style=copy.copy(hdr)
NPB=NCHK+2
put(so,NPB,1,'Panel B: electricity by retailer account, all loaded years (relocated unchanged)',hdr)
for j,h in enumerate(['Retailer account','Site','Lines','$ ex GST (net, all years)'],1): put(so,NPB+1,j,h,hdr)
r=NPB+2
for vals,sts in zip(elec,elec_st):
    for j,v in enumerate(vals,1):
        c=so.cell(r,j); c.value=v; c._style=copy.copy(sts[j-1])
    r+=1
NPB_END=r-1
print(f'Sites: PanelA {PA_START}:{PA_END} ({len(order)}), groups {GRP_START}:{GRP_END}, total {NTOT}, chk {NCHK}, PB {NPB+2}:{NPB_END}')

# ---------------------------------------------------------------- 6. Vendor_Series / Config / narrative
vs=wb['Vendor_Series']; vrow={str(vs.cell(i,1).value):i for i in range(5,vs.max_row+1) if vs.cell(i,1).value}
for lbl,abn in [('CINC EDSS Pty Ltd','17 099 805 277'),('Electrical Data & Security Services Pty Ltd','17 099 805 277')]:
    if lbl in vrow: vs.cell(vrow[lbl],2).value=abn
cf=wb['Config']; cr={str(cf.cell(i,1).value):i for i in range(1,cf.max_row+1) if cf.cell(i,1).value}
for k,v in [('WORKBOOK_VERSION','v46'),('CL_DATA',f'5:{CL_END}'),('SITES_PANELA',f'{PA_START}:{PA_END}'),
            ('SITES_GROUP_ROWS',f'{GRP_START}:{GRP_END}'),('SITES_TOTAL_ROW',NTOT),('SITES_CHECK_ROW',NCHK),
            ('SITES_PANELB',f'{NPB+2}:{NPB_END}'),('PARK_COUNT',len(order))]:
    if k in cr: cf.cell(cr[k],2).value=v
ca=copy.copy(cf.cell(cf.max_row,1)._style); cb=copy.copy(cf.cell(cf.max_row,2)._style); nr=cf.max_row+1
for k,v in [('SITE_CROSSWALK_DATA',f'5:{XW_END}'),('IDENT_ROWS_V46',n_id),('IDENT_HELD_V46',len(P['held']))]:
    a=cf.cell(nr,1); a.value=k; a._style=copy.copy(ca); b=cf.cell(nr,2); b.value=v; b._style=copy.copy(cb); nr+=1
h=wb['Handover']; hr=h.max_row+1; hs=[copy.copy(h.cell(h.max_row,c)._style) for c in range(1,4)]
h.cell(hr,1).value=(f'v46 change (18-Aug-2026): identification build plus site corrections. {n_id} register rows identified from five APLEDGER creditor histories '
 f'(ELE032, QPO001, SEC022, SEA029, HIB003), ${sum(t["amount"] for t in TGT.values()):,.2f} ex GST, every one amount-corroborated at incl/1.1 within 2c; '
 f'{len(P["held"])} references held on the amount test. Site: "Shailer Park" removed from the gazetteer (a SUBURB and the vendor\u2019s business address, {n_rm} rows), '
 f'three user overrides applied, named non-park Council facilities recognised. NEW sheet Site_Crosswalk holds the manual decisions and outranks the matcher permanently. '
 f'Column T never written; no green block touched; control total unchanged at ${CT:,.2f}.')
h.cell(hr,1)._style=hs[0]
h.cell(hr,2).value=f'Creditor_Lines +{len(P["cred"])} rows to {CL_END}. Sites Panel A {len(order)} sites. Skipped {P["skipped"]}.'
h.cell(hr,2)._style=hs[1]; h.cell(hr,3)._style=hs[2]
m=wb['Method']; mr=m.max_row+1; ms=copy.copy(m.cell(m.max_row,1)._style)
METH=[f"45.0 Batch 41 identification (v46). Five APLEDGER creditor histories: ELE032 Electrical Data & Security Services Pty Ltd (4,444 invoices, ABN 17 099 805 277), QPO001 Q Power (Qld) Pty Ltd (1,131, ABN 82 067 507 591), SEC022 (24, ABN 97 613 325 769), SEA029 Sea-Crete (1, ABN 84 127 054 291), HIB003 (2, ABN 30 306 518 188). {n_id} register rows identified, ${sum(t['amount'] for t in TGT.values()):,.2f}. Amount corroboration is the gate (v35): {len(P['held'])} references matched on reference but failed the incl/1.1 test within 2c and are HELD. {P['skipped'].get('confirmed',0)} Confirmed rows skipped. Closes four of the six unidentified series: FY2023/24 7-digit and FY2024/25 7-digit (ELE032, shared with SEC022 - the 7-digit series is multi-vendor), FY2024/25 and FY2025/26 5-digit (QPO001, resolving Open Item #89 - the amount test separates Q Power from Logan Plumbing cleanly).",
 f"45.1 A published address is a starting point, not an authority (v46, standing rule). The Parks Directory lists Ewing Park at Netball Drive and Rainbow Park at Ewing Road, so an address match sent all 237 lines at 99 Ewing Rd Woodridge to Rainbow Park. The user corrected it to Ewing Park. Street-ONLY matching is now banned outright: it sent '293 Logan St Eagleby' to Doug Larsen Park, which is 41 Logan Street BEENLEIGH, a cross-suburb error on 42 lines. Only street PLUS suburb may be proposed, and only as a flagged proposal.",
 f"45.2 NEW sheet Site_Crosswalk (v46). Manual, user-adjudicated site decisions with the basis recorded. The crosswalk OUTRANKS the automatic matcher and is applied on every build, so local knowledge survives a rebuild. Seeded with seven decisions including the removal of 'Shailer Park', which is a SUBURB and David Robinson Landscaping's business address (15 Helios St, Shailer Park, per PO 802122): it had taken $50,140 across {n_rm} lines and beat 'Homestead Park (Shailer Park)' on an exact match.",
 f"45.3 Non-park Council facilities are sites (v46). Olivers Sports Complex, West Logan Netball Courts, the Historical Society Amenities and the Logan and District Services Club are real Council sites carrying real spend; they are recorded under the basis '{NEWFAC}' rather than forced onto a park or left unresolved. This extends the Parks Depot precedent (v43) from one facility to a class."]
for i,t in enumerate(METH):
    c=m.cell(mr+i,1); c.value=t; c._style=copy.copy(ms)
d=wb['Data_Acquisition']; dr=d.max_row+1; ds=copy.copy(d.cell(d.max_row,1)._style)
c=d.cell(dr,1); c.value=(f' |  F44  |  Batch 41: five APLEDGER creditor histories (ELE032, QPO001, SEC022, SEA029, HIB003), {len(P["cred"])} invoice rows, appended to Creditor_Lines at v46. '
 f'Identification only; no invoice capture, sighted count unchanged.  |  1  |  Captured  | '); c._style=copy.copy(ds)
oi=wb['Open_Items']; orow=oi.max_row+1; osx=[copy.copy(oi.cell(oi.max_row,c)._style) for c in range(1,8)]
ITEMS=[(144,'All','Closed','Unidentified series',325,160484.68,'CLOSED at v46. The FY2023/24 and FY2024/25 7-digit series is Electrical Data & Security Services Pty Ltd, APLEDGER ELE032, ABN 17 099 805 277. Note the 7-digit series is MULTI-VENDOR: SEC022 also posts 7-digit references. CINC EDSS Pty Ltd and Electrical Data & Security Services Pty Ltd are one creditor account and should merge on Vendor_Series.'),
 (145,'All','Closed','Unidentified series',141,123109.75,'CLOSED at v46. The FY2024/25, FY2025/26 and FY2026/27 5-digit series is Q Power (Qld) Pty Ltd, APLEDGER QPO001, ABN 82 067 507 591. This also resolves Open Item #89: the amount test at incl/1.1 separates Q Power from Logan Plumbing Service cleanly, 171 of 187 references corroborating.'),
 (146,'All','Open','APLEDGER SEC022 / HIB003',18,0,'SUPPLIERS UNNAMED. SEC022 (ABN 97 613 325 769, 24 invoices, narrations name Underwood Park and Oliver Sports Complex) and HIB003 (ABN 30 306 518 188, 2 invoices, Tygum Pk solar and basketball lighting) are identified by ABN but not by name. Ask Finance for the creditor names, or run an ABR lookup.'),
 (147,'All','Open','Site dimension',len(P['held']),0,'IDENTIFICATION HELD ON AMOUNT. 20 references matched a creditor history by reference but the transaction amount does not tie the register ex GST at 1.1 within 2c. Reference alone is not evidence (v35). Held for invoice sighting.'),
 (148,'All','Open','Site dimension',0,112153.00,'STREET-ONLY ADDRESS MATCHES HELD FOR USER ADJUDICATION. 121 narration segments, 885 lines, $112,153, resolve to a park on street name alone. Street-only matching is banned (Method 45.1) because it produced cross-suburb errors. Each needs one line of local knowledge; the list ships as a decision sheet.'),
 (149,'All','Open','Gazetteer',0,0,'GAZETTEER ADDITIONS PENDING. From the creditor narrations: Crestmead Park, Flagstone Regional Park, Stoneleigh Reserve Park (alias Stoneleigh Park), Brough Place Park, East Beaumont Park, Everleigh Park, Kurrajong Park, Kohlmann Park, Sebring Park, Quandong Park. Also de-hyphenate narration line-wraps ("Underw-ood Park").')]
for i,it in enumerate(ITEMS):
    for j,v in enumerate(it,1):
        c=oi.cell(orow+i,j); c.value=v; c._style=copy.copy(osx[j-1])

pre=f'{OUT}/{NEW}'; wb.save(pre); print('saved')
prof=f'{ROOT}/lo_profile/user'; os.makedirs(prof,exist_ok=True)
open(f'{prof}/registrymodifications.xcu','w').write(open('/home/claude/b43/lo_profile/user/registrymodifications.xcu').read())
t0=time.time()
res=subprocess.run(['soffice',f'-env:UserInstallation=file://{ROOT}/lo_profile','--headless','--calc','--convert-to','xlsx','--outdir',RC,pre],capture_output=True,text=True,timeout=1800)
final=f'{RC}/{NEW}'; assert os.path.exists(final),res.stdout+res.stderr; print(f'recalc {time.time()-t0:.0f}s')
json.dump(dict(final=final,CT=CT,PA=[PA_START,PA_END],GRP=[GRP_START,GRP_END],TOT=NTOT,CHK=NCHK,PB=[NPB+2,NPB_END],
  parks=len(order),n_id=n_id,ident_rows=sorted(TGT),CL_END=CL_END,XW_END=XW_END,CLASS=CLASS,MA=MA,NEWFAC=NEWFAC,
  XW_BASIS=XW_BASIS,gb=int((df.iloc[:,87].notna()).sum())),open(f'{ROOT}/b46_pos.json','w'),indent=1)
r=subprocess.run([sys.executable,f'{ROOT}/b46_verify.py'],capture_output=True,text=True); print(r.stdout); print(r.stderr[-1200:])
if 'FAILS: none' in r.stdout:
    subprocess.run(['cp',final,f'/mnt/user-data/outputs/{NEW}']); print('SHIPPED',NEW)
else: print('NOT SHIPPED'); sys.exit(1)
