"""b45_build.py  v44 -> v45  (Batch 40: qpower_brizsouth_20260818, 18-Aug-2026)
Rule 17 capture: 44 invoices (Q Power 13, BrizSouth Locksmiths 31), 121 printed lines.
EI append + EIL append + recon relocation (per-row E-form preserved) + green blocks + companion notes
+ site writes (printed Site / Job Location, Tier 1) + Vendor_Boilerplate + Config/Handover/Method/DA/Open_Items.
Column T never written; control total unchanged. One save -> one convert recalc -> one exact-TRUE verify -> ship.
"""
import json, copy, re, os, sys, subprocess, time, collections
import openpyxl, pandas as pd
from openpyxl.utils import get_column_letter
from python_calamine import CalamineWorkbook

ROOT='/home/claude/b45'; SRC='/mnt/user-data/outputs/PS_WP_Transaction_Register_3FY_v44_HANDOVER.xlsx'
NEW='PS_WP_Transaction_Register_3FY_v45_HANDOVER.xlsx'
OUT=f'{ROOT}/out'; RC=f'{ROOT}/out/recalc'; os.makedirs(RC,exist_ok=True)
CT=11377312.51
P=json.load(open(f'{ROOT}/b45_parse.json')); F=P['facts']; L=P['lines']; M=P['match']
cw=CalamineWorkbook.from_path(SRC)
cfg={str(r[0]):r[1] for r in cw.get_sheet_by_name('Config').to_python() if r and r[0]}
EI0,EI1=[int(x) for x in str(cfg['EI_DATA']).split(':')]; EI_TOT=int(cfg['EI_TOTALS_ROW'])
C0,C1=[int(x) for x in str(cfg['EI_CONTROL_ROWS']).split(':')]
EIL0,EIL_END=5,int(cfg['EIL_END']); R0,R1=[int(x) for x in str(cfg['RECON_ROWS']).split(':')]
SIGHT=int(cfg['SIGHTED_COUNT']); PA0,PA1=[int(x) for x in str(cfg['SITES_PANELA']).split(':')]
SA0,SA1=[int(x) for x in str(cfg['SITE_ALLOC_DATA']).split(':')]
GRP0,GRP1=[int(x) for x in str(cfg['SITES_GROUP_ROWS']).split(':')]
GAZ=[str(x[0]).strip() for x in cw.get_sheet_by_name('Sites').to_python()[PA0-1:PA1]]

INV=sorted(F, key=lambda k:(F[k]['vendor'],k))
EVID={k:k for k in INV}                                    # C#### and 5-digit ids stay as printed (schema §5)
NEW_EIL=sum(len(L[k]) for k in INV); NEW_EI=len(INV)
BP={'Q Power (Qld) Pty Ltd':('BP:QP-P1','BP:QP-T1'),'BrizSouth Locksmiths Pty Ltd':('BP:C-P1','BP:C-T1')}
BP_TEXT={'BP:QP-P1':('Payment details','To avoid surcharge please pay via Direct Deposit. Credit Card (MasterCard or Visa) Direct Deposit | Pay Online qpow.simprosuite.com/payment/ | Bank Suncorp | Acc. Name Q Power | BSB 484-799 | Acc. No. 611134396'),
          'BP:QP-T1':('Terms & notices','Please call 07 3155 4225 to pay over the phone. | Card Holder\u2019s Name: CCV: | Expiry Date: / Signature: | Powered by TCPDF (www.tcpdf.org)'),
          'BP:C-T1':('Terms & notices','WE APPRECIATE YOUR BUSINESS')}

# ---------------- positions after append
NEW_EI_END=EI1+NEW_EI; SHIFT_EI=NEW_EI
NEW_EI_TOT=EI_TOT+SHIFT_EI; NEW_C0=C0+SHIFT_EI; NEW_C1=C1+SHIFT_EI
NEW_EIL_END=EIL_END+NEW_EIL; SHIFT_EIL=NEW_EIL
NEW_R0=R0+SHIFT_EIL; NEW_R1=R1+SHIFT_EIL+NEW_EI
NEW_SIGHT=SIGHT+NEW_EI

wb=openpyxl.load_workbook(SRC)
reg=wb['Register']; assert reg.max_row==21475
we=wb['Evidence_Invoices']; wl=wb['Evidence_Invoice_Lines']

def sty(c,t): c._style=copy.copy(t._style)
# ---------------- 1. EIL: capture old recon block, clear, write new lines, relocate recon
old_recon=[[wl.cell(r,c).value for c in range(1,7)] for r in range(R0,R1+1)]
old_recon_st=[[copy.copy(wl.cell(r,c)._style) for c in range(1,7)] for r in range(R0,R1+1)]
rec_hdr=[[wl.cell(R0-1,c).value for c in range(1,7)],[copy.copy(wl.cell(R0-1,c)._style) for c in range(1,7)]]
for r in range(R0-1,R1+1):
    for c in range(1,7): wl.cell(r,c).value=None
tmplL=[wl.cell(6,c) for c in range(1,14)]
r=EIL_END+1
for k in INV:
    for o in L[k]:
        vals=[EVID[k],o['n'],o['desc'],(f"{o['qty']:.2f}" if o['qty'] is not None else '(not printed)'),
              (o['unit'] if o['unit'] is not None else '(not printed)'),o['gst'],o['amt'],
              (F[k]['pk'] or '(not printed)'),(F[k]['pk'] or ''),M[k]['target_lk'],
              f"Batch 40 capture; printed line {o['n']} of {len(L[k])}.",
              (F[k]['site_printed'] or ''),('Printed line names one park' if F[k]['gaz_sites'] and len(F[k]['gaz_sites'])==1 else 'No park named on the printed line')]
        for c,v in enumerate(vals,1):
            cell=wl.cell(r,c); cell.value=v; sty(cell,tmplL[c-1])
        r+=1
assert r-1==NEW_EIL_END,(r-1,NEW_EIL_END)
# recon block relocated + new rows
rr=NEW_R0-1
for c,(v,s) in enumerate(zip(rec_hdr[0],rec_hdr[1]),1):
    cell=wl.cell(rr,c); cell.value=v; cell._style=copy.copy(s)
rr=NEW_R0
recon_row={}
for src_r,(vals,sts) in enumerate(zip(old_recon,old_recon_st),R0):
    for c,(v,s) in enumerate(zip(vals,sts),1):
        if isinstance(v,str) and v.startswith('='):
            # retarget the EIL data range and repoint the row self-reference; PRESERVE the row's own E-form
            v=re.sub(r'\$A\$5:\$A\$%d\b'%EIL_END, f'$A$5:$A${NEW_EIL_END}', v)
            v=re.sub(r'\$G\$5:\$G\$%d\b'%EIL_END, f'$G$5:$G${NEW_EIL_END}', v)
            v=re.sub(r'\bD%d\b'%src_r, f'D{rr}', v)
            v=re.sub(r'\bC%d\b'%src_r, f'C{rr}', v)
        cell=wl.cell(rr,c); cell.value=v; cell._style=copy.copy(s)
    if vals[2]: recon_row[str(vals[2])]=rr
    rr+=1
tmplR=[copy.copy(wl.cell(NEW_R0,c)._style) for c in range(1,7)]
for k in INV:
    tgt=F[k]['sub']
    for c,v in enumerate(['','',EVID[k],
        f'=ROUND(SUMIF($A${EIL0}:$A${NEW_EIL_END},C{rr},$G${EIL0}:$G${NEW_EIL_END}),2)',
        f'=IF(D{rr}={tgt},"TRUE","FALSE")',''],1):
        cell=wl.cell(rr,c); cell.value=(v if v!='' else None); cell._style=copy.copy(tmplR[c-1])
    recon_row[EVID[k]]=rr; rr+=1
assert rr-1==NEW_R1,(rr-1,NEW_R1)

# ---------------- 2. EI append (tail relocation)
tail=[[we.cell(r,c).value for c in range(1,33)] for r in range(EI_TOT,C1+1)]
tail_st=[[copy.copy(we.cell(r,c)._style) for c in range(1,33)] for r in range(EI_TOT,C1+1)]
for r in range(EI_TOT,C1+1):
    for c in range(1,33): we.cell(r,c).value=None
tmplE=[we.cell(EI1,c) for c in range(1,33)]
r=EI1+1
for k in INV:
    f=F[k]; m=M[k]
    vals=[EVID[k],f['vendor'],f['abn'],f['date'],'(not printed)',k,(f['order'] or '(not printed)'),
          (f['site_printed'] or '(not printed)'),len(L[k]),f['sub'],f['gst'],f['tot'],
          f"Lines reconcile to the printed subtotal to the cent; reconciliation at Evidence_Invoice_Lines row {recon_row[EVID[k]]}. Rule 17 checks live on Register row {m['target_row']}.",
          'Complete on the register line (green block, columns CJ to DW)']+['']*18
    for c,v in enumerate(vals,1):
        cell=we.cell(r,c); cell.value=(v if v!='' else None); sty(cell,tmplE[c-1])
    r+=1
assert r-1==NEW_EI_END
for i,(vals,sts) in enumerate(zip(tail,tail_st)):
    for c,(v,s) in enumerate(zip(vals,sts),1):
        cell=we.cell(NEW_EI_TOT+i,c); cell.value=v; cell._style=copy.copy(s)
# totals row explicit (self-inclusive SUM trap)
for c,col in ((9,'I'),(10,'J'),(11,'K'),(12,'L')):
    we.cell(NEW_EI_TOT,c).value=f'=SUM({col}{EI0}:{col}{NEW_EI_END})'
# control formulas: retarget ranges
for r in range(NEW_C0,NEW_C1+1):
    for c in range(1,33):
        v=we.cell(r,c).value
        if isinstance(v,str) and v.startswith('='):
            v=re.sub(r'\$A\$5:\$A\$%d\b'%EIL_END,f'$A$5:$A${NEW_EIL_END}',v)
            v=re.sub(r'\$G\$5:\$G\$%d\b'%EIL_END,f'$G$5:$G${NEW_EIL_END}',v)
            v=re.sub(r'\$E\$%d:\$E\$%d\b'%(R0,R1),f'$E${NEW_R0}:$E${NEW_R1}',v)
            v=re.sub(r'\$C\$%d:\$C\$%d\b'%(R0,R1),f'$C${NEW_R0}:$C${NEW_R1}',v)
            v=re.sub(r'\$A\$5:\$A\$%d\b'%EI1,f'$A$5:$A${NEW_EI_END}',v)
            v=re.sub(r'\bJ%d\b'%EI_TOT,f'J{NEW_EI_TOT}',v); v=re.sub(r'\bI%d\b'%EI_TOT,f'I{NEW_EI_TOT}',v)
            # control-block self references (C{old ctl row}) and the EI data tail range
            for oc in range(C0,C1+1):
                v=re.sub(r'\bC%d\b'%oc, f'C{oc+SHIFT_EI}', v)
            v=re.sub(r'\$AF\$5:\$AF\$%d\b'%(EI_TOT-1), f'$AF$5:$AF${NEW_EI_TOT-1}', v)
            v=re.sub(r'\$N\$5:\$N\$11\b','$N$5:$N$11',v)
            we.cell(r,c).value=v
        if isinstance(v,(int,float)) and float(v)==float(SIGHT): we.cell(r,c).value=NEW_SIGHT
# the two count rows and the recon-count / sighted-count literals inside formulas
for r in range(NEW_C0,NEW_C1+1):
    for c in range(1,33):
        v=we.cell(r,c).value
        if isinstance(v,str) and v.startswith('='):
            v=v.replace(f'={SIGHT},',f'={NEW_SIGHT},').replace(f'={int(cfg["RECON_COUNT"])},',f'={int(cfg["RECON_COUNT"])+NEW_EI},')
            we.cell(r,c).value=v

# ---------------- 3. Register green blocks + companions + site
def gb(k):
    f=F[k]; m=M[k]; r=m['target_row']
    items=' '.join(f"{o['n']}. {o['desc']} | {o['qty'] if o['qty'] is not None else '-'} @ ${o['unit'] if o['unit'] is not None else '-'} = ${o['amt']:.2f}." for o in L[k])
    pay,terms=BP[f['vendor']]
    vals={88:EVID[k],89:f['vendor'],90:f['abn'],91:'(not printed)',
          92:('8/27 Allgas Street, Slacks Creek QLD 4127' if f['vendor'].startswith('Q Power') else '7/35 Tradelink Drive, Hillcrest QLD 4118'),
          93:('Tel. 07 3155 4225 | qpower.com.au | Licence # 69807' if f['vendor'].startswith('Q Power') else 'Mobile: 0407 691 755 | Phone: 38007744'),
          94:f['date'],95:'(not printed)',96:(f['order'] or '(not printed)'),97:'(not printed)',
          98:('Logan City Council, PO Box 3226, Logan City DC QLD 4114' if f['vendor'].startswith('Q Power') else 'Logan City Council Parks Services, 177 Chambers Flat Road, MARSDEN QLD 4132'),
          99:'(not printed)',100:(f['officer'] or '(not printed)'),101:'(not printed)',102:'(not printed)',
          103:(f['cr'] or '(not printed)'),104:(f['pk'] or '(not printed)'),105:(f['pk'] or '(not printed)'),
          106:(f['site_printed'] or '(not printed)'),107:'(not printed)',
          108:(f['site_printed'] or '(not printed)'),109:(f['tech'] or '(not printed)'),
          110:len(L[k]),111:items,112:(f['works'] or '(not printed)'),
          113:f['sub'],114:f['gst'],115:f['tot'],116:'(not printed)',117:'(not printed)',
          118:pay,119:terms,120:f['pages'],
          121:f'=ROUND(SUMIF(Evidence_Invoice_Lines!$A${EIL0}:$A${NEW_EIL_END},CJ{r},Evidence_Invoice_Lines!$G${EIL0}:$G${NEW_EIL_END}),2)',
          122:f'=IF(DQ{r}=DI{r},"TRUE","FALSE")',123:f'=IF(ROUND(T{r},2)=DI{r},"TRUE","FALSE")',
          124:f'=IF(DJ{r}=ROUND(DI{r}*0.1,2),"TRUE","FALSE")',
          125:'Batch 40 (qpower_brizsouth_20260818; QPower.pdf + Brizsouth Locksmiths Pty Ltd.pdf, 64 pages)',
          126:'18-Aug-2026, rule 17 capture (v45)',
          127:('Printed total incl GST restated as subtotal + printed GST: the corpus field held the subtotal (defect family F1, all 31 BrizSouth invoices). ' if f['f1_total'] else '')
             +('No PK printed on the invoice face; PK taken from the register posting. ' if not f['pk'] else '')}
    return r,vals
tmplG=[reg.cell(14937,c) for c in range(88,128)]
gb_rows=[]
for k in INV:
    r,vals=gb(k); gb_rows.append(r)
    for c,v in vals.items():
        cell=reg.cell(r,c); cell.value=v; sty(cell,tmplG[c-88])
    # analysis columns
    reg.cell(r,24).value='Sighted invoice line'                      # Nature Basis col 27? -> keep col 27
    reg.cell(r,27).value='Sighted invoice line'
    reg.cell(r,29).value='Tier 1'
    reg.cell(r,33).value='Confirmed'
    reg.cell(r,34).value=None
    reg.cell(r,32).value=f"Invoice sighted (rule 17): {F[k]['vendor']}, {F[k]['site_printed'] or 'site not printed'}; printed subtotal ${F[k]['sub']:,.2f}."
    if F[k]['gaz_sites'] and len(F[k]['gaz_sites'])==1:
        reg.cell(r,129).value=F[k]['gaz_sites'][0]; reg.cell(r,130).value='Sighted invoice site (printed)'
    for comp in M[k]['companions']:
        reg.cell(comp['row'],32).value=f"$0.00 companion of the sighted line {M[k]['target_lk']} (reference {k}); no green block per schema section 8."
        if F[k]['gaz_sites'] and len(F[k]['gaz_sites'])==1:
            reg.cell(comp['row'],129).value=F[k]['gaz_sites'][0]; reg.cell(comp['row'],130).value='Sighted invoice site (printed, companion of the sighted sibling)'
# Nature Category/Detail restated (Amendment 1: must not be generic)
for k in INV:
    r=M[k]['target_row']
    cat='Electrical & data services' if F[k]['vendor'].startswith('Q Power') else 'Security & access'   # MUST exist in Theme_Map A5:A31 (rule 4)
    reg.cell(r,24).value=cat
    reg.cell(r,25).value='; '.join(o['desc'] for o in L[k])[:250]

# ---------------- 4. Vendor_Boilerplate
vb=wb['Vendor_Boilerplate']; vrow={str(vb.cell(i,1).value):i for i in range(5,vb.max_row+1) if vb.cell(i,1).value}
vs=[copy.copy(vb.cell(5,c)._style) for c in range(1,7)]
nb=vb.max_row+1; added=[]
for key,(fld,text) in BP_TEXT.items():
    if key in vrow: continue
    ex=next(k for k in INV if BP[F[k]['vendor']][0]==key or BP[F[k]['vendor']][1]==key)
    for c,v in enumerate([key,fld,F[ex]['vendor'],EVID[ex],0,text],1):
        cell=vb.cell(nb,c); cell.value=v; cell._style=copy.copy(vs[c-1])
    vrow[key]=nb; added.append(key); nb+=1
for key,i in vrow.items():
    n=sum(1 for k in INV if key in BP[F[k]['vendor']])
    if n and isinstance(vb.cell(i,5).value,(int,float)): vb.cell(i,5).value=int(vb.cell(i,5).value)+n
BPK=int(cfg['BOILERPLATE_KEYS'])+len(added)

# ---------------- 5. Config / narrative sheets
cf=wb['Config']; cr={str(cf.cell(i,1).value):i for i in range(1,cf.max_row+1) if cf.cell(i,1).value}
for k,v in [('WORKBOOK_VERSION','v45'),('EI_DATA',f'5:{NEW_EI_END}'),('EI_TOTALS_ROW',NEW_EI_TOT),
            ('EI_CONTROL_ROWS',f'{NEW_C0}:{NEW_C1}'),('EIL_DATA',f'5:{NEW_EIL_END}'),('EIL_END',NEW_EIL_END),
            ('RECON_ROWS',f'{NEW_R0}:{NEW_R1}'),('RECON_COUNT',int(cfg['RECON_COUNT'])+NEW_EI),
            ('SIGHTED_COUNT',NEW_SIGHT),('BOILERPLATE_KEYS',BPK)]:
    cf.cell(cr[k],2).value=v
h=wb['Handover']; hr=h.max_row+1; hs=[copy.copy(h.cell(h.max_row,c)._style) for c in range(1,4)]
h.cell(hr,1).value=(f'v45 change (18-Aug-2026): Batch 40 rule 17 capture. 44 invoices sighted (Q Power (Qld) Pty Ltd 13, BrizSouth Locksmiths Pty Ltd 31), '
  f'121 printed lines, ${sum(F[k]["sub"] for k in INV):,.2f} ex GST. 44 green blocks; 28 $0.00 companion rows cross-referenced; '
  f'{sum(1 for k in INV if F[k]["gaz_sites"])} rows gain a Tier 1 printed site. Capture build: column T untouched, control total unchanged at ${CT:,.2f}.')
h.cell(hr,1)._style=hs[0]
h.cell(hr,2).value=f'Sighted count {SIGHT} -> {NEW_SIGHT}. Corpus sha256 e3a4a0fb...9416, extraction Microsoft 365 Copilot / GPT-5. 2 duplicate-copy records screened out (15181 pp26-27, C13513 p24). F1 recurred: printed_total_incl_gst held the SUBTOTAL on all 31 BrizSouth invoices.'
h.cell(hr,2)._style=hs[1]; h.cell(hr,3)._style=hs[2]
m=wb['Method']; mr=m.max_row+1; ms=copy.copy(m.cell(m.max_row,1)._style)
METH=[f"44.0 Batch 40 capture (v45). 46 corpus records, 44 primary and 2 duplicate copies. All 44 self-tie to the printed subtotal; every priced record is sourced from a TEXT line (none OCR-only); 64 of 64 pages represented; 121 of 121 priced records carry a numeric line amount. Rule 12 screen prefix-aware over all 28 EvID prefixes on Evidence_Invoices plus the bare id: 0 re-sightings. Corpus-internal normalised-text screen (v40 rule): 0 collisions. Every invoice matched to exactly ONE non-zero register row at the printed subtotal within 2c (amount corroboration, v35 rule); 28 $0.00 companion rows carry a cross-reference and no green block.",
 f"44.1 Defect family F1 recurs, third form (v45). `printed_total_incl_gst` held the SUBTOTAL on all 31 BrizSouth Locksmiths invoices (the printed TOTAL prints below the 'PLUS 10% GST' line and the same-line regex returned the subtotal). Restated as subtotal + printed GST and re-tied arithmetically on all 31; the restatement is stated in each anomalies field. Queue with the v33/v34/v36 findings for extraction prompt v3.",
 f"44.2 Site is printed on the invoice face for both vendors (v45). Q Power prints 'Site:' and 'Site Address:' as header rows; BrizSouth prints 'Job Location: <park>, <SUBURB>' in the description block. 41 of 44 invoices name a gazetteer park directly, so this batch is unusually rich for the site dimension: the printed site is written to Register DY with basis 'Sighted invoice site (printed)' on the sighted row and on its $0.00 companion. Two invoices print no site (C13293, C13403) and seven print no PK (15182, C13175, C13403, C13447, C13537, C13580, C13591) - Open Items #140.",
 f"44.3 Boilerplate (v45). Q Power prints one payment block and one terms notice, minted BP:QP-P1 / BP:QP-T1. BrizSouth reuses the existing BP:C-P1 (first sighted C13564) and gains BP:C-T1 for the 'WE APPRECIATE YOUR BUSINESS' notice. One key per cell (v34 trap)."]
for i,t in enumerate(METH):
    c=m.cell(mr+i,1); c.value=t; c._style=copy.copy(ms)
d=wb['Data_Acquisition']; dr=d.max_row+1; ds=copy.copy(d.cell(d.max_row,1)._style)
c=d.cell(dr,1); c.value=(f' |  F43  |  Batch 40: corpus_qpower_brizsouth_20260818.json (sha256 e3a4a0fbaf05e644f144523c53ba5b0aeb750163ff9dce0d08902e92f8999416, md5 {P["md5"]}), '
 f'extraction tool Microsoft 365 Copilot / GPT-5 reasoning model, 2 source PDFs, 64 pages, 46 records, 4,985 lines. Captured in full at v45.  |  1  |  Captured  | '); c._style=copy.copy(ds)
oi=wb['Open_Items']; orow=oi.max_row+1; osx=[copy.copy(oi.cell(oi.max_row,c)._style) for c in range(1,8)]
noPK=[k for k in INV if not F[k]['pk']]
ITEMS=[(140,'All','Open','Q Power / BrizSouth',len(noPK),round(sum(F[k]['sub'] for k in noPK),2),'NO PK ON THE INVOICE FACE. '+', '.join(noPK)+'. PK taken from the register posting and the printed form recorded as "(not printed)". Ask both suppliers to print the PK, as the rest of the series does.'),
 (141,'FY2026/27','Open','Q Power (Qld) Pty Ltd',1,300.00,'PRINTED PK DIVERGES FROM THE CHARGE. Invoice 15134 prints PK000442 (Doug Larsen Park) and posts to PK000412. Same pattern as Open Items #100/#101/#102.'),
 (142,'All','Open','Q Power (Qld) Pty Ltd',3,0,'THREE Q POWER REGISTER ROWS CARRY NO CREDITOR CODE (contractor label "Q Power (Qld) Pty Ltd (series-inferred)"). Now sighted at rule 17, so the identity is Tier 1; the creditor code QPO001 should be back-filled from APLEDGER.'),
 (143,'All','Open','Batch 40 evidence gaps',6,0,'REFERENCED ATTACHMENTS NOT SUPPLIED. 15134, 15135, 15148, 15150, 15182 reference quotes; 15198 references a quote, site mud map and site photographs. Request if the works scope is queried.')]
for i,it in enumerate(ITEMS):
    for j,v in enumerate(it,1):
        c=oi.cell(orow+i,j); c.value=v; c._style=copy.copy(osx[j-1])

# ---------------- 6. save / recalc / verify / ship
pre=f'{OUT}/{NEW}'; os.makedirs(OUT,exist_ok=True); wb.save(pre); print('saved')
prof=f'{ROOT}/lo_profile/user'; os.makedirs(prof,exist_ok=True)
open(f'{prof}/registrymodifications.xcu','w').write(open('/home/claude/b43/lo_profile/user/registrymodifications.xcu').read())
t=time.time()
res=subprocess.run(['soffice',f'-env:UserInstallation=file://{ROOT}/lo_profile','--headless','--calc','--convert-to','xlsx','--outdir',RC,pre],capture_output=True,text=True,timeout=1800)
final=f'{RC}/{NEW}'; assert os.path.exists(final),res.stdout+res.stderr; print(f'recalc {time.time()-t:.0f}s')
json.dump(dict(final=final,CT=CT,EI_END=NEW_EI_END,EI_TOT=NEW_EI_TOT,C=[NEW_C0,NEW_C1],EIL_END=NEW_EIL_END,
  R=[NEW_R0,NEW_R1],SIGHT=NEW_SIGHT,gb=len(gb_rows),gb_rows=gb_rows,invs=INV,BPK=BPK,
  companions=[c['row'] for k in INV for c in M[k]['companions']]),open(f'{ROOT}/b45_pos.json','w'),indent=1)
r=subprocess.run([sys.executable,f'{ROOT}/b45_verify.py'],capture_output=True,text=True); print(r.stdout); print(r.stderr[-1500:])
if 'FAILS: none' in r.stdout:
    subprocess.run(['cp',final,f'/mnt/user-data/outputs/{NEW}']); print('SHIPPED',NEW)
else: print('NOT SHIPPED'); sys.exit(1)
