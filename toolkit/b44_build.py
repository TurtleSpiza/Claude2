"""b44_build.py  v43 -> v44 (18-Aug-2026)
(1) Evidence_Invoice_Lines gains cols L Site (printed line) and M Site basis - site read per PRINTED LINE.
(2) New sheet Site_Allocation: every register row whose printed lines name more than one park is exploded per park
    with the PRINTED line amounts; residue row forces the LineKey to tie to the register amount. Multi-site money is
    allocated on evidence, not held. Sites Panel A sums Register (single) + Site_Allocation (multi).
(3) Electricity: supply-point addresses matched to THREE Council address records (rates parcel narrations, Parks
    Strategy Layer House_Add, Parks Directory). number+street exact & unique -> written; street unique -> written and
    flagged AMBER for user check; ambiguous / unknown -> stay in the address-only class. Every write listed.
(4) Gazetteer gains the water-park printed site names. Column T never written; control total unchanged.
"""
import json, copy, collections, re, os, sys, subprocess, time
import openpyxl, pandas as pd
from openpyxl.utils import get_column_letter
from python_calamine import CalamineWorkbook

ROOT='/home/claude/b44'; os.makedirs(ROOT,exist_ok=True)
SRC='/mnt/user-data/outputs/PS_WP_Transaction_Register_3FY_v43_HANDOVER.xlsx'
NEW='PS_WP_Transaction_Register_3FY_v44_HANDOVER.xlsx'
OUT_DIR=f'{ROOT}/out'; RECALC_DIR=f'{ROOT}/out/recalc'; os.makedirs(RECALC_DIR,exist_ok=True)
CT=11377312.51; SC='DY'; BC='DZ'; SITE_COL=129; BASIS_COL=130
MULTI='Multiple sites named (not allocated)'
MULTI_ALLOC='Multiple sites named (allocated per printed line, see Site_Allocation)'
RESIDUE='(residue: printed lines naming no park or several)'
L_ADDR='Supply-point address only (address route pending)'
E_GREEN='Electricity supply point (address = Council parcel record, number and street exact)'
E_AMBER='Electricity supply point (address = Council parcel record, street unique) [USER CHECK]'
FYS=['FY2023/24','FY2024/25','FY2025/26','FY2026/27']
WATER={'flagstone water play':'Flagstone Water Play','flagstone':'Flagstone Water Play','beenleigh water play':'Beenleigh Water Play',
       'forest glen water feature':'Forest Glen Water Feature','forest glen':'Forest Glen Water Feature','underwood water feature':'Underwood Park','logan gardens':'Logan Gardens'}
NEW_GAZ=['Flagstone Water Play','Beenleigh Water Play','Forest Glen Water Feature']

cwb=CalamineWorkbook.from_path(SRC)
cfg={str(r[0]):r[1] for r in cwb.get_sheet_by_name('Config').to_python() if r and r[0]}
pa0,pa1=[int(x) for x in str(cfg['SITES_PANELA']).split(':')]
gaz=[str(x[0]).strip() for x in cwb.get_sheet_by_name('Sites').to_python()[pa0-1:pa1]]
gaz=sorted(set(gaz)|set(NEW_GAZ))
CANON={g.lower():g for g in gaz}
GP=re.compile('('+'|'.join(re.escape(g) for g in sorted(gaz,key=len,reverse=True) if len(g)>5)+')',re.I)
CLASS_LABELS=json.load(open('/home/claude/b43/b43_pos.json'))['CLASS_LABELS']

df=pd.read_excel(SRC,sheet_name='Register',header=3,engine='calamine').iloc[:21469]
amt=pd.to_numeric(df.iloc[:,19],errors='coerce').fillna(0); assert abs(amt.sum()-CT)<0.005
lk=df.iloc[:,0].astype(str); fyv=df.iloc[:,4].astype(str); ref=df.iloc[:,7].astype(str); pkv=df.iloc[:,16].astype(str)
site_old=df.iloc[:,SITE_COL-1].fillna('').astype(str); basis_old=df.iloc[:,BASIS_COL-1].fillna('').astype(str)
row_of={k:i+5 for i,k in enumerate(lk)}; amt_of={k:float(a) for k,a in zip(lk,amt)}; fy_of=dict(zip(lk,fyv))

# ---------------------------------------------------------------- 1. EIL line-level site
eil_end=int(cfg['EIL_END'])
eil=cwb.get_sheet_by_name('Evidence_Invoice_Lines').to_python()
def line_site(desc):
    names={CANON.get(x.lower(),x) for x in GP.findall(desc)}
    if not names:
        dl=desc.lower()
        for t,lab in WATER.items():
            if t in dl: names.add(lab); break
    return sorted(names)
eil_site=[]; per_lk=collections.defaultdict(lambda:collections.Counter()); per_lk_res=collections.Counter(); per_lk_n=collections.Counter()
for r in eil[4:eil_end]:
    desc=str(r[2]) if r[2] not in (None,'') else ''
    k=str(r[9]).strip().replace('.0','') if r[9] not in (None,'') else ''
    a=float(r[6]) if isinstance(r[6],(int,float)) else 0.0
    names=line_site(desc)
    if len(names)==1: s,b=names[0],'Printed line names one park'
    elif len(names)>1: s,b='; '.join(names),'Printed line names several parks (held)'
    else: s,b='','No park named on the printed line'
    eil_site.append((s,b))
    if k:
        per_lk_n[k]+=1
        if len(names)==1: per_lk[k][names[0]]+=a
        else: per_lk_res[k]+=a
print('EIL lines classified:',collections.Counter(b for _,b in eil_site))

# ---------------------------------------------------------------- 2. Site_Allocation for multi rows
new_site=list(site_old); new_basis=list(basis_old)
alloc_rows=[]; n_alloc_lk=0; n_multi_held=0
for i in range(len(df)):
    if basis_old.iloc[i]!=MULTI: continue
    k=lk.iloc[i]
    if k not in per_lk_n or not per_lk[k]:
        n_multi_held+=1; continue                       # no printed lines tied, or none names a park -> stays held
    A=amt_of[k]; parts=dict(per_lk[k])
    allocated=round(sum(parts.values()),2)
    residue=round(A-allocated,2)
    for p,v in sorted(parts.items(),key=lambda x:-x[1]):
        alloc_rows.append([k,row_of[k],ref.iloc[i],fy_of[k],pkv.iloc[i],A,p,round(v,2),'Printed line amounts (Evidence_Invoice_Lines) summed per park'])
    if abs(residue)>=0.005:
        alloc_rows.append([k,row_of[k],ref.iloc[i],fy_of[k],pkv.iloc[i],A,RESIDUE,residue,'Register amount less the park-named printed lines'])
    new_basis[i]=MULTI_ALLOC; n_alloc_lk+=1
alloc_total=round(sum(r[7] for r in alloc_rows),2)
alloc_reg=round(sum(amt_of[k] for k in {r[0] for r in alloc_rows}),2)
assert abs(alloc_total-alloc_reg)<0.01,(alloc_total,alloc_reg)
print(f'Site_Allocation: {n_alloc_lk} register rows -> {len(alloc_rows)} allocation rows, ${alloc_total:,.2f}; multi still held {n_multi_held}')

# ---------------------------------------------------------------- 3. Electricity crosswalk (from elec_xwalk.py result)
R=pd.read_pickle('/home/claude/elec_res.pkl')
n_g=n_a=0; e_writes=[]
for _,r in R.iterrows():
    i=int(r.row)-5
    assert basis_old.iloc[i]==L_ADDR
    if r.flag=='GREEN': new_site[i]=r.park; new_basis[i]=E_GREEN; n_g+=1
    elif r.flag=='AMBER': new_site[i]=r.park; new_basis[i]=E_AMBER; n_a+=1
    else: continue
    e_writes.append(dict(row=int(r.row),LineKey=lk.iloc[i],amount=float(amt.iloc[i]),supply_text=r.text,park=r.park,flag=r.flag,basis=r.why))
n_red=int((R.flag=='RED').sum())
print(f'electricity: green {n_g}, amber {n_a}, red (kept in class) {n_red}')
# case-fold uniqueness of park labels
fold=collections.defaultdict(set)
for s,b in zip(new_site,new_basis):
    if s and b not in CLASS_LABELS and b!=MULTI_ALLOC: fold[s.lower()].add(s)
for rr in alloc_rows:
    if rr[6]!=RESIDUE: fold[rr[6].lower()].add(rr[6])
assert all(len(v)==1 for v in fold.values()),{k:v for k,v in fold.items() if len(v)>1}

# ---------------------------------------------------------------- 4. openpyxl writes
wb=openpyxl.load_workbook(SRC)
reg=wb['Register']; assert reg.max_row==21475
tmpl=copy.copy(reg.cell(5,42)._style); n_reg=0
for i in range(len(df)):
    if new_site[i]==site_old.iloc[i] and new_basis[i]==basis_old.iloc[i]: continue
    r=i+5
    a=reg.cell(r,SITE_COL); a.value=new_site[i] or None; a._style=copy.copy(tmpl)
    b=reg.cell(r,BASIS_COL); b.value=new_basis[i]; b._style=copy.copy(tmpl)
    n_reg+=1
print('register rows rewritten:',n_reg)
# EIL cols L, M
we=wb['Evidence_Invoice_Lines']
hs=copy.copy(we.cell(4,11)._style); ds=copy.copy(we.cell(6,11)._style)
we.cell(4,12,'Site (printed line)')._style=copy.copy(hs); we.cell(4,13,'Site basis')._style=copy.copy(hs)
we.column_dimensions['L'].width=34; we.column_dimensions['M'].width=38
for j,(s,b) in enumerate(eil_site):
    r=5+j
    c=we.cell(r,12); c.value=s or None; c._style=copy.copy(ds)
    c=we.cell(r,13); c.value=b; c._style=copy.copy(ds)
# Site_Allocation sheet
sa=wb.create_sheet('Site_Allocation', wb.sheetnames.index('Sites')+1)
sa.sheet_view.showGridLines=False
so=wb['Sites']; ttl=copy.copy(so.cell(1,1)._style); sub=copy.copy(so.cell(2,1)._style); hdr=copy.copy(so.cell(4,1)._style)
dat=copy.copy(so.cell(5,1)._style); mon=copy.copy(so.cell(5,3)._style)
def put(ws,r,c,v,st=None,fmt=None):
    cell=ws.cell(r,c); cell.value=v
    if st is not None: cell._style=copy.copy(st)
    if fmt: cell.number_format=fmt
    return cell
put(sa,1,1,'Site allocation of multi-site register lines, from printed invoice line amounts (v44)',ttl)
put(sa,2,1,'One row per (register LineKey, park). Where a register row\'s sighted invoice prints several sites, each printed line carries its own amount and its own site; those amounts are summed per park here. A residue row carries the register amount less the park-named printed lines, so every LineKey ties to Register!T to the cent by construction. This is rule 16 line-level evidence read at line level; nothing is apportioned. Register column '+BC+' on these rows reads "'+MULTI_ALLOC+'". Sites Panel A sums this sheet alongside the single-site register rows.',sub)
SA_H=['Register LineKey','Register row','Reference','FY (Doc)','PK charged','Register $ ex GST','Park / Site','Allocated $ ex GST','Basis']
for j,h in enumerate(SA_H,1): put(sa,4,j,h,hdr)
for j,w in enumerate([30,10,16,11,12,16,40,16,60],1): sa.column_dimensions[get_column_letter(j)].width=w
r=5
for rr in alloc_rows:
    for j,v in enumerate(rr,1):
        put(sa,r,j,v,mon if j in (6,8) else dat,'$#,##0.00;($#,##0.00);"-"' if j in (6,8) else None)
    r+=1
SA_END=r-1; SA_TOT=r+1
put(sa,SA_TOT,1,'TOTAL allocated (ties to the sum of the register amounts of the LineKeys above)',hdr)
put(sa,SA_TOT,8,f'=SUM(H5:H{SA_END})',hdr,'$#,##0.00;($#,##0.00);"-"')
put(sa,SA_TOT+1,1,'Tie check: SUMIF over Register column '+BC+' for the allocated-multi basis = total allocated',hdr)
put(sa,SA_TOT+1,8,f'=IF(ROUND(SUMIF(Register!${BC}$5:${BC}$21473,"{MULTI_ALLOC}",Register!$T$5:$T$21473),2)=ROUND(H{SA_TOT},2),"TRUE","FALSE")',hdr)
# ---- Sites Panel A: rebuild formulas to add Site_Allocation, add new gazetteer rows, add group row for residue
ws=so
# find group rows and total from config
g0,g1=[int(x) for x in str(cfg['SITES_GROUP_ROWS']).split(':')]; TOT_ROW=int(cfg['SITES_TOTAL_ROW']); CHK_ROW=int(cfg['SITES_CHECK_ROW'])
pb0,pb1=[int(x) for x in str(cfg['SITES_PANELB']).split(':')]
# capture panel B + group rows + notes, then rebuild from row 5
elec=[[ws.cell(r,c).value for c in range(1,5)] for r in range(pb0,pb1+1)]
elec_st=[[copy.copy(ws.cell(r,c)._style) for c in range(1,5)] for r in range(pb0,pb1+1)]
grp_notes={ws.cell(r,1).value:ws.cell(r,9).value for r in range(g0,g1+1)}
num_style=copy.copy(ws.cell(5,2)._style); mon_style=copy.copy(ws.cell(5,3)._style)
# park universe now = register single sites + allocation parks
allparks=set(s for s,b in zip(new_site,new_basis) if s and b not in CLASS_LABELS and b!=MULTI_ALLOC)|{rr[6] for rr in alloc_rows if rr[6]!=RESIDUE}
# order by total $
ptot=collections.Counter()
for s,b,a in zip(new_site,new_basis,amt):
    if s and b not in CLASS_LABELS and b!=MULTI_ALLOC: ptot[s]+=a
for rr in alloc_rows:
    if rr[6]!=RESIDUE: ptot[rr[6]]+=rr[7]
parks=sorted(allparks,key=lambda p:-ptot[p])
# per-park primary pk/basis
ppk=collections.defaultdict(collections.Counter); pbas=collections.defaultdict(collections.Counter)
for s,b,p in zip(new_site,new_basis,pkv):
    if s in allparks and b not in CLASS_LABELS and b!=MULTI_ALLOC: ppk[s][p]+=1; pbas[s][b]+=1
for rr in alloc_rows:
    if rr[6]!=RESIDUE: ppk[rr[6]][rr[4]]+=1; pbas[rr[6]]['Site_Allocation (printed line amounts)']+=1
# clear rows 4.. to end
for row in ws.iter_rows(min_row=4,max_row=ws.max_row):
    for c in row: c.value=None
ws.cell(2,1).value=str(ws.cell(2,1).value)+' v44: Panel A now sums BOTH the single-site register rows (Register!'+SC+') AND the per-park allocation of multi-site invoices from printed line amounts (Site_Allocation!G:H). Lines counts register rows only; the allocated multi rows are counted once on their group row. Electricity supply points matched to Council parcel address records are attributed (number+street exact) or attributed-and-flagged (street unique; basis carries [USER CHECK]).'
put(ws,3,1,'Panel A: cost per park / site (live; register single-site + Site_Allocation)',hdr)
HDRS=['Park / Site','Register lines (live)','$ ex GST (live)']+FYS+['Primary PK','Primary basis','Bases']
for j,h in enumerate(HDRS,1): put(ws,4,j,h,hdr)
r=5; PA_START=5
def sf(fy,label_cell):
    return (f'=SUMIFS(Register!$T$5:$T$21473,Register!${SC}$5:${SC}$21473,{label_cell},Register!$E$5:$E$21473,"{fy}")'
            f'+SUMIFS(Site_Allocation!$H$5:$H${SA_END},Site_Allocation!$G$5:$G${SA_END},{label_cell},Site_Allocation!$D$5:$D${SA_END},"{fy}")')
for p in parks:
    put(ws,r,1,p,dat)
    put(ws,r,2,f'=COUNTIF(Register!${SC}$5:${SC}$21473,$A{r})',num_style,'#,##0;-#,##0;"-"')
    put(ws,r,3,f'=SUMIF(Register!${SC}$5:${SC}$21473,$A{r},Register!$T$5:$T$21473)+SUMIF(Site_Allocation!$G$5:$G${SA_END},$A{r},Site_Allocation!$H$5:$H${SA_END})',mon_style,'$#,##0.00;($#,##0.00);"-"')
    for k,fy in enumerate(FYS): put(ws,r,4+k,sf(fy,f'$A{r}'),mon_style,'$#,##0.00;($#,##0.00);"-"')
    put(ws,r,8,max(ppk[p].items(),key=lambda x:x[1])[0] if ppk[p] else '',dat)
    put(ws,r,9,max(pbas[p].items(),key=lambda x:x[1])[0] if pbas[p] else '',dat)
    put(ws,r,10,len(pbas[p]),num_style,'#,##0')
    r+=1
PA_END=r-1
GRP_START=r
# group rows: residue (from Site_Allocation), then the classes on register basis
put(ws,r,1,RESIDUE,dat)
put(ws,r,2,f'=COUNTIF(Register!${BC}$5:${BC}$21473,"{MULTI_ALLOC}")',num_style,'#,##0;-#,##0;"-"')
put(ws,r,3,f'=SUMIF(Site_Allocation!$G$5:$G${SA_END},$A{r},Site_Allocation!$H$5:$H${SA_END})',mon_style,'$#,##0.00;($#,##0.00);"-"')
for k,fy in enumerate(FYS): put(ws,r,4+k,f'=SUMIFS(Site_Allocation!$H$5:$H${SA_END},Site_Allocation!$G$5:$G${SA_END},$A{r},Site_Allocation!$D$5:$D${SA_END},"{fy}")',mon_style,'$#,##0.00;($#,##0.00);"-"')
put(ws,r,9,f'{n_alloc_lk} multi-site register rows allocated per printed line; this row is only the un-named residue',dat); r+=1
GROUP_LABELS=[MULTI]+[c for c in CLASS_LABELS if c!=MULTI]
E_ROWS=[E_AMBER]   # amber electricity gets its own group so the check-flagged $ are visible; green counts as attributed
for lbl in GROUP_LABELS:
    put(ws,r,1,lbl,dat)
    put(ws,r,2,f'=COUNTIF(Register!${BC}$5:${BC}$21473,$A{r})',num_style,'#,##0;-#,##0;"-"')
    put(ws,r,3,f'=SUMIF(Register!${BC}$5:${BC}$21473,$A{r},Register!$T$5:$T$21473)',mon_style,'$#,##0.00;($#,##0.00);"-"')
    for k,fy in enumerate(FYS): put(ws,r,4+k,f'=SUMIFS(Register!$T$5:$T$21473,Register!${BC}$5:${BC}$21473,$A{r},Register!$E$5:$E$21473,"{fy}")',mon_style,'$#,##0.00;($#,##0.00);"-"')
    put(ws,r,9,grp_notes.get(lbl,''),dat); r+=1
GRP_END=r-1; TOT_ROW=r
put(ws,r,1,'TOTAL (must tie to the register)',hdr)
put(ws,r,2,f'=SUM(B{PA_START}:B{GRP_END})',hdr,'#,##0')
put(ws,r,3,f'=SUM(C{PA_START}:C{GRP_END})',hdr,'$#,##0.00;($#,##0.00);"-"')
for k in range(4,8): put(ws,r,k,f'=SUM({get_column_letter(k)}{PA_START}:{get_column_letter(k)}{GRP_END})',hdr,'$#,##0.00;($#,##0.00);"-"')
CHK_ROW=r+1
put(ws,CHK_ROW,1,'Tie check (lines = 21,469 and $ = Register!$T$21475)',hdr)
put(ws,CHK_ROW,2,f'=IF(B{TOT_ROW}=21469,"TRUE","FALSE")',hdr)
put(ws,CHK_ROW,3,f'=IF(ROUND(C{TOT_ROW},2)=ROUND(Register!$T$21475,2),"TRUE","FALSE")',hdr)
put(ws,CHK_ROW,4,f'=IF(ROUND(SUM(D{TOT_ROW}:G{TOT_ROW}),2)=ROUND(C{TOT_ROW},2),"TRUE","FALSE")',hdr)
put(ws,CHK_ROW+1,1,'Electricity flagged for user check (amber): lines / $',hdr)
put(ws,CHK_ROW+1,2,f'=COUNTIF(Register!${BC}$5:${BC}$21473,"{E_AMBER}")',hdr,'#,##0')
put(ws,CHK_ROW+1,3,f'=SUMIF(Register!${BC}$5:${BC}$21473,"{E_AMBER}",Register!$T$5:$T$21473)',hdr,'$#,##0.00;($#,##0.00);"-"')
PB_TITLE=CHK_ROW+3
put(ws,PB_TITLE,1,'Panel B: electricity by retailer account, all loaded years (relocated unchanged)',hdr)
for j,h in enumerate(['Retailer account','Site','Lines','$ ex GST (net, all years)'],1): put(ws,PB_TITLE+1,j,h,hdr)
r=PB_TITLE+2
for vals,sts in zip(elec,elec_st):
    for j,v in enumerate(vals,1): put(ws,r,j,v,sts[j-1])
    r+=1
PB_END=r-1
print(f'Sites: PanelA {PA_START}:{PA_END} ({len(parks)}), groups {GRP_START}:{GRP_END}, total {TOT_ROW}, chk {CHK_ROW}, PB {PB_TITLE+2}:{PB_END}')

# ---------------------------------------------------------------- 5. Config / Handover / Method / DA / Open_Items
cf=wb['Config']; crow={str(cf.cell(i,1).value):i for i in range(1,cf.max_row+1) if cf.cell(i,1).value}
cf.cell(crow['WORKBOOK_VERSION'],2).value='v44'
cf.cell(crow['SITES_PANELA'],2).value=f'{PA_START}:{PA_END}'; cf.cell(crow['SITES_GROUP_ROWS'],2).value=f'{GRP_START}:{GRP_END}'
cf.cell(crow['SITES_TOTAL_ROW'],2).value=TOT_ROW; cf.cell(crow['SITES_CHECK_ROW'],2).value=CHK_ROW
cf.cell(crow['SITES_PANELB'],2).value=f'{PB_TITLE+2}:{PB_END}'; cf.cell(crow['PARK_COUNT'],2).value=len(parks)
n_single=sum(1 for s,b in zip(new_site,new_basis) if s and b not in CLASS_LABELS and b!=MULTI_ALLOC)
cf.cell(crow['SITE_ATTRIBUTED_LINES'],2).value=n_single; cf.cell(crow['SITE_MULTI_LINES'],2).value=n_multi_held
cst_a=copy.copy(cf.cell(cf.max_row,1)._style); cst_b=copy.copy(cf.cell(cf.max_row,2)._style); nr=cf.max_row+1
for k,v in [('SITE_ALLOC_DATA',f'5:{SA_END}'),('SITE_ALLOC_TOTAL_ROW',SA_TOT),('SITE_ALLOC_LINEKEYS',n_alloc_lk),('SITE_ALLOC_ROWS',len(alloc_rows)),
            ('EIL_SITE_COLS','12:13 (L:M)'),('ELEC_ADDR_GREEN',n_g),('ELEC_ADDR_AMBER',n_a),('ELEC_ADDR_UNRESOLVED',n_red)]:
    a=cf.cell(nr,1); a.value=k; a._style=copy.copy(cst_a); b=cf.cell(nr,2); b.value=v; b._style=copy.copy(cst_b); nr+=1
h=wb['Handover']; hr=h.max_row+1; hs2=[copy.copy(h.cell(h.max_row,c)._style) for c in range(1,4)]
h.cell(hr,1).value=f'v44 change (18-Aug-2026): site dimension read at PRINTED-LINE level. Evidence_Invoice_Lines gains Site (L) and Site basis (M) per printed line; new sheet Site_Allocation explodes {n_alloc_lk} multi-site register rows into {len(alloc_rows)} (LineKey, park) rows on the printed line amounts (${alloc_total:,.2f}); Sites Panel A sums register single-site rows plus the allocation. Electricity supply points matched to Council parcel address records: {n_g} lines written exact, {n_a} written and flagged [USER CHECK], {n_red} unresolved. Gazetteer +3 water-park printed sites. Enrichment build; column T untouched; control total ${CT:,.2f}.'; h.cell(hr,1)._style=hs2[0]
h.cell(hr,2).value=f'{n_single} single-site lines on {len(parks)} parks; {n_multi_held} multi rows still held (no printed lines tied); allocation residue ${sum(r[7] for r in alloc_rows if r[6]==RESIDUE):,.2f}. Writes listed in b44_electricity_writes.xlsx (Open Items #135).'; h.cell(hr,2)._style=hs2[1]; h.cell(hr,3)._style=hs2[2]
m=wb['Method']; mr=m.max_row+1; ms=copy.copy(m.cell(m.max_row,1)._style)
METH=[f"43.0 Site read at printed-line level (v44). The v43 build read the site from the register row's line-items field (col 111), which prints EVERY line of the invoice, and so held any invoice naming several parks. But Evidence_Invoice_Lines already carries each printed line with its own amount, printed site and printed PK (rule 16). Where each printed line names one park, the invoice's split across sites is FIXED BY EVIDENCE and is written, not held: {n_alloc_lk} register rows explode to {len(alloc_rows)} (LineKey, park) rows on Site_Allocation, ${alloc_total:,.2f}, every LineKey tying to Register!T to the cent by construction (residue row = register amount less the park-named lines). Precedent: Pool Shop QLD INV-5278, one register line $37,850.00 on PK000022, fifteen printed lines - Logan Gardens Water Play $5,512.58 PK000469, Flagstone Water Play $6,309.56 PK000470, Darlington Parklands $4,836.00 PK000441, Beenleigh Water Play $3,334.82 PK000419, Forest Glen Water Feature $4,270.36. The Water Parks section becomes readable per site for the first time (Method 9 PoolShop cross-posting). Holding these at v43 was over-caution, not rigour: the v32 partial-scope principle says write where evidence fixes the line.",
 f"43.1 Two-level site model (v44). Level 1, Evidence_Invoice_Lines cols L/M: site per printed line, basis 'Printed line names one park' / 'several parks (held)' / 'No park named'. Level 2, Register {SC}/{BC}: single-site rows unchanged; multi-site rows with tied printed lines read '{MULTI_ALLOC}' and their dollars sit on Site_Allocation; multi-site rows with no tied printed lines stay '{MULTI}'. Sites Panel A = SUMIF(Register {SC}) + SUMIF(Site_Allocation G:H) per park; a group row carries the allocation residue; the tie check stays exhaustive.",
 f"43.2 Electricity supply points reconciled to site by Council address records (v44). Three internal address indexes: the S07 internal-rates parcel narrations (parcel id, street, park name - Council's own record), the Parks Strategy Layer House_Add, and the Parks Directory address. Rule: number+street exact and unique across the three -> written ({n_g} lines, basis '{E_GREEN}'); street unique but no number match -> written AND flagged ({n_a} lines, basis '{E_AMBER}', own group row on Sites, listed line by line for the user); street serving several parks, or not in any Council record -> stays '{L_ADDR}' ({n_red} lines; 99 Ewing Rd Woodridge alone is 217 lines and resolves to Booran / Rainbow / Willow Park). External geocoding stays REJECTED (Method 42.5); the unresolved rows await G-NAF property points or a Property_Key cadastre join.",
 f"43.3 Gazetteer additions (v44): Flagstone Water Play, Beenleigh Water Play, Forest Glen Water Feature - the water-park sites as PRINTED on the Pool Shop line items, absent from all three park datasets under those names. Underwood Water Feature -> Underwood Park; Logan Gardens Water Play -> Logan Gardens. Requeue: confirm the parent park of each water-play site with Parks (Flagstone Water Play sits in which named park?)."]
for i,t in enumerate(METH): c=m.cell(mr+i,1); c.value=t; c._style=copy.copy(ms)
d=wb['Data_Acquisition']; dr=d.max_row+1; ds2=copy.copy(d.cell(d.max_row,1)._style)
c=d.cell(dr,1); c.value=f'v44 update (18-Aug-2026): enrichment build; sighted count 1,877 and control total ${CT:,.2f} unchanged. Site read at printed-line level (Evidence_Invoice_Lines L:M, Site_Allocation {len(alloc_rows)} rows); electricity address crosswalk {n_g} exact / {n_a} flagged / {n_red} unresolved.'; c._style=copy.copy(ds2)
oi=wb['Open_Items']; orow=oi.max_row+1; os_=[copy.copy(oi.cell(oi.max_row,c)._style) for c in range(1,8)]
ITEMS=[(135,'All','Open','Electricity supply points',n_a,round(float(R[R.flag=="AMBER"].amt.sum()),2),f'ELECTRICITY SITE WRITES FLAGGED FOR USER CHECK. {n_a} lines attributed on a street-unique Council address record (no house-number match). Basis carries [USER CHECK]; listed in b44_electricity_writes.xlsx with the supply text, matched park and record source. Failures revert to the address-only class.'),
       (136,'All','Open','Electricity supply points',n_red,round(float(R[R.flag=="RED"].amt.sum()),2),'ELECTRICITY SUPPLY POINTS UNRESOLVED BY ADDRESS. Street serves several parks (99 Ewing Rd Woodridge 217 lines; Loganlea Rd 84; 77 Couldery Ct 61; 143 Meakin Rd 54) or the street is in no Council address record. Route: G-NAF property points -> point-in-polygon on LoganHub layer 121, or a Property_Key cadastre join. Ask GIS which parcel each NMI sits on.'),
       (137,'All','Open','Site dimension',n_multi_held,0,f'MULTI-SITE ROWS STILL HELD. {n_multi_held} register rows name several parks in a narration but have no printed invoice lines tied (unsighted). Resolve by capture.'),
       (138,'All','Open','Site_Allocation',0,round(float(sum(r[7] for r in alloc_rows if r[6]==RESIDUE)),2),'ALLOCATION RESIDUE. Printed lines on allocated multi-site invoices that name no park (chemicals, call-outs, freight) or name two. Sits on the residue group row; allocate only if the invoice itself apportions.'),
       (139,'All','Open','Water Parks',0,0,'WATER-PARK SITE NAMES. Flagstone Water Play, Beenleigh Water Play and Forest Glen Water Feature are gazetteered as printed; confirm the parent park name of each with Parks / Water Parks and reconcile PK000469/470/441/419 to sites.')]
for i,it in enumerate(ITEMS):
    for j,v in enumerate(it,1): c=oi.cell(orow+i,j); c.value=v; c._style=copy.copy(os_[j-1])

# ---------------------------------------------------------------- 6. save -> recalc -> verify -> ship
pre=f'{OUT_DIR}/{NEW}'; wb.save(pre); print('saved')
prof=f'{ROOT}/lo_profile/user'; os.makedirs(prof,exist_ok=True)
open(f'{prof}/registrymodifications.xcu','w').write(open('/home/claude/b43/lo_profile/user/registrymodifications.xcu').read())
t0=time.time()
res=subprocess.run(['soffice',f'-env:UserInstallation=file://{ROOT}/lo_profile','--headless','--calc','--convert-to','xlsx','--outdir',RECALC_DIR,pre],capture_output=True,text=True,timeout=1800)
final=f'{RECALC_DIR}/{NEW}'; assert os.path.exists(final),res.stdout+res.stderr; print(f'recalc {time.time()-t0:.0f}s')
json.dump(dict(final=final,CT=CT,PA=[PA_START,PA_END],GRP=[GRP_START,GRP_END],TOT_ROW=TOT_ROW,CHK_ROW=CHK_ROW,PB=[PB_TITLE+2,PB_END],parks=len(parks),
   single=n_single,multi_held=n_multi_held,alloc_lk=n_alloc_lk,alloc_rows=len(alloc_rows),alloc_total=alloc_total,SA_END=SA_END,SA_TOT=SA_TOT,
   green=n_g,amber=n_a,red=n_red,CLASS_LABELS=CLASS_LABELS,MULTI_ALLOC=MULTI_ALLOC,RESIDUE=RESIDUE,E_GREEN=E_GREEN,E_AMBER=E_AMBER,
   eil_end=eil_end,gb_count=int((df.iloc[:,87].notna()).sum())),open(f'{ROOT}/b44_pos.json','w'),indent=1)
pd.DataFrame(e_writes).sort_values(['flag','amount'],ascending=[True,False]).assign(decision='').to_excel(f'{ROOT}/b44_electricity_writes.xlsx',index=False)
r=subprocess.run([sys.executable,f'{ROOT}/b44_verify.py'],capture_output=True,text=True); print(r.stdout); print(r.stderr[-1500:])
if 'FAILS: none' in r.stdout:
    subprocess.run(['cp',final,f'/mnt/user-data/outputs/{NEW}']); subprocess.run(['cp',f'{ROOT}/b44_electricity_writes.xlsx','/mnt/user-data/outputs/']); print('SHIPPED',NEW)
else: print('NOT SHIPPED'); sys.exit(1)
