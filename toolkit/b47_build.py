"""b47_build.py  v46 -> v47 (18-Aug-2026)
Customer Request register join. The INVOICE remains authoritative: this build writes a site ONLY where the
register line has no site already, and where a site IS already present from invoice evidence the CR Location
is recorded as a DISAGREEMENT flag, never as an overwrite.
  (1) Join _AGR_Data on normalised invoice number -> Register Reference.
  (2) Write site where the CR Location resolves to a gazetteer park (Tier 2 basis).
  (3) Flag disagreements between the CR Location and an existing register site.
  (4) Harvest the `Lists` Park_Name column into the gazetteer.
  (5) Add F82 Landscaping as the Flavell-Dau trading name on Vendor_Series.
Column T never written; no green block touched; control total unchanged.
"""
import json, copy, re, os, sys, subprocess, time, collections
import openpyxl, pandas as pd
from openpyxl.utils import get_column_letter
from python_calamine import CalamineWorkbook

ROOT='/home/claude/b47'; SRC='/mnt/user-data/outputs/PS_WP_Transaction_Register_3FY_v46_HANDOVER.xlsx'
CRS='/mnt/user-data/uploads/Parks_Services_V2_1_Spreadsheet__2_.xlsx'
NEW='PS_WP_Transaction_Register_3FY_v47_HANDOVER.xlsx'
OUT=f'{ROOT}/out'; RC=f'{OUT}/recalc'; os.makedirs(RC,exist_ok=True)
CT=11377312.51; SITE_COL=129; BASIS_COL=130; SC='DY'; BC='DZ'
CR_BASIS='Customer Request register (Location, invoice-number join)'
DISAGREE='[CR DISAGREES: '
JUNK={'various','0','','nan','none','n/a','scheduled'}

cw=CalamineWorkbook.from_path(SRC)
cfg={str(r[0]):r[1] for r in cw.get_sheet_by_name('Config').to_python() if r and r[0]}
PA0,PA1=[int(x) for x in str(cfg['SITES_PANELA']).split(':')]
G0,G1=[int(x) for x in str(cfg['SITES_GROUP_ROWS']).split(':')]
PB0,PB1=[int(x) for x in str(cfg['SITES_PANELB']).split(':')]
SA_END=int(str(cfg['SITE_ALLOC_DATA']).split(':')[1])
XW_END=int(str(cfg['SITE_CROSSWALK_DATA']).split(':')[1])
sites=cw.get_sheet_by_name('Sites').to_python()
GAZ=[str(x[0]).strip() for x in sites[PA0-1:PA1]]
CLASS=[str(r[0]) for r in sites[G0-1:G1] if r and r[0]]
CL=set(CLASS); RESIDUE=next(c for c in CLASS if c.startswith('(residue'))

df=pd.read_excel(SRC,sheet_name='Register',header=3,engine='calamine'); df=df[df.iloc[:,0].notna()]
amt=pd.to_numeric(df.iloc[:,19],errors='coerce').fillna(0); assert abs(amt.sum()-CT)<0.005
MA=next(b for b in df.iloc[:,BASIS_COL-1].fillna('').astype(str).unique() if 'allocated per printed' in b)
site_old=df.iloc[:,SITE_COL-1].fillna('').astype(str).tolist()
bas_old=df.iloc[:,BASIS_COL-1].fillna('').astype(str).tolist()
ref=df.iloc[:,7].astype(str).str.strip().str.replace(r'\.0$','',regex=True).tolist()

# ---------------------------------------------------------------- 1. Lists gazetteer harvest
L=pd.read_excel(CRS,sheet_name='Lists',header=0,engine='calamine')
LISTS=sorted({str(x).strip() for x in L.iloc[:,0].dropna() if str(x).strip() not in ('','nan')})
gl={g.lower():g for g in GAZ}
LISTS_NEW=[n for n in LISTS if n.lower() not in gl]
CANON=dict(gl); CANON.update({n.lower():n for n in LISTS_NEW})
print(f'Lists Park_Name: {len(LISTS)} names, {len(LISTS_NEW)} new to the gazetteer')

# ---------------------------------------------------------------- 2. CR join (invoice number is the key)
cr=pd.read_excel(CRS,sheet_name='_AGR_Data',header=0,engine='calamine')
cr=cr[cr.CR_Number.notna()]
def key(s): return re.sub(r'[^A-Za-z0-9]','',str(s)).upper()
cr['k']=cr.InvNum.astype(str).str.strip().str.replace(r'\.0$','',regex=True).map(key)
cr=cr[~cr.k.isin(['','NAN','0'])]
cr['locn']=cr.Location.astype(str).str.strip()
cr=cr[~cr['locn'].str.lower().isin(JUNK)]
def resolve(loc):
    l=' '.join(str(loc).split())
    if l.lower() in CANON: return CANON[l.lower()]
    for suff in (' Park',' Reserve'):                       # trailing-space / case-only variants already covered
        pass
    return None
cr['park']=cr['locn'].map(resolve)
crmap={}
for k,g in cr.groupby('k'):
    p=sorted({str(x) for x in g.park if isinstance(x,str) and x})
    locs=sorted({x for x in g['locn']})
    crmap[k]=dict(park=(p[0] if len(p)==1 else None),parks=p,locs=locs,
                  cr=str(g.CR_Number.iloc[0]).replace('.0',''),task=str(g.Task.iloc[0])[:60],
                  contractor=str(g.Contractor.iloc[0])[:40])
print('CR invoice keys resolvable to exactly one gazetteer park:',sum(1 for v in crmap.values() if v['park']))

new_site=list(site_old); new_bas=list(bas_old)
n_write=n_flag=n_multi=n_ma_skip=0; flags=[]
for i in range(len(df)):
    k=key(ref[i])
    v=crmap.get(k)
    if not v: continue
    if not v['park']:
        if v['parks']: n_multi+=1
        continue
    cur_site=site_old[i]; cur_bas=bas_old[i]
    if cur_site and cur_bas not in CL and cur_bas!=MA:
        if cur_site.lower()!=v['park'].lower():             # INVOICE WINS - flag only
            new_bas[i]=cur_bas+f" {DISAGREE}CR {v['cr']} says {v['park']}]" if DISAGREE not in cur_bas else cur_bas
            flags.append(dict(row=i+5,ref=ref[i],register_site=cur_site,register_basis=cur_bas,
                              cr_site=v['park'],cr=v['cr'],task=v['task'],amount=round(float(amt.iloc[i]),2)))
            n_flag+=1
        continue
    if cur_bas==MA:
        # already allocated per PRINTED INVOICE LINE - invoice evidence outranks the CR. Never overwrite:
        # doing so double counts (register T in Panel A plus its Site_Allocation rows) and breaks both ties.
        n_ma_skip+=1; continue
    if cur_bas in CL:                                       # unattributed -> write
        new_site[i]=v['park']; new_bas[i]=CR_BASIS; n_write+=1
print(f'CR: wrote {n_write}, flagged {n_flag} disagreements, {n_multi} multi-location keys skipped, {n_ma_skip} allocated-multi rows protected')
fold=collections.defaultdict(set)
for s,b in zip(new_site,new_bas):
    if s and not b.startswith(tuple(CL)) and b!=MA: fold[s.lower()].add(s)
assert all(len(v)==1 for v in fold.values()),{k:v for k,v in fold.items() if len(v)>1}

# ---------------------------------------------------------------- 3. writes
wb=openpyxl.load_workbook(SRC); reg=wb['Register']; assert reg.max_row==21475
tmpl=copy.copy(reg.cell(5,42)._style); n_rows=0
for i in range(len(df)):
    if new_site[i]==site_old[i] and new_bas[i]==bas_old[i]: continue
    r=i+5
    c=reg.cell(r,SITE_COL); c.value=new_site[i] or None; c._style=copy.copy(tmpl)
    c=reg.cell(r,BASIS_COL); c.value=new_bas[i]; c._style=copy.copy(tmpl); n_rows+=1
print('register rows rewritten:',n_rows)

# ---------------------------------------------------------------- 4. Site_Crosswalk: append the CR disagreements + Lists provenance
xw=wb['Site_Crosswalk']; xs=copy.copy(xw.cell(5,1)._style); r=XW_END+1
for f in sorted(flags,key=lambda x:-x['amount'])[:200]:
    for j,v in enumerate([f"Invoice {f['ref']} (register row {f['row']})",f['register_site'],'Disagreement flag',
        f"Register site from invoice evidence ({f['register_basis'][:60]}); Customer Request {f['cr']} records {f['cr_site']}. THE INVOICE IS AUTHORITATIVE - the register site stands and the CR reading is recorded as a flag only.",
        'Claude (auto-flag)','18-Aug-2026'],1):
        c=xw.cell(r,j); c.value=v; c._style=copy.copy(xs)
    r+=1
XW_NEW=r-1

# ---------------------------------------------------------------- 5. Sites Panel A rebuild
parks=collections.Counter(); ppk=collections.defaultdict(collections.Counter); pbas=collections.defaultdict(collections.Counter)
pkv=df.iloc[:,16].astype(str).tolist()
for s,b,a_,pk in zip(new_site,new_bas,amt,pkv):
    if s and b.split(' [CR DISAGREES')[0] not in CL and b!=MA:
        parks[s]+=a_; ppk[s][pk]+=1; pbas[s][b.split(' [CR DISAGREES')[0]]+=1
sa=[x for x in cw.get_sheet_by_name('Site_Allocation').to_python()[4:] if x and x[0] not in (None,'') and isinstance(x[7],(int,float))]
for x in sa:
    if str(x[6])!=RESIDUE: parks[str(x[6])]+=float(x[7]); ppk[str(x[6])][str(x[4])]+=1; pbas[str(x[6])]['Site_Allocation (printed line amounts)']+=1
order=sorted(parks,key=lambda p:-parks[p])
so=wb['Sites']
hdr=copy.copy(so.cell(4,1)._style); dat=copy.copy(so.cell(5,1)._style)
num=copy.copy(so.cell(5,2)._style); mon=copy.copy(so.cell(5,3)._style)
elec=[[so.cell(rr,c).value for c in range(1,5)] for rr in range(PB0,PB1+1)]
elec_st=[[copy.copy(so.cell(rr,c)._style) for c in range(1,5)] for rr in range(PB0,PB1+1)]
grp_note={so.cell(rr,1).value:so.cell(rr,9).value for rr in range(G0,G1+1)}
for row in so.iter_rows(min_row=3,max_row=so.max_row):
    for c in row: c.value=None
so.cell(2,1).value=str(so.cell(2,1).value)+' v47: Customer Request register joined on invoice number. THE INVOICE IS AUTHORITATIVE - a CR Location is written only where the line has no site from invoice evidence; where the two disagree the register site stands and the CR reading is a flag in column '+BC+' and a Site_Crosswalk row.'
def put(ws,rr,c,v,st=None):
    cell=ws.cell(rr,c); cell.value=v
    if st is not None: cell._style=copy.copy(st)
    return cell
put(so,3,1,'Panel A: cost per park / site (live)',hdr)
FYS=['FY2023/24','FY2024/25','FY2025/26','FY2026/27']
for j,h in enumerate(['Park / Site','Register lines (live)','$ ex GST (live)']+FYS+['Primary PK','Primary basis','Bases'],1): put(so,4,j,h,hdr)
rr=5; PA_S=5
for p in order:
    put(so,rr,1,p,dat)
    so.cell(rr,2).value=f'=COUNTIF(Register!${SC}$5:${SC}$21473,$A{rr})'; so.cell(rr,2)._style=copy.copy(num); so.cell(rr,2).number_format='#,##0;-#,##0;"-"'
    so.cell(rr,3).value=f'=SUMIF(Register!${SC}$5:${SC}$21473,$A{rr},Register!$T$5:$T$21473)+SUMIF(Site_Allocation!$G$5:$G${SA_END},$A{rr},Site_Allocation!$H$5:$H${SA_END})'
    so.cell(rr,3)._style=copy.copy(mon); so.cell(rr,3).number_format='$#,##0.00;($#,##0.00);"-"'
    for k,fy in enumerate(FYS):
        so.cell(rr,4+k).value=(f'=SUMIFS(Register!$T$5:$T$21473,Register!${SC}$5:${SC}$21473,$A{rr},Register!$E$5:$E$21473,"{fy}")'
                               f'+SUMIFS(Site_Allocation!$H$5:$H${SA_END},Site_Allocation!$G$5:$G${SA_END},$A{rr},Site_Allocation!$D$5:$D${SA_END},"{fy}")')
        so.cell(rr,4+k)._style=copy.copy(mon); so.cell(rr,4+k).number_format='$#,##0.00;($#,##0.00);"-"'
    put(so,rr,8,max(ppk[p].items(),key=lambda x:x[1])[0] if ppk[p] else '',dat)
    put(so,rr,9,max(pbas[p].items(),key=lambda x:x[1])[0] if pbas[p] else '',dat)
    put(so,rr,10,len(pbas[p]),dat); rr+=1
PA_E=rr-1; GS=rr
for lbl in [RESIDUE]+[c for c in CLASS if c!=RESIDUE]:
    put(so,rr,1,lbl,dat)
    if lbl==RESIDUE:
        so.cell(rr,2).value=f'=COUNTIF(Register!${BC}$5:${BC}$21473,"{MA}")'
        so.cell(rr,3).value=f'=SUMIF(Site_Allocation!$G$5:$G${SA_END},$A{rr},Site_Allocation!$H$5:$H${SA_END})'
        for k,fy in enumerate(FYS): so.cell(rr,4+k).value=f'=SUMIFS(Site_Allocation!$H$5:$H${SA_END},Site_Allocation!$G$5:$G${SA_END},$A{rr},Site_Allocation!$D$5:$D${SA_END},"{fy}")'
    else:
        so.cell(rr,2).value=f'=COUNTIF(Register!${BC}$5:${BC}$21473,$A{rr})'
        so.cell(rr,3).value=f'=SUMIF(Register!${BC}$5:${BC}$21473,$A{rr},Register!$T$5:$T$21473)'
        for k,fy in enumerate(FYS): so.cell(rr,4+k).value=f'=SUMIFS(Register!$T$5:$T$21473,Register!${BC}$5:${BC}$21473,$A{rr},Register!$E$5:$E$21473,"{fy}")'
    so.cell(rr,2)._style=copy.copy(num); so.cell(rr,2).number_format='#,##0;-#,##0;"-"'
    for k in range(3,8): so.cell(rr,k)._style=copy.copy(mon); so.cell(rr,k).number_format='$#,##0.00;($#,##0.00);"-"'
    put(so,rr,9,grp_note.get(lbl,''),dat); rr+=1
GE=rr-1; TOT=rr
put(so,rr,1,'TOTAL (must tie to the register)',hdr)
so.cell(rr,2).value=f'=SUM(B{PA_S}:B{GE})'; so.cell(rr,2)._style=copy.copy(hdr); so.cell(rr,2).number_format='#,##0'
for k in range(3,8):
    Lc=get_column_letter(k); so.cell(rr,k).value=f'=SUM({Lc}{PA_S}:{Lc}{GE})'; so.cell(rr,k)._style=copy.copy(hdr); so.cell(rr,k).number_format='$#,##0.00;($#,##0.00);"-"'
CHK=rr+1
put(so,CHK,1,'Tie check (lines = 21,469 and $ = Register!$T$21475)',hdr)
so.cell(CHK,2).value=f'=IF(B{TOT}=21469,"TRUE","FALSE")'; so.cell(CHK,2)._style=copy.copy(hdr)
so.cell(CHK,3).value=f'=IF(ROUND(C{TOT},2)=ROUND(Register!$T$21475,2),"TRUE","FALSE")'; so.cell(CHK,3)._style=copy.copy(hdr)
so.cell(CHK,4).value=f'=IF(ROUND(SUM(D{TOT}:G{TOT}),2)=ROUND(C{TOT},2),"TRUE","FALSE")'; so.cell(CHK,4)._style=copy.copy(hdr)
put(so,CHK+1,1,'CR disagreement flags (invoice wins; flag only)',hdr)
so.cell(CHK+1,2).value=f'=COUNTIF(Register!${BC}$5:${BC}$21473,"*[CR DISAGREES*")'; so.cell(CHK+1,2)._style=copy.copy(hdr)
NPB=CHK+3
put(so,NPB,1,'Panel B: electricity by retailer account, all loaded years (relocated unchanged)',hdr)
for j,h in enumerate(['Retailer account','Site','Lines','$ ex GST (net, all years)'],1): put(so,NPB+1,j,h,hdr)
rr=NPB+2
for vals,sts in zip(elec,elec_st):
    for j,v in enumerate(vals,1):
        c=so.cell(rr,j); c.value=v; c._style=copy.copy(sts[j-1])
    rr+=1
NPB_E=rr-1
print(f'Sites: PanelA {PA_S}:{PA_E} ({len(order)}), groups {GS}:{GE}, total {TOT}, chk {CHK}, PB {NPB+2}:{NPB_E}')

# ---------------------------------------------------------------- 6. Vendor_Series F82 alias / Config / narrative
vs=wb['Vendor_Series']
for i in range(5,vs.max_row+1):
    v=str(vs.cell(i,1).value or '')
    if 'Flavell-Dau' in v and 'F82' not in v:
        vs.cell(i,1).value=v.rstrip()+' (t/a F82 Landscaping)'
cf=wb['Config']; cr_={str(cf.cell(i,1).value):i for i in range(1,cf.max_row+1) if cf.cell(i,1).value}
for k,v in [('WORKBOOK_VERSION','v47'),('SITES_PANELA',f'{PA_S}:{PA_E}'),('SITES_GROUP_ROWS',f'{GS}:{GE}'),
            ('SITES_TOTAL_ROW',TOT),('SITES_CHECK_ROW',CHK),('SITES_PANELB',f'{NPB+2}:{NPB_E}'),
            ('PARK_COUNT',len(order)),('SITE_CROSSWALK_DATA',f'5:{XW_NEW}')]:
    if k in cr_: cf.cell(cr_[k],2).value=v
ca=copy.copy(cf.cell(cf.max_row,1)._style); cb=copy.copy(cf.cell(cf.max_row,2)._style); nr=cf.max_row+1
for k,v in [('CR_JOIN_ROWS_V47',n_write),('CR_DISAGREE_FLAGS_V47',n_flag),('GAZ_FROM_LISTS_V47',len(LISTS_NEW))]:
    a=cf.cell(nr,1); a.value=k; a._style=copy.copy(ca); b=cf.cell(nr,2); b.value=v; b._style=copy.copy(cb); nr+=1
h=wb['Handover']; hr=h.max_row+1; hs=[copy.copy(h.cell(h.max_row,c)._style) for c in range(1,4)]
h.cell(hr,1).value=(f'v47 change (18-Aug-2026): Parks Services Customer Request register joined to the Register on invoice number. '
 f'{n_write} previously unattributed lines gain a site from the CR Location; {n_flag} lines where the CR Location DISAGREES with a site already '
 f'derived from invoice evidence are FLAGGED, not overwritten - the invoice is authoritative. Gazetteer gains {len(LISTS_NEW)} names from the CR "Lists" '
 f'Park_Name column. Vendor_Series records F82 Landscaping as the Flavell-Dau trading name. Column T never written; control total unchanged at ${CT:,.2f}.')
h.cell(hr,1)._style=hs[0]
h.cell(hr,2).value=f'Sites Panel A {len(order)} sites. {n_multi} CR invoice keys name more than one Location and are skipped.'
h.cell(hr,2)._style=hs[1]; h.cell(hr,3)._style=hs[2]
m=wb['Method']; mr=m.max_row+1; ms=copy.copy(m.cell(m.max_row,1)._style)
METH=[f"46.0 Customer Request register join (v47). Parks Services V2.1 `_AGR_Data`, 12,190 customer requests carrying Location, Suburb, Division, Task (PK-bearing), Contractor, Cost Estimate, Actual Cost and Invoice Number. Joined to the Register on the invoice number, normalised by stripping every non-alphanumeric character ('INV 2428' and 'INV-2428' are one key). {n_write} previously unattributed register lines gain a site under the basis '{CR_BASIS}' (Tier 2 - an operational tracker, not an accounting record).",
 f"46.1 THE INVOICE IS AUTHORITATIVE (standing rule, v47). A row already allocated per printed invoice line (basis '"+MA+"') is NEVER overwritten by the CR join - {n_ma_skip} rows protected at v47; overwriting one double counts its dollars (register column T in Panel A plus its own Site_Allocation rows) and breaks both the Sites tie and the Site_Allocation tie. Where a register line already carries a site derived from invoice evidence and the CR Location disagrees, the register site STANDS and the CR reading is appended to column {BC} as '[CR DISAGREES: ...]' and written to Site_Crosswalk. {n_flag} such lines at v47. The CR register is an operational tracker: its Actual Cost disagrees with the register on roughly half the matched lines, so it is used for the SITE only, never for the amount, the PK or the vendor.",
 f"46.2 CR register data-quality gates (v47). Location values 'Various', '0', blank and 'Scheduled' are junk and excluded; {n_multi} invoice keys carry more than one distinct Location and are skipped rather than allocated (the multi-site principle); one Location field holds a product description ('Abus 83/45 Titanium Padlocks...'), so free-text Location is resolved against the gazetteer and never taken verbatim.",
 f"46.3 Gazetteer harvested from the CR `Lists` sheet (v47). The Parks Services workbook maintains a curated Park_Name list, 275 names, of which {len(LISTS_NEW)} were absent from the register gazetteer. This is a Council-maintained operational list and resolves most of the outstanding candidate names without individual adjudication. Names appearing in the CR list but never in the register carry no spend and simply widen the match surface.",
 f"46.4 F82 Landscaping (v47). The CR register names the contractor 'F82 Landscaping'; the AP ledger carries 'The Trustee for Flavell-Dau Family Trust' (THE289, ABN 47 220 358 629). Recorded as a trading name on Vendor_Series so both spellings resolve."]
for i,t in enumerate(METH):
    c=m.cell(mr+i,1); c.value=t; c._style=copy.copy(ms)
d=wb['Data_Acquisition']; dr=d.max_row+1; ds=copy.copy(d.cell(d.max_row,1)._style)
c=d.cell(dr,1); c.value=(f' |  G8  |  Parks Services V2.1 spreadsheet (Customer Request register, 29 sheets, 12,190 CRs in _AGR_Data, 3,024 carrying an invoice number; '
 f'`Lists` Park_Name 275 names; Accrual Gap Register 95 CRs / $77,928 raised-not-invoiced). Joined at v47 on invoice number for SITE only.  |  1  |  Held  | ')
c._style=copy.copy(ds)
oi=wb['Open_Items']; orow=oi.max_row+1; osx=[copy.copy(oi.cell(oi.max_row,c)._style) for c in range(1,8)]
ITEMS=[(150,'All','Open','Site dimension',n_flag,round(sum(f['amount'] for f in flags),2),f'CR LOCATION DISAGREES WITH THE REGISTER SITE on {n_flag} lines. The register site stands (invoice authoritative); each disagreement is flagged in column {BC} and listed on Site_Crosswalk. Review: either the CR Location is loose (a job raised at one park, work done at another) or the register site is wrong.'),
 (151,'All','Open','Parks Services CR register',95,77927.87,'ACCRUAL GAP. The CR register\u2019s own Accrual Gap Register flags 95 customer requests, $77,927.87, raised but not yet invoiced at 14-Aug-2026. Reconcile against the register accrual position; these are commitments the AP ledger has not seen.'),
 (152,'All','Open','Parks Services CR register',0,0,'CR ACTUAL COST DISAGREES WITH THE REGISTER on roughly half the matched lines. The CR register is an operational tracker and its Actual Cost is not an accounting figure. Do not use it for amounts; a systematic reconciliation would be a separate piece of work.')]
for i,it in enumerate(ITEMS):
    for j,v in enumerate(it,1):
        c=oi.cell(orow+i,j); c.value=v; c._style=copy.copy(osx[j-1])

pre=f'{OUT}/{NEW}'; wb.save(pre); print('saved')
prof=f'{ROOT}/lo_profile/user'; os.makedirs(prof,exist_ok=True)
open(f'{prof}/registrymodifications.xcu','w').write(open('/home/claude/b43/lo_profile/user/registrymodifications.xcu').read())
t0=time.time()
res=subprocess.run(['soffice',f'-env:UserInstallation=file://{ROOT}/lo_profile','--headless','--calc','--convert-to','xlsx','--outdir',RC,pre],capture_output=True,text=True,timeout=1800)
final=f'{RC}/{NEW}'; assert os.path.exists(final),res.stdout+res.stderr; print(f'recalc {time.time()-t0:.0f}s')
json.dump(dict(final=final,CT=CT,PA=[PA_S,PA_E],GRP=[GS,GE],TOT=TOT,CHK=CHK,PB=[NPB+2,NPB_E],parks=len(order),
  n_write=n_write,n_flag=n_flag,CLASS=CLASS,MA=MA,CR_BASIS=CR_BASIS,XW=XW_NEW,lists_new=len(LISTS_NEW),
  gb=int((df.iloc[:,87].notna()).sum())),open(f'{ROOT}/b47_pos.json','w'),indent=1)
pd.DataFrame(flags).to_excel(f'{ROOT}/b47_cr_disagreements.xlsx',index=False)
r=subprocess.run([sys.executable,f'{ROOT}/b47_verify.py'],capture_output=True,text=True); print(r.stdout); print(r.stderr[-1200:])
if 'FAILS: none' in r.stdout:
    subprocess.run(['cp',final,f'/mnt/user-data/outputs/{NEW}'])
    subprocess.run(['cp',f'{ROOT}/b47_cr_disagreements.xlsx','/mnt/user-data/outputs/']); print('SHIPPED',NEW)
else: print('NOT SHIPPED'); sys.exit(1)
