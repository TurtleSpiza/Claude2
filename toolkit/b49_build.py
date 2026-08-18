"""b49_build.py  v48 -> v49: apply the user's web-verified address-to-park decisions (578 rows) via Site_Crosswalk."""
import json, copy, re, os, sys, subprocess, time, collections
import openpyxl, pandas as pd
from openpyxl.utils import get_column_letter
from python_calamine import CalamineWorkbook
ROOT='/home/claude/b49'; SRC='/mnt/user-data/outputs/PS_WP_Transaction_Register_3FY_v48_HANDOVER.xlsx'
NEW='PS_WP_Transaction_Register_3FY_v49_HANDOVER.xlsx'; OUT=f'{ROOT}/out'; RC=f'{OUT}/recalc'; os.makedirs(RC,exist_ok=True)
CT=11377312.51; SITE=129; BASIS=130; SC='DY'; BC='DZ'
d=pd.read_pickle('/home/claude/user_addr.pkl')
cw=CalamineWorkbook.from_path(SRC)
cfg={str(r[0]):r[1] for r in cw.get_sheet_by_name('Config').to_python() if r and r[0]}
PA0,PA1=[int(x) for x in str(cfg['SITES_PANELA']).split(':')]; G0,G1=[int(x) for x in str(cfg['SITES_GROUP_ROWS']).split(':')]
PB0,PB1=[int(x) for x in str(cfg['SITES_PANELB']).split(':')]; SA_END=int(str(cfg['SITE_ALLOC_DATA']).split(':')[1])
XW_END=int(str(cfg['SITE_CROSSWALK_DATA']).split(':')[1])
sites=cw.get_sheet_by_name('Sites').to_python(); GAZ=[str(x[0]).strip() for x in sites[PA0-1:PA1]]
CLASS=[str(r[0]) for r in sites[G0-1:G1] if r and r[0]]; CL=set(CLASS); RESIDUE=next(c for c in CLASS if c.startswith('(residue'))
df=pd.read_excel(SRC,sheet_name='Register',header=3,engine='calamine'); df=df[df.iloc[:,0].notna()]
amt=pd.to_numeric(df.iloc[:,19],errors='coerce').fillna(0); assert abs(amt.sum()-CT)<0.005
MA=next(b for b in df.iloc[:,BASIS-1].fillna('').astype(str).unique() if 'allocated per printed' in b)
site_old=df.iloc[:,SITE-1].fillna('').astype(str).tolist(); bas_old=df.iloc[:,BASIS-1].fillna('').astype(str).tolist()
narr=df.iloc[:,21].fillna('').astype(str).str.replace('\n',' ',regex=False)
seg=narr.str.split('-').str[-1].str.strip()
canon={g.lower():g for g in GAZ}
BAS_V='User-adjudicated address decision (web-verified, Site_Crosswalk)'
BAS_FAC='Non-park Council facility (named site, not a park)'
BAS_NP='Non-park Council parcel (user-adjudicated: not a park)'
res=d[~d.dec.str.startswith('UNRESOLVED')].copy()
res['park']=res.dec.map(lambda x: canon.get(str(x).lower(),str(x)))
new_site=list(site_old); new_bas=list(bas_old); n_w=n_np=n_skip=0; applied=[]
segmap={str(r.seg).strip().lower():r for _,r in res.iterrows()}
for i in range(len(df)):
    b=bas_old[i]
    if b==MA or (b not in CL): continue                       # never overwrite evidence or allocation
    key=seg.iloc[i].strip().lower()
    r=segmap.get(key)
    if r is None: continue
    if str(r.dec).startswith('NOT A PARK'):
        new_site[i]=''; new_bas[i]=BAS_NP; n_np+=1; continue   # class rows carry NO DY label or Panel A double counts
    fac=any(k in str(r.park).lower() for k in ('complex','community centre','oval','courts','club'))
    new_site[i]=r.park; new_bas[i]=BAS_FAC if fac else BAS_V; n_w+=1
print(f'user decisions applied: {n_w} park writes, {n_np} not-a-park')
fold=collections.defaultdict(set)
for s,b in zip(new_site,new_bas):
    if s and ';' not in s and b not in CL and b!=MA: fold[s.lower()].add(s)
assert all(len(v)==1 for v in fold.values()),{k:v for k,v in fold.items() if len(v)>1}
wb=openpyxl.load_workbook(SRC); reg=wb['Register']; tmpl=copy.copy(reg.cell(5,42)._style); nrw=0
for i in range(len(df)):
    if new_site[i]==site_old[i] and new_bas[i]==bas_old[i]: continue
    c=reg.cell(i+5,SITE); c.value=new_site[i] or None; c._style=copy.copy(tmpl)
    c=reg.cell(i+5,BASIS); c.value=new_bas[i]; c._style=copy.copy(tmpl); nrw+=1
print('rows rewritten',nrw)
# crosswalk: every one of the 578 decisions recorded (resolved and unresolved) so it never gets re-asked
xw=wb['Site_Crosswalk']; xs=copy.copy(xw.cell(5,1)._style); r=XW_END+1
for _,x in d.iterrows():
    typ='Park' if not str(x.dec).startswith(('UNRESOLVED','NOT A PARK')) else ('Not a park' if str(x.dec).startswith('NOT') else 'Unresolved (public search)')
    for j,v in enumerate([str(x.seg)[:120],str(x.dec)[:80],typ,f"{str(x.status)[:140]} | source: {str(x.source)[:120]}",'User (web-verified)','18-Aug-2026'],1):
        c=xw.cell(r,j); c.value=v; c._style=copy.copy(xs)
    r+=1
XW_NEW=r-1
# Sites rebuild (same shape as v48)
parks=collections.Counter(); ppk=collections.defaultdict(collections.Counter); pbas=collections.defaultdict(collections.Counter)
pkv=df.iloc[:,16].astype(str).tolist()
for s,b,a_,pk in zip(new_site,new_bas,amt,pkv):
    if s and ';' not in s and b not in CL and b!=MA and b!='Multiple sites named (not allocated)': parks[s]+=a_; ppk[s][pk]+=1; pbas[s][b]+=1
sa=[x for x in cw.get_sheet_by_name('Site_Allocation').to_python()[4:] if x and x[0] not in (None,'') and isinstance(x[7],(int,float))]
for x in sa:
    if str(x[6])!=RESIDUE: parks[str(x[6])]+=float(x[7]); ppk[str(x[6])][str(x[4])]+=1; pbas[str(x[6])]['Site_Allocation (printed line amounts)']+=1
order=sorted(parks,key=lambda p:-parks[p])
NEWCLASS=CLASS+([BAS_NP] if BAS_NP not in CLASS else [])
so=wb['Sites']; hdr=copy.copy(so.cell(4,1)._style); dat=copy.copy(so.cell(5,1)._style); num=copy.copy(so.cell(5,2)._style); mon=copy.copy(so.cell(5,3)._style)
elec=[[so.cell(rr,c).value for c in range(1,5)] for rr in range(PB0,PB1+1)]; elec_st=[[copy.copy(so.cell(rr,c)._style) for c in range(1,5)] for rr in range(PB0,PB1+1)]
gnote={so.cell(rr,1).value:so.cell(rr,9).value for rr in range(G0,G1+1)}
for row in so.iter_rows(min_row=3,max_row=so.max_row):
    for c in row: c.value=None
so.cell(2,1).value=str(so.cell(2,1).value)+' v49: 578 user web-verified address decisions applied and recorded on Site_Crosswalk (110 resolved to a site, 1 not a park, 467 unresolved and now permanently marked so).'
def put(ws,rr,c,v,st=None):
    cell=ws.cell(rr,c); cell.value=v
    if st is not None: cell._style=copy.copy(st)
FYS=['FY2023/24','FY2024/25','FY2025/26','FY2026/27']
put(so,3,1,'Panel A: cost per park / site (live)',hdr)
for j,h in enumerate(['Park / Site','Register lines (live)','$ ex GST (live)']+FYS+['Primary PK','Primary basis','Bases'],1): put(so,4,j,h,hdr)
rr=5; PA_S=5
for p in order:
    put(so,rr,1,p,dat)
    so.cell(rr,2).value=f'=COUNTIF(Register!${SC}$5:${SC}$21473,$A{rr})'; so.cell(rr,2)._style=copy.copy(num); so.cell(rr,2).number_format='#,##0;-#,##0;"-"'
    so.cell(rr,3).value=f'=SUMIF(Register!${SC}$5:${SC}$21473,$A{rr},Register!$T$5:$T$21473)+SUMIF(Site_Allocation!$G$5:$G${SA_END},$A{rr},Site_Allocation!$H$5:$H${SA_END})'
    so.cell(rr,3)._style=copy.copy(mon); so.cell(rr,3).number_format='$#,##0.00;($#,##0.00);"-"'
    for k,fy in enumerate(FYS):
        so.cell(rr,4+k).value=(f'=SUMIFS(Register!$T$5:$T$21473,Register!${SC}$5:${SC}$21473,$A{rr},Register!$E$5:$E$21473,"{fy}")+SUMIFS(Site_Allocation!$H$5:$H${SA_END},Site_Allocation!$G$5:$G${SA_END},$A{rr},Site_Allocation!$D$5:$D${SA_END},"{fy}")')
        so.cell(rr,4+k)._style=copy.copy(mon); so.cell(rr,4+k).number_format='$#,##0.00;($#,##0.00);"-"'
    put(so,rr,8,max(ppk[p].items(),key=lambda x:x[1])[0] if ppk[p] else '',dat); put(so,rr,9,max(pbas[p].items(),key=lambda x:x[1])[0] if pbas[p] else '',dat); put(so,rr,10,len(pbas[p]),dat); rr+=1
PA_E=rr-1; GS=rr
for lbl in [RESIDUE]+[c for c in NEWCLASS if c!=RESIDUE]:
    put(so,rr,1,lbl,dat)
    if lbl==RESIDUE:
        so.cell(rr,2).value=f'=COUNTIF(Register!${BC}$5:${BC}$21473,"{MA}")'; so.cell(rr,3).value=f'=SUMIF(Site_Allocation!$G$5:$G${SA_END},$A{rr},Site_Allocation!$H$5:$H${SA_END})'
        for k,fy in enumerate(FYS): so.cell(rr,4+k).value=f'=SUMIFS(Site_Allocation!$H$5:$H${SA_END},Site_Allocation!$G$5:$G${SA_END},$A{rr},Site_Allocation!$D$5:$D${SA_END},"{fy}")'
    else:
        so.cell(rr,2).value=f'=COUNTIF(Register!${BC}$5:${BC}$21473,$A{rr})'; so.cell(rr,3).value=f'=SUMIF(Register!${BC}$5:${BC}$21473,$A{rr},Register!$T$5:$T$21473)'
        for k,fy in enumerate(FYS): so.cell(rr,4+k).value=f'=SUMIFS(Register!$T$5:$T$21473,Register!${BC}$5:${BC}$21473,$A{rr},Register!$E$5:$E$21473,"{fy}")'
    so.cell(rr,2)._style=copy.copy(num); so.cell(rr,2).number_format='#,##0;-#,##0;"-"'
    for k in range(3,8): so.cell(rr,k)._style=copy.copy(mon); so.cell(rr,k).number_format='$#,##0.00;($#,##0.00);"-"'
    put(so,rr,9,gnote.get(lbl,'110 Evans Road Meadowbrook: user web check identifies a Council-rated parcel that is not a park' if lbl==BAS_NP else ''),dat); rr+=1
GE=rr-1; TOT=rr
put(so,rr,1,'TOTAL (must tie to the register)',hdr); so.cell(rr,2).value=f'=SUM(B{PA_S}:B{GE})'; so.cell(rr,2)._style=copy.copy(hdr); so.cell(rr,2).number_format='#,##0'
for k in range(3,8):
    L=get_column_letter(k); so.cell(rr,k).value=f'=SUM({L}{PA_S}:{L}{GE})'; so.cell(rr,k)._style=copy.copy(hdr); so.cell(rr,k).number_format='$#,##0.00;($#,##0.00);"-"'
CHK=rr+1
put(so,CHK,1,'Tie check (lines = 21,469 and $ = Register!$T$21475)',hdr)
so.cell(CHK,2).value=f'=IF(B{TOT}=21469,"TRUE","FALSE")'; so.cell(CHK,2)._style=copy.copy(hdr)
so.cell(CHK,3).value=f'=IF(ROUND(C{TOT},2)=ROUND(Register!$T$21475,2),"TRUE","FALSE")'; so.cell(CHK,3)._style=copy.copy(hdr)
so.cell(CHK,4).value=f'=IF(ROUND(SUM(D{TOT}:G{TOT}),2)=ROUND(C{TOT},2),"TRUE","FALSE")'; so.cell(CHK,4)._style=copy.copy(hdr)
NPB=CHK+2; put(so,NPB,1,'Panel B: electricity by retailer account (relocated unchanged)',hdr)
for j,h in enumerate(['Retailer account','Site','Lines','$ ex GST (net, all years)'],1): put(so,NPB+1,j,h,hdr)
rr=NPB+2
for vals,sts in zip(elec,elec_st):
    for j,v in enumerate(vals,1):
        c=so.cell(rr,j); c.value=v; c._style=copy.copy(sts[j-1])
    rr+=1
NPB_E=rr-1
cf=wb['Config']; cr={str(cf.cell(i,1).value):i for i in range(1,cf.max_row+1) if cf.cell(i,1).value}
for k,v in [('WORKBOOK_VERSION','v49'),('SITES_PANELA',f'{PA_S}:{PA_E}'),('SITES_GROUP_ROWS',f'{GS}:{GE}'),('SITES_TOTAL_ROW',TOT),('SITES_CHECK_ROW',CHK),('SITES_PANELB',f'{NPB+2}:{NPB_E}'),('PARK_COUNT',len(order)),('SITE_CROSSWALK_DATA',f'5:{XW_NEW}')]:
    if k in cr: cf.cell(cr[k],2).value=v
ca=copy.copy(cf.cell(cf.max_row,1)._style); cb=copy.copy(cf.cell(cf.max_row,2)._style); nr=cf.max_row+1
for k,v in [('USER_ADDR_DECISIONS_V49',len(d)),('USER_ADDR_WRITES_V49',n_w)]:
    a=cf.cell(nr,1); a.value=k; a._style=copy.copy(ca); b=cf.cell(nr,2); b.value=v; b._style=copy.copy(cb); nr+=1
h=wb['Handover']; hr=h.max_row+1; hs=[copy.copy(h.cell(h.max_row,c)._style) for c in range(1,4)]
h.cell(hr,1).value=(f'v49 change (18-Aug-2026): {len(d)} user web-verified address-to-park decisions applied through Site_Crosswalk. {n_w} register lines gain a site (VERIFIED 5 / SUPPORTED 104 / PROBABLE 1 segments); '
 f'110 Evans Road ({n_np} lines) recorded as a Council parcel that is not a park; 467 segments recorded as UNRESOLVED on public search so they are never re-asked. Control total unchanged at ${CT:,.2f}.')
h.cell(hr,1)._style=hs[0]; h.cell(hr,2).value=f'Sites Panel A {len(order)} sites.'; h.cell(hr,2)._style=hs[1]; h.cell(hr,3)._style=hs[2]
m=wb['Method']; mr=m.max_row+1; ms=copy.copy(m.cell(m.max_row,1)._style)
c=m.cell(mr,1); c.value=(f"48.0 User web-verified address decisions (v49). The 578 address-to-park questions were adjudicated by the user against public sources (Council directory, LPS2015 ArcGIS, property records) with a status per row: VERIFIED (5, official Council source), SUPPORTED (104, unique directory match), PROBABLE (1), AMBIGUOUS (28, more than one park on the street), NO UNIQUE PUBLIC MATCH (439), NOT A PARK (1: 90-110 Evans Road Meadowbrook, a Council-rated parcel - note Open Item #123 sits on the same parcel). "
 f"Every decision is written to Site_Crosswalk with its status and source URL, so it outranks the matcher permanently and unresolved rows are never re-asked. Only rows with no site from evidence were written; allocated-multi rows untouched."); c._style=copy.copy(ms)
oi=wb['Open_Items']; orow=oi.max_row+1; osx=[copy.copy(oi.cell(oi.max_row,c)._style) for c in range(1,8)]
for j,v in enumerate([155,'All','Open','Site dimension',467,round(float(d[d.dec.str.startswith('UNRESOLVED')].dollars.sum()),2),'ADDRESS SEGMENTS UNRESOLVED ON PUBLIC SEARCH. 439 no unique public match + 28 ambiguous streets. Route now: Origin split statements (site named per NMI) for the electricity rows; G-NAF property points or Council cadastre for the rest. Recorded on Site_Crosswalk so they are not re-asked.'],1):
    c=oi.cell(orow,j); c.value=v; c._style=copy.copy(osx[j-1])
pre=f'{OUT}/{NEW}'; wb.save(pre); print('saved')
prof=f'{ROOT}/lo_profile/user'; os.makedirs(prof,exist_ok=True)
open(f'{prof}/registrymodifications.xcu','w').write(open('/home/claude/b43/lo_profile/user/registrymodifications.xcu').read())
t0=time.time(); res_=subprocess.run(['soffice',f'-env:UserInstallation=file://{ROOT}/lo_profile','--headless','--calc','--convert-to','xlsx','--outdir',RC,pre],capture_output=True,text=True,timeout=1800)
final=f'{RC}/{NEW}'; assert os.path.exists(final),res_.stdout+res_.stderr; print(f'recalc {time.time()-t0:.0f}s')
json.dump(dict(final=final,CT=CT,PA=[PA_S,PA_E],GRP=[GS,GE],TOT=TOT,CHK=CHK,PB=[NPB+2,NPB_E],parks=len(order),n_w=n_w,CLASS=NEWCLASS,MA=MA,gb=int((df.iloc[:,87].notna()).sum()),suburbs=json.load(open('/home/claude/b48/b48_pos.json'))['suburbs']),open(f'{ROOT}/b49_pos.json','w'),indent=1)
r=subprocess.run([sys.executable,'/home/claude/b48/b48_verify.py'],capture_output=True,text=True,env={**os.environ,'B48_POS':f'{ROOT}/b49_pos.json'}); print(r.stdout[-2500:]); print(r.stderr[-800:])
if 'FAILS: none' in r.stdout:
    subprocess.run(['cp',final,f'/mnt/user-data/outputs/{NEW}']); print('SHIPPED',NEW)
else: print('NOT SHIPPED'); sys.exit(1)
