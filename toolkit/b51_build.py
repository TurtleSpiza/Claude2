"""b51_build.py  v50 -> v51  (18-Aug-2026)
Batch 42: Q Power historical corpus (qpower_historical_20260818, 51 invoices): 48 green blocks + 3 rule-16-only captures.
Batch 43: Origin collective statements A-FE35E476-022..-027 reconciled (register groups vs statement sub-accounts).
Repair: 13 v45 Q Power green blocks carried the PLEASE PAY BY date as the invoice date (F1) - restated from the retained Batch 40 text.
Sites: 30 user-adjudicated candidate names (spelling variants -> gazetteer), 17 Peacock Avenue = Logan River Parklands (Logan River Park label merged),
71 Cinderella Drive = Springwood Park (already applied, recorded), Q Power printed sites written; Sites Panel A rebuilt.
Column T never written; control total unchanged. One save -> one convert-route recalc -> one exact-TRUE verify -> ship.
"""
import json, copy, re, os, sys, subprocess, time, collections, datetime
import openpyxl
from openpyxl.utils import get_column_letter
from python_calamine import CalamineWorkbook

ROOT='/home/claude/b51'; SRC='/mnt/user-data/uploads/PS_WP_Transaction_Register_3FY_v50_HANDOVER.xlsx'
NEW='PS_WP_Transaction_Register_3FY_v51_HANDOVER.xlsx'; OUT=f'{ROOT}/out'; RC=f'{OUT}/recalc'; os.makedirs(RC,exist_ok=True)
CT=11377312.51; SITE=129; BASIS=130; SC='DY'; BC='DZ'; VER='v51'; TODAY='18-Aug-2026'
P=json.load(open(f'{ROOT}/b51_parse.json')); F=P['facts']; L=P['lines']; M=P['match']; UNM=P['unmatched']; REP=P['repair']
E=json.load(open(f'{ROOT}/b51_elec.json'))
cw=CalamineWorkbook.from_path(SRC)
cfg={str(r[0]):r[1] for r in cw.get_sheet_by_name('Config').to_python() if r and r[0]}
EI0,EI1=[int(x) for x in str(cfg['EI_DATA']).split(':')]; EI_TOT=int(cfg['EI_TOTALS_ROW'])
C0,C1=[int(x) for x in str(cfg['EI_CONTROL_ROWS']).split(':')]
EIL0,EIL_END=5,int(cfg['EIL_END']); R0,R1=[int(x) for x in str(cfg['RECON_ROWS']).split(':')]
SIGHT=int(cfg['SIGHTED_COUNT']); RECN=int(cfg['RECON_COUNT'])
PA0,PA1=[int(x) for x in str(cfg['SITES_PANELA']).split(':')]; G0,G1=[int(x) for x in str(cfg['SITES_GROUP_ROWS']).split(':')]
PB0,PB1=[int(x) for x in str(cfg['SITES_PANELB']).split(':')]; SA_END=int(str(cfg['SITE_ALLOC_DATA']).split(':')[1]); XW_END=int(str(cfg['SITE_CROSSWALK_DATA']).split(':')[1])
sites=cw.get_sheet_by_name('Sites').to_python(); GAZ=[str(x[0]).strip() for x in sites[PA0-1:PA1]]
CLASS=[str(r[0]) for r in sites[G0-1:G1] if r and r[0]]; RESIDUE=next(c for c in CLASS if c.startswith('(residue')); MULTI='Multiple sites named (not allocated)'
MA='Multiple sites named (allocated per printed line, see Site_Allocation)'; DEPOT='Parks Depot, Marsden (non-park facility)'
HARVEST='Invoice sighted, printed site not in gazetteer (candidate harvest)'
SUBURBS={s.lower() for s in json.load(open('/home/claude/up/left/toolkit/b48_suburbs.json'))}-{'regents park'}
data=cw.get_sheet_by_name('Register').to_python()[4:21473]; assert len(data)==21469
INV=sorted(F, key=lambda k:int(k)); MATCHED=[k for k in INV if k in M]; assert len(MATCHED)+len(UNM)==51
EVID={k:k for k in INV}                                   # 5-digit Q Power ids stay as printed (schema section 5, v45 convention)
NEW_EIL=sum(len(L[k]) for k in INV); NEW_EI=len(INV); NEW_GB=len(MATCHED)
NEW_EI_END=EI1+NEW_EI; NEW_EI_TOT=EI_TOT+NEW_EI; NEW_C0=C0+NEW_EI; NEW_C1=C1+NEW_EI
NEW_EIL_END=EIL_END+NEW_EIL; NEW_R0=R0+NEW_EIL; NEW_R1=R1+NEW_EIL+NEW_EI; NEW_SIGHT=SIGHT+NEW_GB; NEW_RECN=RECN+NEW_EI
OI_UNM=158
def dt(iso): y,m,d=iso.split('-'); return datetime.datetime(int(y),int(m),int(d))
def condense(s): s=re.sub(r'[ \t]{3,}','  ',s); s=re.sub(r'\n{3,}','\n\n',s); return s.replace('\f','')[:32000]

wb=openpyxl.load_workbook(SRC)
reg=wb['Register']; assert reg.max_row==21475
we=wb['Evidence_Invoices']; wl=wb['Evidence_Invoice_Lines']
def sty(c,t): c._style=copy.copy(t._style)

# ================= 1. EIL: capture old recon block, clear, write new lines, relocate recon
old_recon=[[wl.cell(r,c).value for c in range(1,7)] for r in range(R0,R1+1)]
old_recon_st=[[copy.copy(wl.cell(r,c)._style) for c in range(1,7)] for r in range(R0,R1+1)]
rec_hdr=[[wl.cell(R0-1,c).value for c in range(1,7)],[copy.copy(wl.cell(R0-1,c)._style) for c in range(1,7)]]
for r in range(R0-1,R1+1):
    for c in range(1,7): wl.cell(r,c).value=None
tmplL=[wl.cell(EIL_END,c) for c in range(1,14)]
r=EIL_END+1
for k in INV:
    f=F[k]; site1=f['gaz_sites'][0] if len(f['gaz_sites'])==1 else ''
    for o in L[k]:
        note=(f"Batch 42 capture; printed line {o['n']} of {len(L[k])}." if o['src']=='TEXT' else
              "Batch 42 capture; TEMPLATE PRINTS NO LINE TABLE (Q Power work-note template): single-scope line = the printed Description block, amount = the printed Sub-Total ex GST (BLC precedent, v40).")
        if len(o['desc_rows'])>1 and o['src']=='TEXT': note+=f" Description printed over {len(o['desc_rows'])} layout rows (F12), joined ' | ', every row asserted verbatim in the retained page text."
        if k in UNM: note+=f" No register line (rule 16 only, Open Items #{OI_UNM}); LineKey blank."
        vals=[EVID[k],o['n'],o['desc'],(f"{o['qty']:.2f}" if o['qty'] is not None else '(not printed)'),
              (o['unit'] if o['unit'] is not None else '(not printed)'),o['gst'],o['amt'],
              (f['pk'] or '(not printed)'),(f['pk'] if f['pk'] and re.match(r'^PK\d{6,7}$',f['pk']) else ''),
              (M[k]['target_lk'] if k in M else ''),note,site1,
              ('Printed line names one park' if site1 else 'No park named on the printed line')]
        for c,v in enumerate(vals,1):
            cell=wl.cell(r,c); cell.value=v; sty(cell,tmplL[c-1])
        r+=1
assert r-1==NEW_EIL_END,(r-1,NEW_EIL_END)
rr=NEW_R0-1
for c,(v,s) in enumerate(zip(rec_hdr[0],rec_hdr[1]),1):
    cell=wl.cell(rr,c); cell.value=v; cell._style=copy.copy(s)
rr=NEW_R0; recon_row={}
for src_r,(vals,sts) in enumerate(zip(old_recon,old_recon_st),R0):
    for c,(v,s) in enumerate(zip(vals,sts),1):
        if isinstance(v,str) and v.startswith('='):
            v=re.sub(r'\$A\$5:\$A\$%d\b'%EIL_END,f'$A$5:$A${NEW_EIL_END}',v)
            v=re.sub(r'\$G\$5:\$G\$%d\b'%EIL_END,f'$G$5:$G${NEW_EIL_END}',v)
            v=re.sub(r'\bD%d\b'%src_r,f'D{rr}',v); v=re.sub(r'\bC%d\b'%src_r,f'C{rr}',v)
        cell=wl.cell(rr,c); cell.value=v; cell._style=copy.copy(s)
    if vals[2]: recon_row[str(vals[2])]=rr
    rr+=1
tmplR=[copy.copy(wl.cell(NEW_R0,c)._style) for c in range(1,7)]
for k in INV:
    tgt=F[k]['sub']
    for c,v in enumerate(['','',EVID[k],f'=ROUND(SUMIF($A${EIL0}:$A${NEW_EIL_END},C{rr},$G${EIL0}:$G${NEW_EIL_END}),2)',f'=IF(D{rr}={tgt},"TRUE","FALSE")',''],1):
        cell=wl.cell(rr,c); cell.value=(v if v!='' else None); cell._style=copy.copy(tmplR[c-1])
    recon_row[EVID[k]]=rr; rr+=1
assert rr-1==NEW_R1,(rr-1,NEW_R1)
hv=wl.cell(NEW_R0-1,3).value
if isinstance(hv,str): wl.cell(NEW_R0-1,3).value=re.sub(r'\(v\d+ block; all ranges full: \$A\$5:\$A\$\d+\)',f'({VER} block; all ranges full: $A$5:$A${NEW_EIL_END})',hv)

# ================= 2. Evidence_Invoices append (tail relocation)
tail=[[we.cell(r,c).value for c in range(1,33)] for r in range(EI_TOT,C1+1)]
tail_st=[[copy.copy(we.cell(r,c)._style) for c in range(1,33)] for r in range(EI_TOT,C1+1)]
for r in range(EI_TOT,C1+1):
    for c in range(1,33): we.cell(r,c).value=None
tmplE=[we.cell(EI1,c) for c in range(1,33)]
BATCH_LBL='Batch 42 (qpower_historical_20260818: Q Power.pdf, 118 pages, 51 invoices)'
def bp_terms(k): return 'BP:QP-T2' if F[k]['svc']=='Service - Quoted' else 'BP:QP-T1'
r=EI1+1
for k in INV:
    f=F[k]
    site_txt=f['site'] or '(not printed)'; cq=(f['order'] or '(not printed)')
    if k in M:
        ver=f"Lines reconcile to the printed subtotal to the cent; reconciliation at Evidence_Invoice_Lines row {recon_row[EVID[k]]}. Rule 17 checks live on Register row {M[k]['target_row']}."
        vals=[EVID[k],f['vendor'],f['abn'],dt(f['inv_iso']),dt(f['due_iso']),k,cq,site_txt,len(L[k]),f['sub'],f['gst'],f['tot'],ver,
              'Complete on the register line (green block, columns CJ to DW)']+['']*18
    else:
        why={'14406':'site Jimboomba Soccer Club, Kurrajong Road (order 709263 - MS016398): the APLEDGER QPO001 history shows the invoice paid 29-May-2025 ($457.84 incl GST) but no AP line carries the reference or amount inside 4090240/4090260 - posts outside the PS/WP scope',
             '14865':'site Station Road, Woodridge (order 710663): paid 27-Feb-2026 per the QPO001 history ($374.00 incl GST); no AP line on reference, amount or date inside the loaded scope',
             '15222':'dated 12-Aug-2026, paid per the QPO001 history on 12-Aug-2026 ($5,860.73 incl GST) - later than the loaded FY2026/27 extract (27SLACT P1-P2)'}[k]
        ver=f"Lines reconcile to the printed subtotal to the cent; reconciliation at Evidence_Invoice_Lines row {recon_row[EVID[k]]}. NO register line matched - held on the evidence sheets (Open Items #{OI_UNM}): {why}. Register row (no register match)."
        anom=('Template prints no line table; single-scope line tied to the printed Sub-Total ex GST. ' if f['tie']!='TIE' else '')+(f"Corpus invoice_date {f['corpus_date']} restated to the printed INVOICE DATE {f['inv_lbl']} (F1). " if f['f1_date'] else '')+(f"Corpus notes: {f['notes']}." if f['notes'] else '')
        vals=[EVID[k],f['vendor'],f['abn'],dt(f['inv_iso']),dt(f['due_iso']),f'(no register match; Open Items #{OI_UNM})',cq,site_txt,len(L[k]),f['sub'],f['gst'],f['tot'],ver,
              '8/27 Allgas Street, Slacks Creek QLD 4127','Tel. 07 3155 4225 | qpower.com.au | Licence # 69807','(not printed)','Logan City Council, PO Box 3226, Logan City DC QLD 4114','(not printed)',
              (f['officer'] or '(not printed)'),'(not printed)','(not printed)',(f['cr'] or '(not printed)'),'(not printed)',(' + '.join(f['techs']) if f['techs'] else '(not printed)'),
              (f['works'] or '(not printed)'),'(not printed)','(not printed)','BP:QP-P1',bp_terms(k),f['pages'],condense(anom),
              f"Staged on this sheet (columns N to AE); no register line exists for this invoice (rule 16 capture only, Open Items #{OI_UNM}). {BATCH_LBL}, captured {TODAY} ({VER})"]
    for c,v in enumerate(vals,1):
        cell=we.cell(r,c); cell.value=(v if v!='' else None); sty(cell,tmplE[c-1])
    r+=1
assert r-1==NEW_EI_END
for i,(vals,sts) in enumerate(zip(tail,tail_st)):
    for c,(v,s) in enumerate(zip(vals,sts),1):
        cell=we.cell(NEW_EI_TOT+i,c); cell.value=v; cell._style=copy.copy(s)
for c,col in ((9,'I'),(10,'J'),(11,'K'),(12,'L')): we.cell(NEW_EI_TOT,c).value=f'=SUM({col}{EI0}:{col}{NEW_EI_END})'
for r in range(NEW_C0,NEW_C1+1):
    for c in range(1,33):
        v=we.cell(r,c).value
        if isinstance(v,str) and v.startswith('='):
            v=re.sub(r'\$A\$5:\$A\$%d\b'%EIL_END,f'$A$5:$A${NEW_EIL_END}',v); v=re.sub(r'\$G\$5:\$G\$%d\b'%EIL_END,f'$G$5:$G${NEW_EIL_END}',v)
            v=re.sub(r'\$E\$%d:\$E\$%d\b'%(R0,R1),f'$E${NEW_R0}:$E${NEW_R1}',v); v=re.sub(r'\$C\$%d:\$C\$%d\b'%(R0,R1),f'$C${NEW_R0}:$C${NEW_R1}',v)
            v=re.sub(r'\$A\$5:\$A\$%d\b'%EI1,f'$A$5:$A${NEW_EI_END}',v)
            v=re.sub(r'\bJ%d\b'%EI_TOT,f'J{NEW_EI_TOT}',v); v=re.sub(r'\bI%d\b'%EI_TOT,f'I{NEW_EI_TOT}',v)
            for oc in range(C0,C1+1): v=re.sub(r'\bC%d\b'%oc,f'C{oc+NEW_EI}',v)
            v=re.sub(r'\$AF\$5:\$AF\$%d\b'%(EI_TOT-1),f'$AF$5:$AF${NEW_EI_TOT-1}',v)
            v=v.replace(f'={SIGHT},',f'={NEW_SIGHT},').replace(f'={RECN},',f'={NEW_RECN},')
            we.cell(r,c).value=v
        elif isinstance(v,str):
            v=re.sub(r'must be [\d,]+',f'must be {NEW_SIGHT:,}',v); v=re.sub(r'the count is [\d,]+',f'the count is {NEW_SIGHT:,}',v)
            v=re.sub(r'all TRUE, [\d,]+ of [\d,]+',f'all TRUE, {NEW_RECN:,} of {NEW_RECN:,}',v); we.cell(r,c).value=v

# ================= 3. Register green blocks + analysis columns + site
tmplG=[reg.cell(14937,c) for c in range(88,128)]
tmplS=copy.copy(reg.cell(5,42)._style)
gb_rows=[]; pk_div=[]; noPK=[]; attach=[]
site_written=collections.Counter()
for k in MATCHED:
    f=F[k]; m=M[k]; r=m['target_row']; assert reg.cell(r,88).value in (None,''), r
    items=' '.join(f"{o['n']}. {o['desc']} | {(f'{o['qty']:.2f}' if o['qty'] is not None else '-')} @ ${(f'{o['unit']:.2f}' if o['unit'] is not None else '-')} = ${o['amt']:.2f}." for o in L[k])
    anom=[]
    if f['tie']!='TIE': anom.append('TEMPLATE PRINTS NO LINE TABLE (Q Power work-note template, Aug-2024 to Jun-2025): the invoice prints a Description / Work Note narrative and a totals block only, no priced item grid (confirmed on the OCR layer as well as the text layer). Captured as ONE single-scope line = the printed Description block at the printed Sub-Total ex GST (BLC precedent, v40); rule 16 basis stated here.')
    if f['f1_date']: anom.append(f"Corpus invoice_date {f['corpus_date']} was the work-note or pay-by date; Invoice Date restated to the printed INVOICE DATE {f['inv_lbl']} and Due Date to the printed PLEASE PAY BY {f['due_lbl']} from the header row (F1, fourth form); the register Doc Date agrees.")
    if not f['pk']: anom.append('No PK printed on the invoice face; PK left as (not printed), the register posting stands.'); noPK.append(k)
    elif not re.match(r'^PK\d{6,7}$',f['pk']): anom.append(f"Printed PK '{f['pk']}' is a 5-digit typo (not normalisable from print alone); the register posts {m['pk']} - recorded, not normalised (Open Items #157).")
    elif f['pk']!=m['pk']: anom.append(f"PRINTED PK {f['pk']} DIVERGES from the register charge {m['pk']} (Open Items #157)."); pk_div.append((k,f['pk'],m['pk']))
    if f['notes'] and 'refers to' in f['notes']: attach.append(k)
    if len(f['desc'] or '')>0 and 'PK' in (f['desc'] or '') and f['pk'] is None: pass
    if 'Site: Doug Larsen Park / Beenleigh' in (f['site'] or ''): pass
    if len(f['gaz_sites'])==1 and f['gaz_sites'][0]=='Flindersia Riverside Park': anom.append("Printed site 'Flindersia Riverside Park' added to the gazetteer as printed (v44 precedent); possible same site as the rates-narration label 'Logan Reserve Riverside Park' - Open Items #162.")
    if f['notes'] and 'refers to' in f['notes']: anom.append('Corpus note: '+f['notes'].split(';')[0].strip()+'.' if 'refers to' in f['notes'].split(';')[0] else 'Corpus note: '+f['notes']+'.')
    vals={88:EVID[k],89:f['vendor'],90:f['abn'],91:'(not printed)',92:'8/27 Allgas Street, Slacks Creek QLD 4127',93:'Tel. 07 3155 4225 | qpower.com.au | Licence # 69807',
          94:dt(f['inv_iso']),95:dt(f['due_iso']),96:(f['order'] or '(not printed)'),97:'(not printed)',98:'Logan City Council, PO Box 3226, Logan City DC QLD 4114',99:'(not printed)',
          100:(f['officer'] or '(not printed)'),101:'(not printed)',102:'(not printed)',103:(f['cr'] or '(not printed)'),104:(f['pk'] or '(not printed)'),
          105:(f['pk'] if f['pk'] and re.match(r'^PK\d{6,7}$',f['pk']) else (f"{f['pk']} (as printed; 5-digit typo, not normalisable)" if f['pk'] else '(not printed)')),
          106:(f['site'] or '(not printed)'),107:'(not printed)',108:(f['desc'] or '(not printed)'),109:(' + '.join(f['techs']) if f['techs'] else '(not printed)'),
          110:len(L[k]),111:condense(items),112:condense(f['works'] or '(not printed)'),113:f['sub'],114:f['gst'],115:f['tot'],116:'(not printed)',117:'(not printed)',
          118:'BP:QP-P1',119:bp_terms(k),120:f['pages'],
          121:f'=ROUND(SUMIF(Evidence_Invoice_Lines!$A${EIL0}:$A${NEW_EIL_END},CJ{r},Evidence_Invoice_Lines!$G${EIL0}:$G${NEW_EIL_END}),2)',
          122:f'=IF(DQ{r}=DI{r},"TRUE","FALSE")',123:f'=IF(ROUND(T{r},2)=DI{r},"TRUE","FALSE")',124:f'=IF(DJ{r}=ROUND(DI{r}*0.1,2),"TRUE","FALSE")',
          125:f"{BATCH_LBL} ({f['pages']})",126:f'{TODAY}, rule 17 capture ({VER})',127:condense(' '.join(anom))}
    for c,v in vals.items():
        cell=reg.cell(r,c); cell.value=v; sty(cell,tmplG[c-88])
    reg.cell(r,24).value='Electrical & data services'                    # rule 4: must exist in Theme_Map A5:A31
    reg.cell(r,25).value=('; '.join(o['desc'] for o in L[k]) if f['tie']=='TIE' else (f['desc'] or ''))[:250]
    reg.cell(r,27).value='Sighted invoice line'; reg.cell(r,29).value='Tier 1'; reg.cell(r,33).value='Confirmed'; reg.cell(r,34).value=None
    reg.cell(r,32).value=f"Invoice sighted (rule 17): Q Power (Qld) Pty Ltd, {f['site'] or 'site not printed'}; printed subtotal ${f['sub']:,.2f}; {f['svc'] or 'work-note template (no line table, single-scope line)'}."
    reg.cell(r,28).value=(str(reg.cell(r,28).value or '').rstrip()+f" Sighted tax invoice {k}, Q Power (Qld) Pty Ltd, ABN 82 067 507 591; line capture Evidence_Invoice_Lines; reconciliation row {recon_row[EVID[k]]}.").strip()
    gb_rows.append(r)
    if len(f['gaz_sites'])==1:
        reg.cell(r,SITE).value=f['gaz_sites'][0]; reg.cell(r,BASIS).value='Sighted invoice site (printed)'; site_written[f['gaz_sites'][0]]+=1
    elif f['site']:
        reg.cell(r,SITE).value=None; reg.cell(r,BASIS).value=HARVEST
    for c in (SITE,BASIS): reg.cell(r,c)._style=copy.copy(tmplS)

# ================= 4. v45 date repair (13 green blocks: Invoice Date held the PLEASE PAY BY date)
rep_rows=[]
evrow={str(reg.cell(r,88).value):r for r in range(5,21474) if reg.cell(r,88).value not in (None,'')}
eirow={str(we.cell(r,1).value):r for r in range(EI0,NEW_EI_END+1) if we.cell(r,1).value not in (None,'')}
for k,d in REP.items():
    r=evrow[k]; old=reg.cell(r,94).value
    oldd=old.date() if hasattr(old,'date') else datetime.datetime.strptime(str(old),'%d-%b-%Y').date()
    assert oldd.isoformat()==d['due_iso'] and abs((oldd-dt(d['inv_iso']).date()).days)==14, (k, old)
    assert str(reg.cell(r,4).value)[:10]==d['inv_iso'], (k,'register Doc Date corroboration')
    reg.cell(r,94).value=dt(d['inv_iso']); reg.cell(r,95).value=dt(d['due_iso'])
    reg.cell(r,127).value=condense((reg.cell(r,127).value or '')+f" [repair {VER}] Invoice Date restated from {d['corpus_date']} to the printed INVOICE DATE {d['inv_lbl']} and Due Date written as the printed PLEASE PAY BY {d['due_lbl']}, both from the retained Batch 40 page text (the corpus invoice_date held the pay-by date, F1); previously the due date was recorded as (not printed).")
    er=eirow[k]; we.cell(er,4).value=dt(d['inv_iso']); we.cell(er,5).value=dt(d['due_iso']); rep_rows.append(r)
assert len(rep_rows)==13
# ---- rule 19.10: rewrite BOTH citation limbs from the authoritative maps (EvID -> recon row; EvID -> register rows carrying the green block)
evrows=collections.defaultdict(list)
for rr_ in range(5,21474):
    if reg.cell(rr_,88).value not in (None,''): evrows[str(reg.cell(rr_,88).value)].append(rr_)
n_cit_ei=n_cit_reg=0
for er in range(EI0,NEW_EI_END+1):
    e=str(we.cell(er,1).value or '').strip()
    if not e or e not in recon_row: continue
    v=str(we.cell(er,13).value or ''); nv=v
    if re.search(r'Evidence_Invoice_Lines row \d+',nv): nv=re.sub(r'Evidence_Invoice_Lines row \d+',f'Evidence_Invoice_Lines row {recon_row[e]}',nv)
    elif 'Evidence_Invoice_Lines reconciliation block' in nv: nv=nv.replace('Evidence_Invoice_Lines reconciliation block',f'Evidence_Invoice_Lines row {recon_row[e]}')
    else: nv=(nv.rstrip()+f' Reconciliation at Evidence_Invoice_Lines row {recon_row[e]}.').strip()
    if e in evrows:
        rows_txt=(str(evrows[e][0]) if len(evrows[e])==1 else ', '.join(str(x) for x in evrows[e][:-1])+' and '+str(evrows[e][-1]))
        nv=re.sub(r'([Rr]egister row)s? [\d, and]+(?=[.;,)]|$)',lambda m_: f'{m_.group(1)}{"s" if len(evrows[e])>1 else ""} {rows_txt}',nv)
        if not re.search(r'[Rr]egister row',nv): nv=nv.rstrip()+f' Rule 17 checks live on Register row {rows_txt}.'
    if nv!=v: we.cell(er,13).value=nv; n_cit_ei+=1
for e,rows_ in evrows.items():
    if e not in recon_row: continue
    for rr_ in rows_:
        v=str(reg.cell(rr_,28).value or ''); nv=re.sub(r'((?:reconciliation|recon) row) \d+',lambda m_: f'{m_.group(1)} {recon_row[e]}',v)
        if not re.search(r'(reconciliation|recon) row \d+',nv): nv=(nv.rstrip()+f' Reconciliation row {recon_row[e]}.').strip()
        if nv!=v: reg.cell(rr_,28).value=nv; n_cit_reg+=1
print(f'citations rewritten: EI {n_cit_ei}, register {n_cit_reg}')

# ================= 5. Sites: user decisions (30 candidate names + 2 addresses), alias canonicalisation, Q Power sites, Panel A rebuild
ALIAS={'jamie nicholson park':'Jamie Nicolson Park','alexandra park':'Alexander Park','darlington park':'Darlington Parklands','logan gardens park':'Logan Gardens',
 'springwood dog off leash park':'Springwood Park','harley street shailer park':'Harley St Shailer Park','lavell park':'Lavelle Park','richared wilson park':'Richard Wilson Park',
 'woodland district park':'Woodlands District Park','doug larson park':'Doug Larsen Park','teviot down park':'Teviot Downs Park','tervit down park':'Teviot Downs Park',
 'oliver sports complex':'Olivers Sports Complex',"oliver's sport complex":'Olivers Sports Complex',"oliver's sports complex":'Olivers Sports Complex','vandalism olivers sports complex':'Olivers Sports Complex',
 'colville st green':'Colville Park','skinner park':'Skinners Park','forest glen park':'Forestglen Park','clean up cecil clark park':'Cecil Clark Park','hubener park':'Hubner Park',
 'shaw st oval':'Shaw Street Oval','jamie nicholson park lavelle park':'Jamie Nicolson Park; Lavelle Park','mt warren oval park':'Mt Warren Oval Park','logan reserve':'Logan Reserve Park',
 'sport complex':'Olivers Sports Complex','logan garden park':'Logan Gardens','james smith rec park':'James Smith Recreation Area','logan river park':'Logan River Parklands'}
SELF=['Logan Gardens','Alexander Park','Crestmead Park','Lincoln Green Park','Skinners Park']   # user "Add/Merge": already gazetteer labels, confirmed
GUARD={'logan reserve','sport complex'}   # generic tokens: alias only when the column names no other park (and, for Logan Reserve, no "Logan Reserve Rd")
gaz=[g for g in GAZ if g.lower() not in SUBURBS and g!='Logan River Park']; canon={g.lower():g for g in gaz}
for a in ['Flindersia Riverside Park','Mt Warren Oval Park','Logan Reserve Park','Cecil Clark Park','Richard Wilson Park','James Smith Recreation Area']: canon.setdefault(a.lower(),a)
for k,v in ALIAS.items():
    if ';' not in v: canon.setdefault(v.lower(),v)
UNIVERSE=sorted(set(canon.values())); assert not [u for u in UNIVERSE if u.lower() in SUBURBS]
NAMES=sorted(set(list(canon.keys())+list(ALIAS.keys())),key=len,reverse=True)
GP=re.compile('('+'|'.join(re.escape(n) for n in NAMES if len(n)>5)+')',re.I)
ALP=re.compile('('+'|'.join(re.escape(n) for n in sorted(ALIAS,key=len,reverse=True))+')',re.I)
EMPTY={'','nan','none','(not printed)','(blank as printed)','(no site named)','0','various','various sites','various parks','scheduled','n/a'}
def clean(v):
    if v is None: return ''
    s=' '.join(str(v).replace('\n',' ').split()); s=re.sub(r'(\w)-(\w)',r'\1\2',s)
    return '' if s.strip().lower() in EMPTY else s
def resolve(tok):
    t=tok.lower()
    if t in ALIAS: return [x.strip() for x in ALIAS[t].split(';')]
    return [canon[t]] if t in canon else []
def parks_in(text):
    hits=[(m.start(),m.group(1)) for m in GP.finditer(text)]
    out=[]; toks=set()
    for pos,tk in hits:
        pre=text[:pos]
        if pre and re.search(r'[A-Za-z]$',pre): continue                 # inside a longer word/name
        for p in resolve(tk): out.append(p); toks.add(tk.lower())
    out=sorted(set(out))
    if 'Logan Reserve Park' in out and 'logan reserve' in toks and 'logan reserve park' not in toks and not re.match(r'^logan reserve\s*(\(|-|$)',text.strip().lower()): out.remove('Logan Reserve Park')   # suburb usage ("X, Logan Reserve") is never a park
    out=[p for p in out if 'depot' not in p.lower()]
    if 'Olivers Sports Complex' in out and 'sport complex' in toks and not re.search(r'oliver',text,re.I) and 'olivers sports complex' not in toks: out.remove('Olivers Sports Complex')
    return out
BANNED={12,25,32,92,98,105,SITE,BASIS}; COLS=[c for c in range(1,131) if c not in BANNED]
PRIORITY={111:1,106:2,112:3,108:4,127:5,99:6,42:7,22:8,23:9,80:10}
LBL={111:'Sighted invoice line items (printed)',106:'Sighted invoice site (printed)',112:'Sighted invoice works carried out (printed)',108:'Sighted invoice work description',127:'Sighted invoice anomalies note',99:'Sighted invoice ship-to (printed)',42:'Electricity supply point',22:'GL narration',23:'Enquiry narration',80:'Src User Fld 2'}
MANUAL_PFX=('User-adjudicated','Manual site crosswalk','Customer Request register','Rates parcel narration','Electricity supply point (address','Non-park Council facility','Multi-site list reduced')
site_now=[str(reg.cell(i,SITE).value or '') for i in range(5,21474)]
bas_now=[str(reg.cell(i,BASIS).value or '') for i in range(5,21474)]
changes=[]   # (row, old site, old basis, new site, new basis, why)
def setsite(i,s,b,why):
    r=i+5; changes.append((r,site_now[i],bas_now[i],s,b,why))
    reg.cell(r,SITE).value=(s if s else None); reg.cell(r,BASIS).value=b
    for c in (SITE,BASIS): reg.cell(r,c)._style=copy.copy(tmplS)
    site_now[i]=s or ''; bas_now[i]=b
# 5a. Peacock Avenue: 'Logan River Park' label = 17 Peacock Avenue = Logan River Parklands (user, Council directory)
n_peacock=0
for i in range(21469):
    if site_now[i]=='Logan River Park' or (';' in site_now[i] and 'Logan River Park' in [t.strip() for t in site_now[i].split(';')]):
        if bas_now[i]==MA: continue
        if ';' in site_now[i]:
            keep=sorted({('Logan River Parklands' if t.strip()=='Logan River Park' else t.strip()) for t in site_now[i].split(';') if t.strip()})
            setsite(i,'; '.join(keep) if len(keep)>1 else keep[0],(bas_now[i] if len(keep)>1 else 'Multi-site list reduced to one park (Peacock Avenue label merged into Logan River Parklands)'),'peacock-list'); n_peacock+=1
        else:
            setsite(i,'Logan River Parklands',bas_now[i].split(' [')[0]+' [user-adjudicated 18-Aug-2026: 17 Peacock Avenue = Logan River Parklands (Council directory); the Origin supply label "Logan River Park" names the same site]','peacock'); n_peacock+=1
# 5b. canonicalise alias tokens sitting in DY itself (single or joined lists), every row incl. MA (DY on an MA row is descriptive only)
n_dy=0
for i in range(21469):
    s=site_now[i]
    if not s: continue
    toks=[t.strip() for t in s.split(';') if t.strip()]
    if not any(t.lower() in ALIAS for t in toks): continue
    new=[]
    for t in toks: new+= ([x.strip() for x in ALIAS[t.lower()].split(';')] if t.lower() in ALIAS else [t])
    new=sorted({t for t in new if t.lower() not in SUBURBS}); assert new
    b=bas_now[i]
    if bas_now[i]==MA:
        setsite(i,'; '.join(new),b,'dy-alias-MA')
    elif len(new)==1:
        setsite(i,new[0],('Multi-site list reduced to one park (spelling variant canonicalised, Site_Crosswalk)' if len(toks)>1 else b.split(' [')[0]+' [spelling variant canonicalised, Site_Crosswalk]'),'dy-alias')
    else:
        setsite(i,'; '.join(new),(MULTI if len(toks)>1 else MULTI),'dy-alias-multi')
    n_dy+=1
# 5c. rows whose SOURCE columns carry an alias token: re-decide from the highest-priority column naming any park (alias-extended)
n_src=0; n_src_multi=0; n_skip=0
for i in range(21469):
    if bas_now[i]==MA: continue
    row=data[i]
    hit=False
    for c in COLS:
        t=clean(row[c-1])
        if t and ALP.search(t): hit=True; break
    if not hit: continue
    found=None
    for c in sorted([c for c in COLS if c in PRIORITY],key=lambda c:PRIORITY[c]):
        t=clean(row[c-1])
        if not t: continue
        if c==99 and re.search(r'177\s+Chambers\s+Flat',t,re.I): continue
        ps=parks_in(t)
        if ps: found=(c,ps,bool(ALP.search(t))); break
    if not found or not found[2]: continue                       # alias not in the deciding column: leave the earlier decision
    c,ps,_=found
    cur=sorted({t.strip() for t in site_now[i].split(';') if t.strip()})
    if ps==cur: continue
    if any(bas_now[i].startswith(p) for p in MANUAL_PFX) and site_now[i]: n_skip+=1; continue    # a manual/outranking basis is never overwritten by the matcher
    if len(ps)==1: setsite(i,ps[0],LBL[c]+' [spelling variant canonicalised, Site_Crosswalk]','src-alias'); n_src+=1
    else: setsite(i,'; '.join(ps),MULTI,'src-alias-multi'); n_src_multi+=1
print(f'sites: peacock {n_peacock}, DY alias {n_dy}, source alias {n_src} (+{n_src_multi} multi), skipped manual {n_skip}, Q Power printed sites {sum(site_written.values())}')
# fold check: no two single-park labels share a case-folded key
fold=collections.defaultdict(set)
for s,b in zip(site_now,bas_now):
    if s and ';' not in s and b not in set(CLASS) and b!=MA: fold[s.lower()].add(s)
assert all(len(v)==1 for v in fold.values()),{k:v for k,v in fold.items() if len(v)>1}
assert not [s for s in site_now if s and ';' not in s and s.lower() in SUBURBS]
# 5c2. Site_Allocation col G and Evidence_Invoice_Lines col L: canonicalise alias tokens (Logan River Park -> Logan River Parklands etc.)
wsa=wb['Site_Allocation']; n_sa=n_eilL=0
for rr_ in range(5,SA_END+1):
    v=wsa.cell(rr_,7).value
    if v and str(v).lower() in ALIAS and ';' not in ALIAS[str(v).lower()]: wsa.cell(rr_,7).value=ALIAS[str(v).lower()]; n_sa+=1
for rr_ in range(5,NEW_EIL_END+1):
    v=wl.cell(rr_,12).value
    if v and str(v).lower() in ALIAS and ';' not in ALIAS[str(v).lower()]: wl.cell(rr_,12).value=ALIAS[str(v).lower()]; n_eilL+=1
print(f'Site_Allocation G canonicalised {n_sa}, EIL L canonicalised {n_eilL}')
# 5d. Site_Crosswalk: append the user decisions
xw=wb['Site_Crosswalk']; xs=[copy.copy(xw.cell(5,c)._style) for c in range(1,7)]
xr=XW_END+1
XROWS=[('17 Peacock Avenue','Logan River Parklands','Park','USER DECISION 18-Aug-2026: 17 Peacock Avenue = Logan River Parklands (Council parks directory, logan.qld.gov.au/community/parks-and-gardens/parks-directory/Logan-River-Parklands). Supersedes the UNRESOLVED web-verification row above. The Origin supply label "Logan River Park" (Lot 2 Peacock Ave, Beenleigh) is the same site: the 46 register rows carrying that label are merged into Logan River Parklands.','User (Council directory)',TODAY),
       ('Logan River Park','Logan River Parklands','Park (label variant)','Origin supply-point label for Lot 2 Peacock Ave, Beenleigh; user decision above resolves the address to Logan River Parklands. Label retired from Panel A.','User (Council directory)',TODAY),
       ('71 Cinderella Drive','Springwood Park','Park','USER DECISION 18-Aug-2026 (Council parks directory, logan.qld.gov.au/community/parks-and-gardens/parks-directory/Springwood-Park) confirms the earlier web-verified SUPPORTED decision; all 162 register lines naming Cinderella Drive already read Springwood Park - no write required.','User (Council directory)',TODAY)]
for a,v in ALIAS.items():
    if a in ('logan garden park','james smith rec park','logan river park'): continue
    XROWS.append((a.title().replace("'S","'s"),v,('Park (spelling variant)' if ';' not in v else 'Two parks named (multi-site)'),
        f'USER DECISION 18-Aug-2026 (candidate-name adjudication, Add/Merge): printed/narrated form maps to the gazetteer label {v}.'+(' Guarded: Logan Reserve is also a SUBURB, so the alias applies only where the field itself reads Logan Reserve ("X Park, Logan Reserve" is suburb usage and never attributes).' if a=='logan reserve' else '')+(' Guarded: applied only where the field names Oliver.' if a=='sport complex' else '')+(' Ambiguous typo (Teviot Downs / Tervit) resolved to the same site.' if a=='tervit down park' else '')+(' The requested target "Mount Warren Park" is a Logan SUBURB name and is blocked as a gazetteer label (v48 blocklist; the Regents Park ALLOW exception needs the user to confirm the park\'s official name); added AS PRINTED pending that confirmation.' if a=='mt warren oval park' else ''),'User',TODAY))
for s in SELF: XROWS.append((s,s,'Park','USER DECISION 18-Aug-2026 (candidate-name adjudication, Add/Merge): confirmed as the existing gazetteer label; no change.','User',TODAY))
XROWS.append(('Flindersia Riverside Park','Flindersia Riverside Park','Park (added as printed)','Printed on Q Power 14190 / 14320 / 14321 (Site: Flindersia Riverside Park, Logan Reserve). Added as printed (v44 precedent). Possible same site as the rates-narration label "Logan Reserve Riverside Park" (5 lines) - user to adjudicate (Open Items #162).','Claude (flag)',TODAY))
XROWS.append(('Logan Gardens Park -> Logan Garden Park (v48 variant map)','Logan Gardens','Park (variant map corrected)','The v48 variant map sent "Logan Gardens Park" to a non-existent label "Logan Garden Park"; corrected to Logan Gardens on the user decision.','Claude','TODAY'.replace('TODAY',TODAY)))
for vals in XROWS:
    for c,v in enumerate(vals,1):
        cell=xw.cell(xr,c); cell.value=v; cell._style=copy.copy(xs[c-1])
    xr+=1
NEW_XW_END=xr-1
# 5e. Sites rebuild (Panel A from the register + Site_Allocation; group rows per class; Panel B relocated unchanged)
import pandas as pd
amt=[float(reg.cell(i,20).value or 0) for i in range(5,21474)]; assert abs(sum(amt)-CT)<0.005
pkv=[str(reg.cell(i,17).value) for i in range(5,21474)]
parks=collections.Counter(); ppk=collections.defaultdict(collections.Counter); pbas=collections.defaultdict(collections.Counter); CL=set(CLASS)
for s,b,a_,pk in zip(site_now,bas_now,amt,pkv):
    if s and ';' not in s and b not in CL and b!=MA and b!=MULTI: parks[s]+=a_; ppk[s][pk]+=1; pbas[s][b]+=1
sa=[x for x in cw.get_sheet_by_name('Site_Allocation').to_python()[4:] if x and x[0] not in (None,'') and isinstance(x[7],(int,float))]
for x in sa:
    if str(x[6]).lower() in ALIAS and ';' not in ALIAS[str(x[6]).lower()]: x=list(x); x[6]=ALIAS[str(x[6]).lower()]
    if str(x[6])!=RESIDUE: parks[str(x[6])]+=float(x[7]); ppk[str(x[6])][str(x[4])]+=1; pbas[str(x[6])]['Site_Allocation (printed line amounts)']+=1
order=sorted(parks,key=lambda p:-parks[p])
so=wb['Sites']; hdr=copy.copy(so.cell(4,1)._style); dat=copy.copy(so.cell(5,1)._style); num=copy.copy(so.cell(5,2)._style); mon=copy.copy(so.cell(5,3)._style)
elec=[[so.cell(rr,c).value for c in range(1,5)] for rr in range(PB0,PB1+1)]; elec_st=[[copy.copy(so.cell(rr,c)._style) for c in range(1,5)] for rr in range(PB0,PB1+1)]
gnote={so.cell(rr,1).value:so.cell(rr,9).value for rr in range(G0,G1+1)}
for row in so.iter_rows(min_row=3,max_row=so.max_row):
    for c in row: c.value=None
def put(ws,rr,c,v,st=None):
    cell=ws.cell(rr,c); cell.value=v
    if st is not None: cell._style=copy.copy(st)
FYS=['FY2023/24','FY2024/25','FY2025/26','FY2026/27']
put(so,3,1,'Panel A: cost per park / site (live)',hdr)
for j,h in enumerate(['Park / Site','Register lines (live)','$ ex GST (live)']+FYS+['Primary PK','Primary basis','Bases'],1): put(so,4,j,h,hdr)
rr=5; PA_S=5
for p in order:
    put(so,rr,1,p,dat)
    so.cell(rr,2).value=f'=COUNTIF(Register!${SC}$5:${SC}$21473,$A{rr})'; so.cell(rr,2)._style=copy.copy(num); so.cell(rr,2).number_format='#,##0;-#,##0;"-"'
    so.cell(rr,3).value=f'=SUMIF(Register!${SC}$5:${SC}$21473,$A{rr},Register!$T$5:$T$21473)+SUMIF(Site_Allocation!$G$5:$G${SA_END},$A{rr},Site_Allocation!$H$5:$H${SA_END})'
    so.cell(rr,3)._style=copy.copy(mon); so.cell(rr,3).number_format='$#,##0.00;($#,##0.00);"-"'
    for kk,fy in enumerate(FYS):
        so.cell(rr,4+kk).value=(f'=SUMIFS(Register!$T$5:$T$21473,Register!${SC}$5:${SC}$21473,$A{rr},Register!$E$5:$E$21473,"{fy}")+SUMIFS(Site_Allocation!$H$5:$H${SA_END},Site_Allocation!$G$5:$G${SA_END},$A{rr},Site_Allocation!$D$5:$D${SA_END},"{fy}")')
        so.cell(rr,4+kk)._style=copy.copy(mon); so.cell(rr,4+kk).number_format='$#,##0.00;($#,##0.00);"-"'
    put(so,rr,8,max(ppk[p].items(),key=lambda x:x[1])[0] if ppk[p] else '',dat); put(so,rr,9,max(pbas[p].items(),key=lambda x:x[1])[0] if pbas[p] else '',dat); put(so,rr,10,len(pbas[p]),dat); rr+=1
PA_E=rr-1; GS=rr
for lbl in [RESIDUE]+[c for c in CLASS if c!=RESIDUE]:
    put(so,rr,1,lbl,dat)
    if lbl==RESIDUE:
        so.cell(rr,2).value=f'=COUNTIF(Register!${BC}$5:${BC}$21473,"{MA}")'; so.cell(rr,3).value=f'=SUMIF(Site_Allocation!$G$5:$G${SA_END},$A{rr},Site_Allocation!$H$5:$H${SA_END})'
        for kk,fy in enumerate(FYS): so.cell(rr,4+kk).value=f'=SUMIFS(Site_Allocation!$H$5:$H${SA_END},Site_Allocation!$G$5:$G${SA_END},$A{rr},Site_Allocation!$D$5:$D${SA_END},"{fy}")'
    else:
        so.cell(rr,2).value=f'=COUNTIF(Register!${BC}$5:${BC}$21473,$A{rr})'; so.cell(rr,3).value=f'=SUMIF(Register!${BC}$5:${BC}$21473,$A{rr},Register!$T$5:$T$21473)'
        for kk,fy in enumerate(FYS): so.cell(rr,4+kk).value=f'=SUMIFS(Register!$T$5:$T$21473,Register!${BC}$5:${BC}$21473,$A{rr},Register!$E$5:$E$21473,"{fy}")'
    so.cell(rr,2)._style=copy.copy(num); so.cell(rr,2).number_format='#,##0;-#,##0;"-"'
    for kk in range(3,8): so.cell(rr,kk)._style=copy.copy(mon); so.cell(rr,kk).number_format='$#,##0.00;($#,##0.00);"-"'
    put(so,rr,9,gnote.get(lbl,''),dat); rr+=1
GE=rr-1; TOT=rr
put(so,rr,1,'TOTAL (must tie to the register)',hdr); so.cell(rr,2).value=f'=SUM(B{PA_S}:B{GE})'; so.cell(rr,2)._style=copy.copy(hdr); so.cell(rr,2).number_format='#,##0'
for kk in range(3,8):
    Lc=get_column_letter(kk); so.cell(rr,kk).value=f'=SUM({Lc}{PA_S}:{Lc}{GE})'; so.cell(rr,kk)._style=copy.copy(hdr); so.cell(rr,kk).number_format='$#,##0.00;($#,##0.00);"-"'
CHK=rr+1
put(so,CHK,1,'Tie check (lines = 21,469 and $ = Register!$T$21475)',hdr)
so.cell(CHK,2).value=f'=IF(B{TOT}=21469,"TRUE","FALSE")'; so.cell(CHK,2)._style=copy.copy(hdr)
so.cell(CHK,3).value=f'=IF(ROUND(C{TOT},2)=ROUND(Register!$T$21475,2),"TRUE","FALSE")'; so.cell(CHK,3)._style=copy.copy(hdr)
so.cell(CHK,4).value=f'=IF(ROUND(SUM(D{TOT}:G{TOT}),2)=ROUND(C{TOT},2),"TRUE","FALSE")'; so.cell(CHK,4)._style=copy.copy(hdr)
NPB=CHK+2; put(so,NPB,1,'Panel B: electricity by retailer account (relocated unchanged)',hdr)
for j,h in enumerate(['Retailer account','Site','Lines','$ ex GST (net, all years)'],1): put(so,NPB+1,j,h,hdr)
rr=NPB+2
for vals,sts in zip(elec,elec_st):
    for j,v in enumerate(vals,1):
        c=so.cell(rr,j); c.value=v; c._style=copy.copy(sts[j-1])
    rr+=1
NPB_E=rr-1

# ================= 6. Vendor_Boilerplate sightings
vb=wb['Vendor_Boilerplate']; vrow={str(vb.cell(i,1).value):i for i in range(5,vb.max_row+1) if vb.cell(i,1).value}
bumps={'BP:QP-P1':len(INV),'BP:QP-T1':sum(1 for k in INV if bp_terms(k)=='BP:QP-T1'),'BP:QP-T2':sum(1 for k in INV if bp_terms(k)=='BP:QP-T2')}
for key,n in bumps.items():
    i=vrow[key]; assert isinstance(vb.cell(i,5).value,(int,float)); vb.cell(i,5).value=int(vb.cell(i,5).value)+n
BPK=int(cfg['BOILERPLATE_KEYS'])

# ================= 7. Config / narrative sheets
cf=wb['Config']; cr={str(cf.cell(i,1).value):i for i in range(1,cf.max_row+1) if cf.cell(i,1).value}
n_ties=sum(1 for x in E['recon'] if x['tie']=='TRUE'); elec_lines=sum(x['lines'] for x in E['recon']); elec_ex=round(sum(x['reg_ex'] for x in E['recon']),2)
CFG=[('WORKBOOK_VERSION',VER),('EI_DATA',f'5:{NEW_EI_END}'),('EI_TOTALS_ROW',NEW_EI_TOT),('EI_CONTROL_ROWS',f'{NEW_C0}:{NEW_C1}'),('EIL_DATA',f'5:{NEW_EIL_END}'),('EIL_END',NEW_EIL_END),
     ('RECON_ROWS',f'{NEW_R0}:{NEW_R1}'),('RECON_COUNT',NEW_RECN),('SIGHTED_COUNT',NEW_SIGHT),('BOILERPLATE_KEYS',BPK),
     ('SITES_PANELA',f'{PA_S}:{PA_E}'),('SITES_GROUP_ROWS',f'{GS}:{GE}'),('SITES_TOTAL_ROW',TOT),('SITES_CHECK_ROW',CHK),('SITES_PANELB',f'{NPB+2}:{NPB_E}'),('PARK_COUNT',len(order)),
     ('SITE_CROSSWALK_DATA',f'5:{NEW_XW_END}')]
for k,v in CFG: cf.cell(cr[k],2).value=v
cs=[copy.copy(cf.cell(cf.max_row,c)._style) for c in (1,2)]; nr=cf.max_row+1
for k,v in [('DATE_REPAIRS_V51',13),('SITE_ALIAS_WRITES_V51',len(changes)),('ELEC_STMTS_RECONCILED_V51','A-FE35E476-022:027'),('ELEC_GROUPS_TIED_V51',f'{n_ties}/{len(E["recon"])}'),('ELEC_LINES_TIED_V51',elec_lines)]:
    cf.cell(nr,1).value=k; cf.cell(nr,2).value=v; cf.cell(nr,1)._style=copy.copy(cs[0]); cf.cell(nr,2)._style=copy.copy(cs[1]); nr+=1
tot_sub=sum(F[k]['sub'] for k in MATCHED); tot_all=sum(F[k]['sub'] for k in INV)
h=wb['Handover']; hr=h.max_row+1; hs=[copy.copy(h.cell(h.max_row,c)._style) for c in range(1,4)]
h.cell(hr,1).value=(f'{VER} change ({TODAY}): Batch 42 rule 17 capture (Q Power historical corpus, 51 invoices, ${tot_all:,.2f} ex GST): {NEW_GB} green blocks (${tot_sub:,.2f}), 3 invoices with no register line held under rule 16 (14406, 14865, 15222; Open Items #{OI_UNM}). '
  f'32 of 51 invoices print NO line table (Aug-2024 to Jun-2025 work-note template) and are captured as single-scope lines at the printed Sub-Total (BLC precedent). 13 v45 Q Power green blocks repaired: the invoice date held the PLEASE PAY BY date (F1). '
  f'Batch 43: Origin collective statements A-FE35E476-022 to -027 reconciled - {n_ties} of {len(E["recon"])} register (statement, NMI) groups tie to a statement sub-account to the cent, {elec_lines} lines / ${elec_ex:,.2f}. '
  f'Sites: 30 user-adjudicated candidate names and 17 Peacock Avenue = Logan River Parklands applied through Site_Crosswalk ({len(changes)} DY/DZ writes); Panel A {len(order)} sites. Capture build: column T untouched, control total unchanged at ${CT:,.2f}.')
h.cell(hr,1)._style=hs[0]; h.cell(hr,2).value=f'Sighted count {SIGHT} -> {NEW_SIGHT}. Corpus sha256 {P["sha"][:12]}..., extraction {P["manifest"]["extraction_tool"]}. Register total {CT:,.2f} unchanged.'; h.cell(hr,2)._style=hs[1]; h.cell(hr,3)._style=hs[2]
m=wb['Method']; mr=m.max_row+1; ms=copy.copy(m.cell(m.max_row,1)._style)
METH=[f"49.0 Batch 42 capture ({VER}). Corpus qpower_historical_20260818 (Q Power.pdf, 118 pages, 51 primary documents, 0 duplicate copies; extraction {P['manifest']['extraction_tool']}). Every invoice's printed subtotal, GST and total re-tied (GST = ROUND_HALF_UP(subtotal/10), subtotal + GST = total, header AMOUNT = total); rule 12 screen prefix-aware over all {len(set(e.split('-')[0] for e in [k for k in [str(x[0]) for x in cw.get_sheet_by_name('Evidence_Invoices').to_python()[4:2046]] if '-' in k]))} EvID prefixes plus the bare id: 0 re-sightings; corpus-internal normalised-text screen: 0 collisions. {NEW_GB} invoices matched to exactly ONE non-zero AP row at the printed subtotal within 2c (amount corroboration, v35 rule) with the register Doc Date agreeing to the day; no $0.00 companions. Three invoices carry no register line and are captured under rule 16 only (v36 unmatched pathway).",
 f"49.1 The Q Power work-note template prints NO line table ({VER}, extends the v40 BLC single-scope precedent). 32 of 51 invoices (Aug-2024 to Jun-2025) print a Description / Work Note narrative and a totals block only - no priced item grid on the text layer OR the OCR layer (the extractor's 'OUT' set). Each is captured as ONE single-scope line = the printed Description block at the printed Sub-Total ex GST, qty / unit '(not printed)', with the basis stated in the EIL Notes and the anomalies field; check 1 ties by construction and checks 2 and 3 tie on the printed figures. This is rule 16 line capture at the granularity the invoice prints, not a summary row: the invoice has one scope. If Finance requires a labour / materials breakdown for these 32 invoices it must be requested from Q Power (Open Items #160).",
 f"49.2 F1, fourth form ({VER}): the extractor's invoice_date is the WORK NOTE date or the PLEASE PAY BY date. The Q Power header row prints 'PLEASE PAY BY | AMOUNT | INVOICE DATE' as three values on one row ('Invoicing 05/09/2024 $712.80 22/08/2024'); the corpus invoice_date disagreed with the printed INVOICE DATE on 36 of 51 documents (work-note date) and, in the Batch 40 corpus, on all 13 v45 Q Power captures (the pay-by date). Restated from the header row - due date first, invoice date third - and cross-checked against the register Doc Date on every matched row (48 of 48 agree). The 13 v45 green blocks (15133-15198) are repaired in this build from the retained Batch 40 page text: Invoice Date and Due Date rewritten on the register and on Evidence_Invoices, the repair stated in the anomalies. Seven v28 captures (QP-15124/15137/15138/15147/15158/15159/15169) show the same 14-day pattern but their page text was not retained; flagged, not rewritten (Open Items #159).",
 f"49.3 F12 on the Q Power simPRO line table ({VER}). Wrapped items print over THREE layout rows - description above, part number / item on the priced row, continuation below ('DIN Contactor 1P 8.5A AC3 25A AC1 230V | HAGESC125 | 1NO DIN Rail'). Rule: a NARRATIVE row immediately above a priced row is its prefix; when a prefix exists the NARRATIVE row immediately below is its suffix; every narrative row inside the table block must be consumed (asserted) and every row asserted verbatim in the retained page text. Site / Site Address / Order No. print as label-led rows with continuation rows below; the description block runs from 'Description' to 'Ordered By:'; work notes run from the first '(dd/mm/yyyy) - Work Note' header to the table.",
 f"49.4 Batch 43: Origin collective account summaries A-FE35E476-022 to -027 ({VER}). The six 'Your collective account summary' PDFs (Sep-2025 to Feb-2026 issues) print the CURRENT CHARGES - SUB ACCOUNT SUMMARY: sub account, invoice number, full supply address, GST and charges incl GST for every Council sub-account (233 / 301 / 396 / 202 / 199 / 298 accounts). Parsed with the account count and the printed totals asserted per statement. Every register (statement, NMI) group - {len(E['recon'])} groups, {elec_lines} lines, ${elec_ex:,.2f} ex GST - was matched to a statement sub-account by supply-address prefix (the register supply string is cut at 40 characters, so a truncated tail token is dropped before comparison) and ties register ex GST to (charges incl GST - GST) to the cent: {n_ties} of {len(E['recon'])}. With v42 (-021, -028 to -031) the FY2025/26 collective account is now reconciled on eleven statements; -020 (14 lines) and -033 (148 lines) remain. The full supply addresses add no park name the truncated register labels did not already carry, so no site write follows from Batch 43. Worked in PSWP_Batch43_Electricity_Reconciliation_022_027.xlsx.",
 f"49.5 Site decisions applied ({VER}). Thirty user-adjudicated candidate names (Add/Merge) folded into the alias map and Site_Crosswalk: spelling variants (Jamie Nicholson -> Jamie Nicolson Park, Doug Larson -> Doug Larsen Park, Teviot Down / Tervit Down -> Teviot Downs Park, Hubener -> Hubner Park, Skinner -> Skinners Park, Lavell -> Lavelle Park, Richared Wilson -> Richard Wilson Park, Woodland District -> Woodlands District Park, Forest Glen -> Forestglen Park, Shaw St -> Shaw Street Oval, Oliver / Oliver's Sport(s) / Vandalism Olivers -> Olivers Sports Complex, Colville St Green -> Colville Park, Clean Up Cecil Clark -> Cecil Clark Park, Alexandra -> Alexander Park, Darlington Park -> Darlington Parklands, Logan Gardens Park -> Logan Gardens, Springwood Dog Off Leash Park -> Springwood Park, Harley Street -> Harley St Shailer Park, Mt Warren Oval Park added as printed (the target Mount Warren Park is a suburb name, blocklisted), Logan Reserve -> Logan Reserve Park [guarded], Jamie Nicholson Park Lavelle Park -> two parks); five confirmed as existing labels. Applied three ways: (a) alias tokens sitting in DY itself canonicalised, joined lists re-screened; (b) rows whose deciding source column (priority 111 > 106 > 112 > 108 > 127 > 99 > 42 > 22 > 23 > 80) names an alias re-decided; a manual, CR-register, rates-parcel or user-adjudicated basis is never overwritten by the matcher. {len(changes)} DY/DZ writes. The v48 variant map's 'Logan Gardens Park -> Logan Garden Park' (a non-existent label) is corrected. 17 Peacock Avenue = Logan River Parklands: the Origin label 'Logan River Park' (46 lines, all at Lot 2 Peacock Ave) merged into Logan River Parklands. 71 Cinderella Drive = Springwood Park was already applied (v49); recorded."]
for i,t in enumerate(METH):
    c=m.cell(mr+i,1); c.value=t; c._style=copy.copy(ms)
d=wb['Data_Acquisition']; dr=d.max_row+1; ds=copy.copy(d.cell(d.max_row,1)._style)
c=d.cell(dr,1); c.value=(f' |  F45  |  Batch 42: corpus_qpower_historical_20260818.json (sha256 {P["sha"]}, md5 {P["md5"]}), extraction {P["manifest"]["extraction_tool"]}, source Q Power.pdf 118 pages, 51 documents, 9,820 lines. Captured in full at {VER} (48 rule 17, 3 rule 16 only).  |  1  |  Captured  | '); c._style=copy.copy(ds)
c=d.cell(dr+1,1); c.value=(' |  F46  |  Batch 43: Origin collective account summaries '+', '.join(f"{k} ({v['issue']}, {v['accounts']} accounts, md5 {v['md5']})" for k,v in E['stmt'].items())+f' from A-FE35E476-026-summary (1).zip. Reconciled at {VER}: {n_ties}/{len(E["recon"])} register groups tie; PSWP_Batch43_Electricity_Reconciliation_022_027.xlsx.  |  1  |  Reconciled  | '); c._style=copy.copy(ds)
oi=wb['Open_Items']; orow=oi.max_row+1; osx=[copy.copy(oi.cell(oi.max_row,c)._style) for c in range(1,8)]
ITEMS=[(156,'All','Open','Q Power (Qld) Pty Ltd',len(noPK),round(sum(F[k]['sub'] for k in noPK),2),'NO PK ON THE INVOICE FACE (Batch 42): '+', '.join(noPK)+'. PK recorded as (not printed); the register posting stands. Same pattern as #140 - ask Q Power to print the PK on every invoice.'),
 (157,'All','Open','Q Power (Qld) Pty Ltd',len(pk_div)+1,round(sum(F[k]['sub'] for k,_,_ in pk_div)+F['14190']['sub'],2),'PRINTED PK DIVERGES FROM THE CHARGE: '+'; '.join(f'{k} prints {a}, posts {b}' for k,a,b in pk_div)+'. Plus 14190 prints the 5-digit typo PK00018 (register PK000118) - recorded as printed, not normalised. Same pattern as #100/#101/#102/#141.'),
 (158,'All','Open','Q Power (Qld) Pty Ltd',3,round(sum(F[k]['sub'] for k in UNM),2),'THREE SIGHTED INVOICES WITH NO REGISTER LINE, held on the evidence sheets (rule 16 only): 14406 ($416.22, Jimboomba Soccer Club, Kurrajong Road, order 709263 - MS016398) and 14865 ($340.00, Station Road Woodridge, order 710663) are paid per the QPO001 creditor history but post outside 4090240/4090260; 15222 ($5,327.94, Doug Larsen Park BBQ element replacement, PO 717370) is dated and paid 12-Aug-2026, after the loaded FY2026/27 extract. Match on the next 27SLACT pull.'),
 (159,'FY2026/27','Open','Q Power (Qld) Pty Ltd',7,0,'INVOICE DATE HELD THE PAY-BY DATE (F1) on seven v28 captures QP-15124, QP-15137, QP-15138, QP-15147, QP-15158, QP-15159, QP-15169 (EI/green-block date = register Doc Date + 14 days on every one). Their Batch 25 page text was not retained, so no restatement is possible without re-sighting; the 13 v45 captures with retained text were repaired at v51. Re-sight from QPower.pdf (Batch 25) or the TechOne attachment.'),
 (160,'All','Open','Q Power (Qld) Pty Ltd',32,round(sum(F[k]['sub'] for k in MATCHED if F[k]['tie']!='TIE'),2),'WORK-NOTE TEMPLATE PRINTS NO LINE TABLE on 32 invoices (Aug-2024 to Jun-2025): captured as single-scope lines at the printed subtotal (Method 49.1). If a labour / materials breakdown is required for coding or contract-rate checks, request it from Q Power; the later simPRO template (from Jun-2025) prints Electrician (LCC) / Apprentice hours at $85-$120 and $56-$65 and materials by part number.'),
 (161,'All','Open','Batch 42 evidence gaps',len(attach),0,'REFERENCED ATTACHMENTS NOT SUPPLIED (quotes, drawings, site photographs, mud maps, work orders) on '+str(len(attach))+' invoices: '+', '.join(attach)+'. Request if the works scope is queried.'),
 (162,'All','Open','Site dimension',9,round(sum(F[k]['sub'] for k in MATCHED if F[k]['site'] and not F[k]['gaz_sites']) + sum(F[k]['sub'] for k in MATCHED if F[k]['gaz_sites']==['Flindersia Riverside Park']),2),"GAZETTEER FROM BATCH 42: 'Flindersia Riverside Park' added as printed (14190, 14320, 14321) - possibly the same site as the rates-narration label 'Logan Reserve Riverside Park' (5 lines): user to adjudicate. Candidate names not in the gazetteer, held in the harvest class: Jimboomba Sports Park (14052), Forest of Memories - Logan Village (14137, 14295), Glenlogan park (14389; cf. Glenlogan Lakes Park), 125 City Road Beenleigh (14206), Jimboomba Soccer Club (14406, no register line), Station Road Woodridge (14865, no register line)."),
 (163,'FY2025/26','Open','Origin Energy (collective account)',elec_lines,elec_ex,f'ELECTRICITY RECONCILED (Batch 43): statements A-FE35E476-022 to -027 - {n_ties} of {len(E["recon"])} register (statement, NMI) groups tie to a statement sub-account to the cent. FY2025/26 collective account now reconciled on eleven statements (-021 to -031); -020 (14 lines) and -033 (148 lines, a late FY2025/26 posting) still lack a statement. Request the -020 and -033 summaries. The statements confirm the register supply-point label is the Origin supply address cut at 40 characters.'),
 (164,'All','Closed','Site dimension',n_peacock,round(sum(a for a,s,b in zip(amt,site_now,bas_now) if s=='Logan River Parklands' and 'Peacock' in b),2),'17 PEACOCK AVENUE = LOGAN RIVER PARKLANDS (user, Council directory). The Origin label "Logan River Park" (46 lines) merged into Logan River Parklands; the #155 unresolved count falls by one. 71 Cinderella Drive = Springwood Park confirmed (already applied v49).')]
for i,it in enumerate(ITEMS):
    for j,v in enumerate(it,1):
        c=oi.cell(orow+i,j); c.value=v; c._style=copy.copy(osx[j-1])

# ================= 8. save / recalc / verify / ship
pre=f'{OUT}/{NEW}'; os.makedirs(OUT,exist_ok=True); wb.save(pre); print('saved')
prof=f'{ROOT}/lo_profile/user'; os.makedirs(prof,exist_ok=True)
open(f'{prof}/registrymodifications.xcu','w').write('''<?xml version="1.0" encoding="UTF-8"?>
<oor:items xmlns:oor="http://openoffice.org/2001/registry" xmlns:xs="http://www.w3.org/2001/XMLSchema" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
<item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop></item>
<item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop></item>
</oor:items>''')
t=time.time()
res=subprocess.run(['soffice',f'-env:UserInstallation=file://{ROOT}/lo_profile','--headless','--calc','--convert-to','xlsx','--outdir',RC,pre],capture_output=True,text=True,timeout=1800)
final=f'{RC}/{NEW}'; assert os.path.exists(final),res.stdout+res.stderr; print(f'recalc {time.time()-t:.0f}s')
json.dump(dict(final=final,CT=CT,EI_END=NEW_EI_END,EI_TOT=NEW_EI_TOT,C=[NEW_C0,NEW_C1],EIL_END=NEW_EIL_END,R=[NEW_R0,NEW_R1],SIGHT=NEW_SIGHT,RECN=NEW_RECN,
  gb=NEW_SIGHT,gb_rows=gb_rows,invs=INV,unm=UNM,rep_rows=rep_rows,rep=REP,PA=[PA_S,PA_E],GRP=[GS,GE],TOT=TOT,CHK=CHK,PB=[NPB+2,NPB_E],parks=len(order),CLASS=CLASS,MA=MA,
  suburbs=sorted(SUBURBS),changes=changes,aliases=sorted(ALIAS),xw=[5,NEW_XW_END],peacock=n_peacock,elec=dict(ties=n_ties,groups=len(E['recon']),lines=elec_lines,ex=elec_ex)),open(f'{ROOT}/b51_pos.json','w'),indent=1)
r=subprocess.run([sys.executable,f'{ROOT}/b51_verify.py'],capture_output=True,text=True); print(r.stdout); print(r.stderr[-1500:])
if 'FAILS: none' in r.stdout:
    subprocess.run(['cp',final,f'/mnt/user-data/outputs/{NEW}']); print('SHIPPED',NEW)
else: print('NOT SHIPPED'); sys.exit(1)
