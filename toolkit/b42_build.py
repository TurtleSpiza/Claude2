import json, copy, collections, re, sys
import openpyxl
from openpyxl.utils import get_column_letter

SRC='zipx/PS_WP_Transaction_Register_3FY_v41_HANDOVER.xlsx'
OUT='PS_WP_Transaction_Register_3FY_v42_HANDOVER.xlsx'
S=json.load(open('b39_sites2.json'))
rows=S['rows']; park_fy=S['park_fy']; park_ln=S['park_ln']
park_bas=S['park_bas']; park_pk=S['park_pk']; park_acct=S['park_acct']

SITE_COL=129; BASIS_COL=130            # DY, DZ
SC=get_column_letter(SITE_COL); BC=get_column_letter(BASIS_COL)
MULTI='Multiple sites named (not allocated)'
NONE='(no site named)'
FYS=['FY2023/24','FY2024/25','FY2025/26','FY2026/27']

wb=openpyxl.load_workbook(SRC)
reg=wb['Register']
assert reg.max_row==21475

# ---------- 1. Register: two new columns ----------
tmpl_h=reg.cell(4,42); tmpl_d=reg.cell(5,42)
for col,txt in ((SITE_COL,'Park / Site (normalised)'),(BASIS_COL,'Site basis')):
    c=reg.cell(4,col,txt); c._style=copy.copy(tmpl_h._style)
n_single=n_multi=n_none=0
for o in rows:
    r=o['row']
    if o['park'] and o['park'].startswith('Multiple'):
        site='; '.join(o['all']); basis=MULTI; n_multi+=1
    elif o['park']:
        site=o['park']; basis=o['basis']; n_single+=1
    else:
        site=None; basis=NONE; n_none+=1
    a=reg.cell(r,SITE_COL,site); a._style=copy.copy(tmpl_d._style)
    b=reg.cell(r,BASIS_COL,basis); b._style=copy.copy(tmpl_d._style)
print(f'register: single {n_single}, multi {n_multi}, none {n_none}')

# ---------- 2. Sites sheet rebuild ----------
old=wb['Sites']
elec=[[old.cell(r,c).value for c in range(1,5)] for r in range(5,old.max_row+1)]
elec_styles=[[copy.copy(old.cell(r,c)._style) for c in range(1,5)] for r in range(5,old.max_row+1)]
hdr_style=copy.copy(old.cell(4,1)._style)
ttl_style=copy.copy(old.cell(1,1)._style)
sub_style=copy.copy(old.cell(2,1)._style)
data_style=copy.copy(old.cell(5,1)._style)
num_style=copy.copy(old.cell(5,3)._style)
mon_style=copy.copy(old.cell(5,4)._style)

wb.remove(old)
ws=wb.create_sheet('Sites', wb.sheetnames.index('PK_Listing'))
ws.sheet_view.showGridLines=False
for i,w in enumerate([42,11,17,16,16,16,16,13,34,10],1):
    ws.column_dimensions[get_column_letter(i)].width=w

def put(r,c,v,st=None,fmt=None):
    cell=ws.cell(r,c,v)
    if st is not None: cell._style=copy.copy(st)
    if fmt: cell.number_format=fmt
    return cell

put(1,1,'Site dimension: cost per park / site, all loaded years',ttl_style)
put(2,1,'Panel A keyed on Register column '+SC+' (Park / Site normalised); basis in column '+BC+'. A site is attributed only where the source text carries a FULL PARK NAME (user directive, 18-Aug-2026). Lines naming several sites are held at "'+MULTI+'" and are NEVER allocated to one park - the association is not fixed by evidence (v32 partial-scope principle). Gazetteer: 646 park names, built from the internal-rates parcel narrations (S07 series, park name printed after the street address) and from names recurring three or more times in sighted-invoice site details, work descriptions and electricity supply points.',sub_style)
put(3,1,'Panel A: cost per park / site (live)',hdr_style)
HDRS=['Park / Site','Lines (live)','$ ex GST (live)']+FYS+['Primary PK','Primary basis','Bases']
for j,h in enumerate(HDRS,1): put(4,j,h,hdr_style)

parks=sorted(park_fy.keys(), key=lambda p:-sum(park_fy[p].values()))
PA_START=5; r=PA_START
for p in parks:
    put(r,1,p,data_style)
    put(r,2,f'=COUNTIF(Register!${SC}$5:${SC}$21473,$A{r})',num_style,'#,##0;-#,##0;"-"')
    put(r,3,f'=SUMIF(Register!${SC}$5:${SC}$21473,$A{r},Register!$T$5:$T$21473)',mon_style,'$#,##0.00;($#,##0.00);"-"')
    for k,fy in enumerate(FYS):
        put(r,4+k,f'=SUMIFS(Register!$T$5:$T$21473,Register!${SC}$5:${SC}$21473,$A{r},Register!$E$5:$E$21473,"{fy}")',mon_style,'$#,##0.00;($#,##0.00);"-"')
    put(r,8,max(park_pk[p].items(),key=lambda x:x[1])[0],data_style)
    put(r,9,max(park_bas[p].items(),key=lambda x:x[1])[0],data_style)
    put(r,10,len(park_bas[p]),num_style,'#,##0')
    r+=1
PA_END=r-1
# held + unattributed rows keyed on the BASIS column
for lbl,note in ((MULTI,'Held: several parks named on one line'),(NONE,'No park name in any source field')):
    put(r,1,lbl,data_style)
    put(r,2,f'=COUNTIF(Register!${BC}$5:${BC}$21473,$A{r})',num_style,'#,##0;-#,##0;"-"')
    put(r,3,f'=SUMIF(Register!${BC}$5:${BC}$21473,$A{r},Register!$T$5:$T$21473)',mon_style,'$#,##0.00;($#,##0.00);"-"')
    for k,fy in enumerate(FYS):
        put(r,4+k,f'=SUMIFS(Register!$T$5:$T$21473,Register!${BC}$5:${BC}$21473,$A{r},Register!$E$5:$E$21473,"{fy}")',mon_style,'$#,##0.00;($#,##0.00);"-"')
    put(r,9,note,data_style)
    r+=1
GRP_END=r-1
TOT_ROW=r
put(r,1,'TOTAL (must tie to the register)',hdr_style)
put(r,2,f'=SUM(B{PA_START}:B{GRP_END})',hdr_style,'#,##0')
put(r,3,f'=SUM(C{PA_START}:C{GRP_END})',hdr_style,'$#,##0.00;($#,##0.00);"-"')
for k in range(4,8): put(r,k,f'=SUM({get_column_letter(k)}{PA_START}:{get_column_letter(k)}{GRP_END})',hdr_style,'$#,##0.00;($#,##0.00);"-"')
for k in (8,9,10): put(r,k,None,hdr_style)
CHK_ROW=r+1
put(CHK_ROW,1,'Tie check (lines = 21,469 and $ = Register!$T$21475)',hdr_style)
put(CHK_ROW,2,f'=IF(B{TOT_ROW}=21469,"TRUE","FALSE")',hdr_style)
put(CHK_ROW,3,f'=IF(ROUND(C{TOT_ROW},2)=ROUND(Register!$T$21475,2),"TRUE","FALSE")',hdr_style)
put(CHK_ROW,4,f'=IF(ROUND(SUM(D{TOT_ROW}:G{TOT_ROW}),2)=ROUND(C{TOT_ROW},2),"TRUE","FALSE")',hdr_style)
PB_TITLE=CHK_ROW+2
put(PB_TITLE,1,'Panel B: electricity by retailer account, all loaded years (relocated from the v41 Sites sheet, unchanged)',hdr_style)
put(PB_TITLE+1,1,'Retailer account',hdr_style); put(PB_TITLE+1,2,'Site',hdr_style)
put(PB_TITLE+1,3,'Lines',hdr_style); put(PB_TITLE+1,4,'$ ex GST (net, all years)',hdr_style)
r=PB_TITLE+2
for vals,sts in zip(elec,elec_styles):
    for j,v in enumerate(vals,1): put(r,j,v,sts[j-1])
    r+=1
PB_END=r-1
print(f'Sites: PanelA {PA_START}:{PA_END} ({len(parks)} parks), groups to {GRP_END}, total {TOT_ROW}, check {CHK_ROW}, PanelB {PB_TITLE+2}:{PB_END}')

# ---------- 3. Config ----------
cf=wb['Config']
cf.cell(3,2,'v42')
cst_a=copy.copy(cf.cell(27,1)._style); cst_b=copy.copy(cf.cell(27,2)._style)
newcfg=[('SITE_COL',f'{SITE_COL} ({SC})'),('SITE_BASIS_COL',f'{BASIS_COL} ({BC})'),
        ('SITES_PANELA',f'{PA_START}:{PA_END}'),('SITES_GROUP_ROWS',f'{PA_END+1}:{GRP_END}'),
        ('SITES_TOTAL_ROW',TOT_ROW),('SITES_CHECK_ROW',CHK_ROW),
        ('SITES_PANELB',f'{PB_TITLE+2}:{PB_END}'),('PARK_COUNT',len(parks)),
        ('SITE_ATTRIBUTED_LINES',n_single),('SITE_MULTI_LINES',n_multi)]
r=29
for k,v in newcfg:
    a=cf.cell(r,1,k); a._style=cst_a
    b=cf.cell(r,2,v); b._style=cst_b
    r+=1

# ---------- 4. Handover / Method / Data_Acquisition / Open_Items ----------
h=wb['Handover']; hr=h.max_row+1
hs=[copy.copy(h.cell(h.max_row,c)._style) for c in range(1,4)]
h.cell(hr,1,'v42 change (18-Aug-2026): site dimension added. Register gains two columns - '+SC+' Park / Site (normalised) and '+BC+' Site basis - and the Sites sheet is rebuilt as a live cost-per-park dimension. Capture build: no green block written, column T untouched, control total unchanged at $11,377,312.51.')._style=hs[0]
h.cell(hr,2,f'Batch 39 analysis session: five Origin consolidated statement CSVs (A-FE35E476-021/-028/-029/-030/-031), the Waste branch Public Place Bins Charges 25/26 driver, and the Cleanaway invoice 25009848 email. {n_single} register lines attributed to {len(parks)} named parks; {n_multi} multi-site lines held unallocated; {n_none} carry no park name in any source field.')._style=hs[1]
h.cell(hr,3,None)._style=hs[2]

m=wb['Method']; mr=m.max_row+1
ms=copy.copy(m.cell(m.max_row,1)._style)
METH=[
 "41.0 Site dimension (NEW v42). Cost per park / site is now a first-class register dimension. Column "+SC+" carries the normalised park or site name and column "+BC+" the basis on which it was attributed. Nothing is inferred: a line is attributed only where a source field carries a FULL PARK NAME matching the gazetteer (user directive, 18-Aug-2026). Sources are read in evidence order - rates parcel narration, sighted invoice site details (printed), sighted invoice work description, electricity supply point, GL narration, enquiry narration - and the first hit wins, with the source recorded as the basis.",
 "41.1 The gazetteer. 646 park names. The primary source is the internal-rates S07 narration series, which prints parcel id, street address and park name (1,139 parcels, 583 carrying a park name); these give properly cased, authoritative Logan park names with their addresses. A secondary harvest takes any capitalised multi-word name ending Park / Reserve / Gardens / Wetlands / Forest / Parklands / Oval / Bushland / Green / Square / Common that recurs three or more times across site details, work descriptions and supply points. Generic constructions are blocklisted (General Park, Skate Park, Dog Park, Water Park, Car Park, Memorial Park, Sports Park, District Park, Adventure Park and the like) because they name a facility type, not a site; 'General Park maintenance' had otherwise attributed 152 lines and $72,856 to a park that does not exist.",
 "41.2 Multi-site lines are HELD, never split. 319 lines ($904,116.12) name several parks on one line - a Council-wide grounds contract billing a schedule of sites, a standing order across a maintenance round. Column "+SC+" carries the semicolon-joined list of every park named and column "+BC+" reads '"+MULTI+"'. No allocation is made, because no evidence fixes the split; this is the v32 partial-scope principle applied to the site dimension. Allocating them on a head-count or dollar-weighted basis would be fabrication. Precedent: register row 14039 ($33,335.10, PAR/340E/2026) prints 'Logan Gardens; Flagstone; Darlington Parklands; ...' and an earlier single-park pass had attributed the whole amount to one of them.",
 "41.3 Coverage at v42. "+str(n_single)+" lines ($2,407,766.77, 21.2% of the register) attributed to "+str(len(parks))+" named parks; "+str(n_multi)+" lines ($904,116.12, 7.9%) held multi-site; "+str(n_none)+" lines ($8,065,429.62, 70.9%) carry no park name in any source field. The unattributed majority is not a defect: most Parks spend is coded at PK level against a maintenance round or a whole-of-city contract and never names a site. Coverage rises with sighted-invoice capture, which is the only source that reliably prints a site.",
 "41.4 Batch 39 reconciliations (analysis, no register write). FY2025/26 public place bins: all twelve monthly Parks postings equal the Waste branch driver (LCC DOCS-19541843-v1, row Parks, coded (PK) PK000022-7B115) to the cent, $1,123,128.88. FY2025/26 electricity: 746 register lines on the five uploaded Origin statements group to 234 sites by NMI and every one ties register ex GST to the Origin statement site total to the cent, $27,621.35. Both worked in PSWP_Batch39_Waste_and_Energy_Analysis.xlsx.",
]
for i,t in enumerate(METH): m.cell(mr+i,1,t)._style=ms

d=wb['Data_Acquisition']; dr=d.max_row+1
ds=copy.copy(d.cell(d.max_row,1)._style)
DA=[' |  G1a  |  Batch 39: Origin consolidated statement breakdown CSVs A-FE35E476-021, -028, -029, -030, -031 (Council-wide, 1,267 sub-account sites, 7,449 charge rows). Partially fulfils acquisition item G1. 234 PS/WP sites and 746 register lines reconciled to the cent, $27,621.35. Still outstanding on this account: -022, -023, -025, -026, -032, -033, -034, and the whole A-5AF9E63D series.  |  2  |  Held  | ',
    ' |  G2  |  Batch 39: Waste branch public place bins driver, LCC DOCS-19541843-v1 Public Place Bins Charges 25/26 (bin-lift counts and rates by branch, Corp Property / Parks / RCM / Venues / Water). Ties 12 of 12 months to the register. Request the equivalent 23/24 and 24/25 drivers to complete the three-year series.  |  1  |  Partially held  | ',
    ' |  G3  |  Batch 39: Cleanaway tax invoice 25009848 (8-Jul-2024, contract HEW/02/2021) with the June-2024 Statement of Services workbook, forwarded by Waste branch. Council-wide summary invoice, ten service lines, $1,770,244.85 ex GST; the Parks share is a downstream allocation, so this is NOT a rule 16/17 capture candidate. Held as allocation evidence for register row 199.  |  3  |  Held  | ',
    'v42 update (18-Aug-2026): Batch 39 analysis session. No invoice capture; sighted count unchanged at 1,877 and the control total unchanged at $11,377,312.51. Register gained the site dimension (columns '+SC+' and '+BC+') and the Sites sheet was rebuilt as a live cost-per-park panel over '+str(len(parks))+' parks. Open Items #117-#128 raised from the waste and energy findings.']
for i,t in enumerate(DA): d.cell(dr+i,1,t)._style=ds

oi=wb['Open_Items']; orow=oi.max_row+1
os_=[copy.copy(oi.cell(oi.max_row,c)._style) for c in range(1,8)]
ITEMS=[
 (117,'All','Open','Cleanaway (waste services)',24,999917.80,'NATURAL ACCOUNT INCONSISTENCY. The Parks share of Cleanaway public place bin lifts is coded 73122 Garbage Collect in FY2023/24 and FY2024/25 and 7B115 Internal - Garb from FY2025/26. Trending either account alone is wrong: 73122 falls from $999,918 to $24,702, a 97.5% apparent saving that did not occur. Agree one account with Finance and restate, or record the break in Method and Coverage.'),
 (118,'FY2026/27','Open','Cleanaway (waste services)',0,1120000.00,'TIMING EXPOSURE. At 18-Aug-2026 no FY2026/27 public place bins charge has posted. On the FY2025/26 pattern (July and August charges posted 4-Nov-2025) roughly $1.12M is still to land against PK000022. Any FY2026/27 forecast built on posted actuals understates Parks waste by close to a full year.'),
 (119,'FY2025/26','Open','Cleanaway (waste services)',12,24702.17,'SECOND GARBAGE SERIES UNIDENTIFIED. 73122 continues through FY2025/26 at about $2,058 a month over twelve postings citing real Cleanaway invoice numbers (25130428, 25173104, 25172924, 25255469, 25264357). Scope not established. Note 25130428 is cited both as the FY2024/25 June charge ($97,100.69, row 6096) and the FY2025/26 July charge ($1,808.68, row 17599).'),
 (120,'All','Open','Cleanaway (waste services)',36,0,'SERVICE PERIOD LAGS THE NARRATION BY ABOUT TWO MONTHS. Cleanaway 25009848 is dated 8-Jul-2024 and narrated June, but its Statement of Services covers 28-Apr-2024 to 25-May-2024. FY (Service) on the whole series should be read against the service window, not the narration month. Four narration wordings are used inside FY2025/26 alone (Collection Charges, Internal Garbage Charges, Waste bin and disposal charges, Public Place Bins).'),
 (121,'All','Open','Cleanaway (waste services)',0,0,'RATE STRUCTURE CHANGED. Apr-May 2024 charged every public place bin at a flat $6.127 a lift ($6.691 dog waste). FY2025/26 charges $6.99 for 60L and 140L park bins, $7.61 dog, $7.60 140L and $7.90 240L. The 240L rate rose 28.9%. Confirm the contract variation under HEW/02/2021 and whether the increase was approved.'),
 (122,'FY2026/27','Open','Internal charges (PK000092)',1981,402746.33,'PARK LAND INTERNAL CHARGES RESTRUCTURED. PK000092 appears for the first time in FY2026/27 carrying Fire Levy, Internal Water, Internal Sewerage and Internal Garbage across individual park parcels; internal garbage alone is 328 lines and $177,073.50. Previously PK000022 carried two parcels only. Confirm the restructure was intended and that PK000022 budget was reduced accordingly.'),
 (123,'FY2026/27','Open','Internal charges (PK000022)',1,13987.00,'RATE JUMP. Parcel 9714148, 90-110 Evans Road, charged $1,014.75 a quarter through FY2025/26 and $13,987.00 in 2027-1 (row 20602), a 13.8 times increase. Confirm whether the FY2026/27 posting is annual in advance or a genuine revaluation.'),
 (124,'FY2025/26','Open','Cleanaway (waste services)',2,0,'WATER PARKS WASTE SITS OUTSIDE THE PK STRUCTURE. The driver codes the Water row to (PS)100054-200-7B115, not a PK, at $195.72 to $244.65 a month ($2,628.87 a year). In Jun-2026 $244.65 posted to PK000022 (row 14028) and was correctly reversed the same day (row 14026). Water Parks waste is therefore invisible to the 4090260 ledger.'),
 (125,'FY2023/24','Open','Cleanaway (waste services)',1,161.34,'UNRECONCILED VARIANCE ON THE RECORD. Waste branch email of 10-Jul-2024 (M Commons) states there is $161.34 between Cleanaway and Council data on invoice 25009848 and that a final reconciliation was still to be completed. No later reconciliation is held. Request the closing position.'),
 (126,'FY2025/26','Open','Origin Energy (collective account)',746,27621.35,'TIER 1 UPGRADE PATHWAY NEEDED. All 746 register lines on the five uploaded Origin statements are Status Confirmed at Evidence Tier 2 on Nature Basis "Line narration", yet every one now reconciles to the retailer statement site total to the cent (234 of 234 sites) with NMI, service address, meter number, read type, volume and tariff in hand. The v35 rule bars an identification build from touching a Confirmed row, so pswp_ident_batch.py would skip all 746. Ratify a pathway for upgrading a Confirmed utility line on retailer statement evidence, or leave at Tier 2 and hold the statements as acquisition evidence only.'),
 (127,'All','Open','Site dimension',319,904116.12,'MULTI-SITE LINES HELD. 319 register lines name several parks on one line and are not allocated to any of them (Sites sheet, "'+MULTI+'"). Where a schedule of sites and a per-site rate can be obtained for the underlying contract (PAR/340A, PAR/340E and the standing-order rounds), these become allocable and the site dimension gains $904,116.12 of coverage.'),
 (128,'All','Open','Site dimension',15698,8065429.62,'SITE COVERAGE GAP. 70.9% of register dollars carry no park name in any source field. The only source that reliably prints a site is the sighted invoice, so site coverage tracks capture coverage. Where a vendor prints a site schedule rather than a site, the schedule should be captured to Evidence_Invoice_Lines so the site dimension can read it.'),
]
for i,it in enumerate(ITEMS):
    for j,v in enumerate(it,1):
        c=oi.cell(orow+i,j,v); c._style=os_[j-1]

wb.save(OUT)
print('saved',OUT)
json.dump({'PA':[PA_START,PA_END],'GRP_END':GRP_END,'TOT_ROW':TOT_ROW,'CHK_ROW':CHK_ROW,
           'PB':[PB_TITLE+2,PB_END],'parks':len(parks),'single':n_single,'multi':n_multi,'none':n_none},
          open('b42_pos.json','w'))
