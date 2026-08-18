"""b50_build.py  v49 -> v50: user override - 90-110 Evans Road Meadowbrook IS Riverdale Park. Site_Crosswalk row corrected; class row removed."""
import json, copy, os, sys, subprocess, time, collections
import openpyxl, pandas as pd
from openpyxl.utils import get_column_letter
from python_calamine import CalamineWorkbook
ROOT='/home/claude/b50'; SRC='/mnt/user-data/outputs/PS_WP_Transaction_Register_3FY_v49_HANDOVER.xlsx'
NEW='PS_WP_Transaction_Register_3FY_v50_HANDOVER.xlsx'; OUT=f'{ROOT}/out'; RC=f'{OUT}/recalc'; os.makedirs(RC,exist_ok=True)
CT=11377312.51; SITE=129; BASIS=130; SC='DY'; BC='DZ'
OLD='Non-park Council parcel (user-adjudicated: not a park)'
NEWB='User-adjudicated address decision (local knowledge, Site_Crosswalk)'
cw=CalamineWorkbook.from_path(SRC); cfg={str(r[0]):r[1] for r in cw.get_sheet_by_name('Config').to_python() if r and r[0]}
PA0,PA1=[int(x) for x in str(cfg['SITES_PANELA']).split(':')]; G0,G1=[int(x) for x in str(cfg['SITES_GROUP_ROWS']).split(':')]
PB0,PB1=[int(x) for x in str(cfg['SITES_PANELB']).split(':')]; SA_END=int(str(cfg['SITE_ALLOC_DATA']).split(':')[1]); XW_END=int(str(cfg['SITE_CROSSWALK_DATA']).split(':')[1])
sites=cw.get_sheet_by_name('Sites').to_python(); CLASS=[str(r[0]) for r in sites[G0-1:G1] if r and r[0]]; RESIDUE=next(c for c in CLASS if c.startswith('(residue'))
df=pd.read_excel(SRC,sheet_name='Register',header=3,engine='calamine'); df=df[df.iloc[:,0].notna()]
amt=pd.to_numeric(df.iloc[:,19],errors='coerce').fillna(0); assert abs(amt.sum()-CT)<0.005
MA=next(b for b in df.iloc[:,BASIS-1].fillna('').astype(str).unique() if 'allocated per printed' in b)
bas=df.iloc[:,BASIS-1].fillna('').astype(str).tolist(); site=df.iloc[:,SITE-1].fillna('').astype(str).tolist()
narr=df.iloc[:,21].fillna('').astype(str)
wb=openpyxl.load_workbook(SRC); reg=wb['Register']; tmpl=copy.copy(reg.cell(5,42)._style); n=0
for i in range(len(df)):
    if bas[i]==OLD or (site[i]=='' and bas[i] not in (MA,) and '110 Evans Road' in narr.iloc[i] and bas[i] in set(CLASS)):
        c=reg.cell(i+5,SITE); c.value='Riverdale Park'; c._style=copy.copy(tmpl)
        c=reg.cell(i+5,BASIS); c.value=NEWB; c._style=copy.copy(tmpl); site[i]='Riverdale Park'; bas[i]=NEWB; n+=1
print('rows corrected to Riverdale Park:',n)
xw=wb['Site_Crosswalk']
for r in range(5,XW_END+1):
    if str(xw.cell(r,1).value).startswith('110 Evans Road'):
        xw.cell(r,2).value='Riverdale Park'; xw.cell(r,3).value='Park'
        xw.cell(r,4).value='USER OVERRIDE 18-Aug-2026: 90-110 Evans Road Meadowbrook IS Riverdale Park (local knowledge). The earlier public-property search result (NOT A PARK) is superseded. Local knowledge outranks a public listing; this parcel is Council-rated (S07 9714148) and carries Open Item #123.'
        xw.cell(r,5).value='User (local knowledge)'
CLASS=[c for c in CLASS if c!=OLD]
# Sites rebuild
parks=collections.Counter(); ppk=collections.defaultdict(collections.Counter); pbas=collections.defaultdict(collections.Counter)
pkv=df.iloc[:,16].astype(str).tolist(); CL=set(CLASS)
for s,b,a_,pk in zip(site,bas,amt,pkv):
    if s and ';' not in s and b not in CL and b!=MA and b!='Multiple sites named (not allocated)': parks[s]+=a_; ppk[s][pk]+=1; pbas[s][b]+=1
sa=[x for x in cw.get_sheet_by_name('Site_Allocation').to_python()[4:] if x and x[0] not in (None,'') and isinstance(x[7],(int,float))]
for x in sa:
    if str(x[6])!=RESIDUE: parks[str(x[6])]+=float(x[7]); ppk[str(x[6])][str(x[4])]+=1; pbas[str(x[6])]['Site_Allocation (printed line amounts)']+=1
order=sorted(parks,key=lambda p:-parks[p])
so=wb['Sites']; hdr=copy.copy(so.cell(4,1)._style); dat=copy.copy(so.cell(5,1)._style); num=copy.copy(so.cell(5,2)._style); mon=copy.copy(so.cell(5,3)._style)
elec=[[so.cell(rr,c).value for c in range(1,5)] for rr in range(PB0,PB1+1)]; elec_st=[[copy.copy(so.cell(rr,c)._style) for c in range(1,5)] for rr in range(PB0,PB1+1)]
gnote={so.cell(rr,1).value:so.cell(rr,9).value for rr in range(G0,G1+1)}
for row in so.iter_rows(min_row=3,max_row=so.max_row):
    for c in row: c.value=None
def put(ws,rr,c,v,st=None):
    cell=ws.cell(rr,c); cell.value=v
    if st is not None: cell._style=copy.copy(st)
FYS=['FY2023/24','FY2024/25','FY2025/26','FY2026/27']
put(so,3,1,'Panel A: cost per park / site (live)',hdr)
for j,h in enumerate(['Park / Site','Register lines (live)','$ ex GST (live)']+FYS+['Primary PK','Primary basis','Bases'],1): put(so,4,j,h,hdr)
rr=5; PA_S=5
for p in order:
    put(so,rr,1,p,dat)
    so.cell(rr,2).value=f'=COUNTIF(Register!${SC}$5:${SC}$21473,$A{rr})'; so.cell(rr,2)._style=copy.copy(num); so.cell(rr,2).number_format='#,##0;-#,##0;"-"'
    so.cell(rr,3).value=f'=SUMIF(Register!${SC}$5:${SC}$21473,$A{rr},Register!$T$5:$T$21473)+SUMIF(Site_Allocation!$G$5:$G${SA_END},$A{rr},Site_Allocation!$H$5:$H${SA_END})'
    so.cell(rr,3)._style=copy.copy(mon); so.cell(rr,3).number_format='$#,##0.00;($#,##0.00);"-"'
    for k,fy in enumerate(FYS):
        so.cell(rr,4+k).value=(f'=SUMIFS(Register!$T$5:$T$21473,Register!${SC}$5:${SC}$21473,$A{rr},Register!$E$5:$E$21473,"{fy}")+SUMIFS(Site_Allocation!$H$5:$H${SA_END},Site_Allocation!$G$5:$G${SA_END},$A{rr},Site_Allocation!$D$5:$D${SA_END},"{fy}")')
        so.cell(rr,4+k)._style=copy.copy(mon); so.cell(rr,4+k).number_format='$#,##0.00;($#,##0.00);"-"'
    put(so,rr,8,max(ppk[p].items(),key=lambda x:x[1])[0] if ppk[p] else '',dat); put(so,rr,9,max(pbas[p].items(),key=lambda x:x[1])[0] if pbas[p] else '',dat); put(so,rr,10,len(pbas[p]),dat); rr+=1
PA_E=rr-1; GS=rr
for lbl in [RESIDUE]+[c for c in CLASS if c!=RESIDUE]:
    put(so,rr,1,lbl,dat)
    if lbl==RESIDUE:
        so.cell(rr,2).value=f'=COUNTIF(Register!${BC}$5:${BC}$21473,"{MA}")'; so.cell(rr,3).value=f'=SUMIF(Site_Allocation!$G$5:$G${SA_END},$A{rr},Site_Allocation!$H$5:$H${SA_END})'
        for k,fy in enumerate(FYS): so.cell(rr,4+k).value=f'=SUMIFS(Site_Allocation!$H$5:$H${SA_END},Site_Allocation!$G$5:$G${SA_END},$A{rr},Site_Allocation!$D$5:$D${SA_END},"{fy}")'
    else:
        so.cell(rr,2).value=f'=COUNTIF(Register!${BC}$5:${BC}$21473,$A{rr})'; so.cell(rr,3).value=f'=SUMIF(Register!${BC}$5:${BC}$21473,$A{rr},Register!$T$5:$T$21473)'
        for k,fy in enumerate(FYS): so.cell(rr,4+k).value=f'=SUMIFS(Register!$T$5:$T$21473,Register!${BC}$5:${BC}$21473,$A{rr},Register!$E$5:$E$21473,"{fy}")'
    so.cell(rr,2)._style=copy.copy(num); so.cell(rr,2).number_format='#,##0;-#,##0;"-"'
    for k in range(3,8): so.cell(rr,k)._style=copy.copy(mon); so.cell(rr,k).number_format='$#,##0.00;($#,##0.00);"-"'
    put(so,rr,9,gnote.get(lbl,''),dat); rr+=1
GE=rr-1; TOT=rr
put(so,rr,1,'TOTAL (must tie to the register)',hdr); so.cell(rr,2).value=f'=SUM(B{PA_S}:B{GE})'; so.cell(rr,2)._style=copy.copy(hdr); so.cell(rr,2).number_format='#,##0'
for k in range(3,8):
    L=get_column_letter(k); so.cell(rr,k).value=f'=SUM({L}{PA_S}:{L}{GE})'; so.cell(rr,k)._style=copy.copy(hdr); so.cell(rr,k).number_format='$#,##0.00;($#,##0.00);"-"'
CHK=rr+1
put(so,CHK,1,'Tie check (lines = 21,469 and $ = Register!$T$21475)',hdr)
so.cell(CHK,2).value=f'=IF(B{TOT}=21469,"TRUE","FALSE")'; so.cell(CHK,2)._style=copy.copy(hdr)
so.cell(CHK,3).value=f'=IF(ROUND(C{TOT},2)=ROUND(Register!$T$21475,2),"TRUE","FALSE")'; so.cell(CHK,3)._style=copy.copy(hdr)
so.cell(CHK,4).value=f'=IF(ROUND(SUM(D{TOT}:G{TOT}),2)=ROUND(C{TOT},2),"TRUE","FALSE")'; so.cell(CHK,4)._style=copy.copy(hdr)
NPB=CHK+2; put(so,NPB,1,'Panel B: electricity by retailer account (relocated unchanged)',hdr)
for j,h in enumerate(['Retailer account','Site','Lines','$ ex GST (net, all years)'],1): put(so,NPB+1,j,h,hdr)
rr=NPB+2
for vals,sts in zip(elec,elec_st):
    for j,v in enumerate(vals,1):
        c=so.cell(rr,j); c.value=v; c._style=copy.copy(sts[j-1])
    rr+=1
NPB_E=rr-1
cf=wb['Config']; cr={str(cf.cell(i,1).value):i for i in range(1,cf.max_row+1) if cf.cell(i,1).value}
for k,v in [('WORKBOOK_VERSION','v50'),('SITES_PANELA',f'{PA_S}:{PA_E}'),('SITES_GROUP_ROWS',f'{GS}:{GE}'),('SITES_TOTAL_ROW',TOT),('SITES_CHECK_ROW',CHK),('SITES_PANELB',f'{NPB+2}:{NPB_E}'),('PARK_COUNT',len(order))]:
    if k in cr: cf.cell(cr[k],2).value=v
h=wb['Handover']; hr=h.max_row+1; hs=[copy.copy(h.cell(h.max_row,c)._style) for c in range(1,4)]
h.cell(hr,1).value=f'v50 change (18-Aug-2026): user override. 90-110 Evans Road Meadowbrook IS Riverdale Park (local knowledge, supersedes the public property-search result). {n} lines moved from the not-a-park class to Riverdale Park; Site_Crosswalk row corrected; the one-row class removed. Control total unchanged at ${CT:,.2f}.'
h.cell(hr,1)._style=hs[0]; h.cell(hr,2).value=f'Sites Panel A {len(order)} sites.'; h.cell(hr,2)._style=hs[1]; h.cell(hr,3)._style=hs[2]
m=wb['Method']; mr=m.max_row+1; ms=copy.copy(m.cell(m.max_row,1)._style)
c=m.cell(mr,1); c.value=("48.1 Local knowledge outranks a public listing (v50, standing rule). The web-verification pass marked 90-110 Evans Road Meadowbrook NOT A PARK on a property-portal search; the user's local knowledge is that it IS Riverdale Park. The Council-rated parcel (S07 9714148, Open Item #123) is the park's own parcel. Corrected on Site_Crosswalk with the override recorded and the superseded status retained. Rule: a public property listing describes tenure, not use; it can neither confirm nor deny that a Council parcel is a park."); c._style=copy.copy(ms)
pre=f'{OUT}/{NEW}'; wb.save(pre)
prof=f'{ROOT}/lo_profile/user'; os.makedirs(prof,exist_ok=True)
open(f'{prof}/registrymodifications.xcu','w').write(open('/home/claude/b43/lo_profile/user/registrymodifications.xcu').read())
res_=subprocess.run(['soffice',f'-env:UserInstallation=file://{ROOT}/lo_profile','--headless','--calc','--convert-to','xlsx','--outdir',RC,pre],capture_output=True,text=True,timeout=1800)
final=f'{RC}/{NEW}'; assert os.path.exists(final)
json.dump(dict(final=final,CT=CT,PA=[PA_S,PA_E],GRP=[GS,GE],TOT=TOT,CHK=CHK,PB=[NPB+2,NPB_E],parks=len(order),CLASS=CLASS,MA=MA,gb=int((df.iloc[:,87].notna()).sum()),suburbs=json.load(open('/home/claude/b48/b48_pos.json'))['suburbs']),open(f'{ROOT}/b50_pos.json','w'),indent=1)
r=subprocess.run([sys.executable,'/home/claude/b48/b48_verify.py'],capture_output=True,text=True,env={**os.environ,'B48_POS':f'{ROOT}/b50_pos.json'}); print(r.stdout[-1200:])
if 'FAILS: none' in r.stdout: subprocess.run(['cp',final,f'/mnt/user-data/outputs/{NEW}']); print('SHIPPED',NEW)
else: print('NOT SHIPPED'); sys.exit(1)
